"""Merge an enriched back-fill slice final into a state's canonical
intermediate `{state}_final.xlsx`, deduping by filing_id.

Input  : output/{state}_final_backfill2024.xlsx  (from enrich_backfill.py)
Target : output/{state}_final.xlsx                (existing enriched intermediate
         that run_final_rates.py::load_targets reads)

Only the "Filings" sheet is carried (the sheet run_final_rates reads via
ws.active). Existing rows win on filing_id collision — they hold the original
enrichment; the slice's submission windows are non-overlapping with the
canonical so collisions should be ~0. The pre-merge target is backed up to
`*.prebackfill.bak.xlsx`.

    python merge_backfill_final.py MT
"""
from __future__ import annotations

import shutil
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

OUTPUT_DIR = Path("output")
SLICE_SUFFIX = "_backfill2024"


def _read_filings(path: Path) -> tuple[list, list[list]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    if len(sys.argv) < 2:
        raise SystemExit("usage: python merge_backfill_final.py <STATE>")
    state = sys.argv[1].upper()
    slug = state.lower()
    main_path = OUTPUT_DIR / f"{slug}_final.xlsx"
    slice_path = OUTPUT_DIR / f"{slug}_final{SLICE_SUFFIX}.xlsx"

    if not main_path.exists():
        raise SystemExit(f"missing canonical intermediate: {main_path}")
    if not slice_path.exists():
        raise SystemExit(f"missing slice final: {slice_path}")

    header, main_rows = _read_filings(main_path)
    s_header, slice_rows = _read_filings(slice_path)
    if s_header != header:
        raise SystemExit(f"header mismatch:\n  main : {header}\n  slice: {s_header}")

    idx_fid = header.index("filing_id")
    idx_target = header.index("target_company")
    idx_serff = header.index("serff_tracking_number")

    print(f"[{state}] canonical final : {len(main_rows)} rows")
    print(f"[{state}] slice final     : {len(slice_rows)} rows")

    by_fid: dict[str, list] = {}
    for r in main_rows:
        by_fid[str(r[idx_fid])] = r
    added = 0
    collisions = 0
    for r in slice_rows:
        fid = str(r[idx_fid])
        if fid in by_fid:
            collisions += 1
            continue
        by_fid[fid] = r
        added += 1

    merged = list(by_fid.values())
    merged.sort(key=lambda r: (r[idx_target] or "", r[idx_serff] or "", str(r[idx_fid])))

    print(f"[{state}] new unique      : {added}")
    print(f"[{state}] collisions      : {collisions} (existing kept)")
    print(f"[{state}] merged total    : {len(merged)}")

    bak = main_path.with_suffix(".prebackfill.bak.xlsx")
    if not bak.exists():
        shutil.copy2(main_path, bak)
        print(f"[{state}] backup -> {bak.name}")
    else:
        print(f"[{state}] backup exists ({bak.name}), not overwriting")

    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in merged:
        ws.append(r)
    wb.save(str(main_path))
    print(f"[{state}] wrote -> {main_path}")

    per = Counter(r[idx_target] for r in merged)
    for c in sorted(per, key=lambda x: (x or "")):
        print(f"    {str(c):18s} {per[c]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
