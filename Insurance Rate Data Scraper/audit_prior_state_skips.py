"""Thread 4 — light audit of prior-states enrichment skips.

For ID/WA/CO/OR/UT, collect every "row not found in results" or
"could not open detail" skip across all enrichment runs, classify
each:
  - self-recovered (present in {state}_final_rates.xlsx)
  - out-of-scope (non-target TOI or non-Rate)
  - filing-vehicle excluded (Peerless / Am Economy / LM General etc.)
  - GENUINELY LOST (target-TOI Rate, not in final output, not excluded)

Reads existing logs + workbooks only. No re-scraping.
"""
import re
from collections import defaultdict
from pathlib import Path
import openpyxl

LOGS_DIR = Path("output")

STATE_LOGS = {
    "ID": ["id_full_lookback.log"],
    "WA": ["wa_full_lookback.log"],
    "CO": ["co_full_lookback.log"],
    "OR": ["or_full.log", "or_full_lookback.log"],
    "UT": ["ut_full.log", "ut_full_lookback.log", "ut_full_lookback2.log"],
}

EXCLUDED_SUBSIDIARY_PATTERNS = (
    "lm general insurance company",
    "lm insurance corporation",
    "standard fire insurance",
    "integon ",
    "national general",
    "esurance",
    "drive insurance",
    "united financial",
    "american economy",
    "peerless",
)

SKIP_RE = re.compile(r"\[skip\]\s+(\S+):\s+(row not found in results|could not open detail)")


def collect_skips(state: str) -> dict[str, set[str]]:
    """Return {tracking_number: set(reasons)} across all of this state's logs."""
    skips: dict[str, set[str]] = defaultdict(set)
    for log_name in STATE_LOGS[state]:
        log_path = LOGS_DIR / log_name
        if not log_path.exists():
            continue
        with open(log_path, encoding="utf-8", errors="ignore") as f:
            for line in f:
                m = SKIP_RE.search(line)
                if m:
                    skips[m.group(1)].add(m.group(2))
    return skips


def load_search_metadata(state: str) -> dict[str, dict]:
    """Map tracking_number -> {filing_id, target_company, filing_type, sub_toi, company_name}"""
    path = LOGS_DIR / f"{state.lower()}_all_companies_search.xlsx"
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {n: i for i, n in enumerate(hdr)}
    out = {}
    for r in rows[1:]:
        tk = r[idx["serff_tracking_number"]] or ""
        if not tk:
            continue
        out[tk] = {
            "filing_id": str(r[idx["filing_id"]] or ""),
            "target_company": r[idx["target_company"]] or "",
            "filing_type": r[idx["filing_type"]] or "",
            "sub_type_of_insurance": r[idx["sub_type_of_insurance"]] or "",
            "company_name": r[idx["company_name"]] or "",
        }
    wb.close()
    return out


def load_final_rate_tracking(state: str) -> set[str]:
    """SERFF tracking numbers present in this state's final-rates output."""
    path = LOGS_DIR / f"{state.lower()}_final_rates.xlsx"
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {n: i for i, n in enumerate(hdr)}
    out = {r[idx["serff_tracking_number"]] for r in rows[1:] if r[idx["serff_tracking_number"]]}
    wb.close()
    return out


def load_pdf_effective_dates(state: str, filing_ids: set[str]) -> dict[str, tuple]:
    """For specified filing_ids, peek at cached system PDF for effective dates.
    Returns {filing_id: (eff_new, eff_renewal)} or {} if PDF missing."""
    import pdfplumber
    eff_re_new = re.compile(r"Effective Date \(New\):\s*\n?\s*(\d{2}/\d{2}/\d{4})", re.I)
    eff_re_ren = re.compile(r"Effective Date \(Renewal\):\s*\n?\s*(\d{2}/\d{2}/\d{4})", re.I)
    eff_re_any = re.compile(r"Effective Date\s+(\d{2}/\d{2}/\d{4})")
    out = {}
    for fid in filing_ids:
        pdf_path = LOGS_DIR / "pdfs" / state / fid / "filing_summary.pdf"
        if not pdf_path.exists() or pdf_path.stat().st_size < 5000:
            out[fid] = None
            continue
        try:
            with pdfplumber.open(str(pdf_path)) as pdf:
                text = "\n".join((p.extract_text() or "") for p in pdf.pages[:5])
            new = eff_re_new.search(text)
            ren = eff_re_ren.search(text)
            any_eff = list(eff_re_any.finditer(text))
            out[fid] = (
                new.group(1) if new else (any_eff[0].group(1) if any_eff else None),
                ren.group(1) if ren else (any_eff[1].group(1) if len(any_eff) > 1 else None),
            )
        except Exception:
            out[fid] = None
    return out


def in_window(date_str: str | None) -> bool | None:
    """Is date in [01/01/2025, 04/17/2026]? Returns None if unparseable."""
    if not date_str:
        return None
    from datetime import datetime
    try:
        d = datetime.strptime(date_str, "%m/%d/%Y")
    except ValueError:
        return None
    return datetime(2025, 1, 1) <= d <= datetime(2026, 4, 17)


def classify(state: str, tk: str, meta: dict, in_final: bool, eff_data: tuple | None) -> str:
    sub_toi = meta.get("sub_type_of_insurance", "") or ""
    ftype = meta.get("filing_type", "") or ""
    cname = meta.get("company_name", "") or ""

    # Check target-TOI
    target_toi = sub_toi.startswith("19.") or sub_toi.startswith("04.")
    target_filing = ftype in ("Rate", "Rate/Rule")

    if not (target_toi and target_filing):
        return f"OUT-OF-SCOPE ({ftype} / {sub_toi[:25]})"

    if in_final:
        return "SELF-RECOVERED (in final output)"

    # Filing vehicle?
    n = cname.lower()
    if any(p in n for p in EXCLUDED_SUBSIDIARY_PATTERNS):
        return f"FILING-VEHICLE excluded ({cname})"

    # Effective-date window check
    if eff_data is None:
        return f"PDF NOT CACHED — undetermined (would need recovery to verify)"
    eff_new, eff_ren = eff_data
    if not eff_new and not eff_ren:
        return f"PDF cached but no eff date parsed — undetermined"
    iw_new = in_window(eff_new)
    iw_ren = in_window(eff_ren)
    if iw_new is True or iw_ren is True:
        return f"GENUINELY LOST — in-scope target-Rate, eff_new={eff_new} eff_ren={eff_ren}"
    return f"OUT-OF-WINDOW (eff_new={eff_new} eff_ren={eff_ren})"


def main() -> int:
    overall = {"genuinely_lost": [], "self_recovered": 0, "scope_excluded": 0,
               "filing_vehicle": 0, "out_of_window": 0, "undetermined": 0, "total_skips": 0}

    for state in ["ID", "WA", "CO", "OR", "UT"]:
        skips = collect_skips(state)
        if not skips:
            print(f"\n=== {state} === no skips found")
            continue

        search_meta = load_search_metadata(state)
        final_tracking = load_final_rate_tracking(state)

        # Look up filing_id for each tracking to peek at PDF
        filing_ids_for_undetermined = set()
        for tk in skips:
            meta = search_meta.get(tk, {})
            fid = meta.get("filing_id", "")
            if fid and tk not in final_tracking:
                filing_ids_for_undetermined.add(fid)

        eff_data = load_pdf_effective_dates(state, filing_ids_for_undetermined)

        print(f"\n=== {state}: {len(skips)} unique skipped filings ===")
        breakdown = defaultdict(int)
        lost_in_state = []
        for tk in sorted(skips):
            meta = search_meta.get(tk, {})
            fid = meta.get("filing_id", "")
            in_final = tk in final_tracking
            eff = eff_data.get(fid) if fid else None
            cls = classify(state, tk, meta, in_final, eff)
            cls_key = cls.split(" ")[0]
            breakdown[cls_key] += 1
            if "GENUINELY LOST" in cls:
                lost_in_state.append((tk, meta, cls))

        # Summary
        for k, v in sorted(breakdown.items(), key=lambda x: -x[1]):
            print(f"  {v:3d}  {k}")

        # Detail any genuinely lost
        if lost_in_state:
            print(f"\n  GENUINELY LOST in {state}:")
            for tk, meta, cls in lost_in_state:
                print(f"    {tk:24s} target={meta.get('target_company',''):16s} sub_toi={meta.get('sub_type_of_insurance','')[:35]:35s}")
                print(f"      ->{cls}")
            overall["genuinely_lost"].extend([(state, tk, meta) for tk, meta, _ in lost_in_state])

        overall["total_skips"] += len(skips)
        overall["self_recovered"] += breakdown.get("SELF-RECOVERED", 0)
        overall["scope_excluded"] += breakdown.get("OUT-OF-SCOPE", 0)
        overall["filing_vehicle"] += breakdown.get("FILING-VEHICLE", 0)
        overall["out_of_window"] += breakdown.get("OUT-OF-WINDOW", 0)
        overall["undetermined"] += breakdown.get("PDF", 0)

    print(f"\n\n=== OVERALL ACROSS 5 PRIOR STATES ===")
    print(f"  Total unique skipped filings:           {overall['total_skips']}")
    print(f"  Self-recovered (in final output):       {overall['self_recovered']}")
    print(f"  Out-of-scope (non-target TOI/type):     {overall['scope_excluded']}")
    print(f"  Filing-vehicle excluded:                {overall['filing_vehicle']}")
    print(f"  Out-of-effective-window scope exclude:  {overall['out_of_window']}")
    print(f"  Undetermined (no PDF / no eff date):    {overall['undetermined']}")
    print(f"  GENUINELY LOST:                         {len(overall['genuinely_lost'])}")
    if overall["genuinely_lost"]:
        print()
        for state, tk, meta in overall["genuinely_lost"]:
            print(f"    {state}  {tk}  target={meta.get('target_company','')}  sub_toi={meta.get('sub_type_of_insurance','')[:40]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
