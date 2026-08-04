"""B22 retro-sweep — re-harvest all shipped states with the 2026-08-03 parser.

Applies EVERY guard accumulated since each state's original harvest in one
offline pass (the sweep re-derives deliverables from cached PDFs):
  - B22 (CA-surfaced, the reason for this sweep): explicit
    rate-data-does-NOT-apply alongside a complete parsed company table —
    semantic backstop now trusts the table (ph>0 + impact present, logged
    OVERRIDDEN). This exclusion fired in every state since day one.
  - B21 parser kills (exhibit-arithmetic debris at the source; wrapped-name
    prefix fragments) — previously re-applied only in NJ/NV/LA/NY.
  - All B20-era fixes for any state whose last rebuild predates them.

CA is EXCLUDED: harvested 2026-08-03 with this exact parser + the b23 CDI
overlay already applied; re-sweeping it would only churn the workbook. If CA
is ever added here, b23_apply_ca_cdi.py MUST run after its rebuild.

OFFLINE: --burst 0 schedules zero Begin-Searches (a WARN fires if any spend
is detected). AL/NY prose_extracted rows are re-applied after their rebuilds
(the harvest rewrites the workbook from scratch — B20 lesson).

Detached-safe; resumable (skips states whose delta is already logged).
"""
from __future__ import annotations

import json
import subprocess
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402

STATES = ["ID", "WA", "CO", "OR", "UT", "AZ", "MT", "WY", "NV", "NM", "GA",
          "VA", "OH", "IL", "WV", "NH", "VT", "HI", "AK", "MA", "ND", "RI",
          "ME", "SD", "DE", "CT", "KS", "NJ", "MS", "IA", "SC", "NE", "MD",
          "OK", "AR", "KY", "WI", "PA", "MI", "IN", "MO", "MN", "TN", "AL",
          "LA", "TX", "NY"]  # CA excluded (harvested with this parser today)
DONE_LOG = Path("output/b22_sweep_results.log")


def log(msg: str) -> None:
    line = f"[{time.strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)
    with open(DONE_LOG, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def rowcount(st: str) -> int:
    p = Path(f"output/{st.lower()}_final_rates.xlsx")
    if not p.exists():
        return -1
    wb = openpyxl.load_workbook(p, read_only=True)
    n = wb.active.max_row - 1
    wb.close()
    return n


def reapply_prose(st: str) -> int:
    """Re-append prose_extracted rows lost to the workbook rebuild."""
    if st == "AL":
        rows = json.loads(Path("output/al_prose_extracted.json").read_text(encoding="utf-8"))
        entries = [dict(state="AL", effective_date=e["eff"] or None, company_name=e["company"],
                        line_of_business={"19": "19.0 Personal Auto", "04": "04.0 Homeowners"}.get(e["sub_type"][:2], e["sub_type"][:4]),
                        sub_type_of_insurance=e["sub_type"], overall_rate_impact=f"{e['pct']:.3f}%",
                        rate_activity="rate_change", serff_tracking_number=e["tracking"],
                        source_pdf=f"output/pdfs/AL/{e['tracking'].split('-')[-1]}/filing_summary.pdf",
                        validation_tier="pipeline_only",
                        source="original" if (e["eff"] or "")[-4:] >= "2025" else "extension",
                        external_validation="prose_extracted") for e in rows]
    elif st == "NY":
        entries = [
            dict(state="NY", effective_date="06/15/2024", company_name="Encompass Indemnity Company",
                 line_of_business="19.0 Personal Auto", sub_type_of_insurance="19.0001 Private Passenger Auto (PPA)",
                 overall_rate_impact="16.900%", rate_activity="rate_change",
                 serff_tracking_number="ALSE-134016536", source_pdf="output/pdfs/NY/134016536/filing_summary.pdf",
                 validation_tier="pipeline_only", source="extension", external_validation="prose_extracted"),
            dict(state="NY", effective_date="03/28/2024", company_name="Travelers Personal Insurance Company",
                 line_of_business="04.0 Homeowners", sub_type_of_insurance="04.0000 Homeowners Sub-TOI Combinations",
                 overall_rate_impact="10.000%", rate_activity="rate_change",
                 serff_tracking_number="TRVD-133865900", source_pdf="output/pdfs/NY/133865900/filing_summary.pdf",
                 validation_tier="pipeline_only", source="extension", external_validation="prose_extracted"),
        ]
    else:
        return 0
    p = f"output/{st.lower()}_final_rates.xlsx"
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    existing = {(str(r[hdr.index('serff_tracking_number')].value), str(r[hdr.index('company_name')].value))
                for r in ws.iter_rows(min_row=2)}
    added = 0
    for v in entries:
        if (v["serff_tracking_number"], v["company_name"]) in existing:
            continue
        ws.append([v.get(h) for h in hdr])
        added += 1
    wb.save(p)
    return added


done = set()
if DONE_LOG.exists():
    for line in DONE_LOG.read_text(encoding="utf-8").splitlines():
        if "] DONE " in line:
            done.add(line.split("] DONE ")[1].split(":")[0])

for st in STATES:
    if st in done:
        log(f"SKIP {st} (already swept)")
        continue
    before = rowcount(st)
    if before < 0:
        log(f"DONE {st}: no workbook, skipped")
        continue
    r = subprocess.run([sys.executable, "-u", "run_final_rates.py", st,
                        "--burst", "0", "--wall-stop", "2"],
                       capture_output=True, text=True)
    overrides = [l for l in (r.stdout or "").splitlines() if "OVERRIDDEN" in l]
    spent = next((l for l in (r.stdout or "").splitlines() if "Begin-Searches spent" in l), "")
    if spent and "spent this run: 0/" not in spent:
        log(f"WARN {st}: non-zero SERFF spend detected ({spent.strip()}) — check cache!")
    after = rowcount(st)
    prose = reapply_prose(st)
    final = rowcount(st)
    delta = final - before
    log(f"DONE {st}: {before} -> {final} (delta {delta:+d}; harvest {after}, prose re-applied {prose}); overrides: {len(overrides)}")
    for o in overrides:
        log(f"   {st} {o.strip()}")

log("B22 SWEEP COMPLETE")
