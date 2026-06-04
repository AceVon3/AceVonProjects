"""Enrich ONLY a state's back-fill search slice, reusing the tested
src.detail.enrich_filings orchestrator (same detail-page scraping + PDF
download path as run_{state}_full.py, grouped by (state, target_company)).

Reads  output/{state}_all_companies_search_backfill2024.xlsx  (the new
2023-07-01 -> 2024-06-30 slice produced by run_search.py --out-suffix)
Writes output/{state}_final_backfill2024.xlsx

Merge the result into output/{state}_final.xlsx with merge_backfill_final.py,
then re-emit with run_final_rates.py {state}.

Enriching only the slice (vs re-running run_{state}_full on the merged
workbook) avoids re-enriching the already-cached 2024-07+ filings. PDFs are
downloaded here for Rate/Rule filings; "Rate"-type and any-missing target PDFs
are backfilled by run_final_rates.download_all_pdfs at emit time (unchanged).

    python enrich_backfill.py MT
"""
from __future__ import annotations

import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

sys.stdout.reconfigure(encoding="utf-8")

from src.config import HEADLESS, OUTPUT_DIR, REQUEST_DELAY, USER_AGENT
from src.detail import enrich_filing
from src.models import Filing
from src.output import write_excel
from src.search import _parse_date, _set_rows_per_page_100, _submit_search

SLICE_SUFFIX = "_backfill2024"


def _parse_any_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return _parse_date(s)


def _load_filings(path: Path, default_state: str) -> list[Filing]:
    """Mirror of run_{state}_full.py::_load_filings (state-agnostic)."""
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {name: i for i, name in enumerate(header)}

    def g(r, col):
        i = idx.get(col)
        return r[i] if i is not None else None

    filings: list[Filing] = []
    for r in rows[1:]:
        naic_raw = g(r, "naic_codes") or ""
        naic = [c.strip() for c in str(naic_raw).split(";") if c.strip()]
        filings.append(
            Filing(
                state=g(r, "state") or default_state,
                serff_tracking_number=g(r, "serff_tracking_number") or "",
                filing_id=str(g(r, "filing_id") or ""),
                company_name=g(r, "company_name") or "",
                target_company=g(r, "target_company") or "",
                naic_codes=naic,
                product_name=g(r, "product_name"),
                type_of_insurance=g(r, "type_of_insurance"),
                sub_type_of_insurance=g(r, "sub_type_of_insurance"),
                filing_type=g(r, "filing_type"),
                filing_status=g(r, "filing_status"),
                submission_date=_parse_any_date(g(r, "submission_date")),
                detail_url=g(r, "detail_url"),
            )
        )
    wb.close()
    return filings


def main() -> int:
    if len(sys.argv) < 2:
        raise SystemExit("usage: python enrich_backfill.py <STATE>")
    state = sys.argv[1].upper()
    slug = state.lower()
    src = Path(OUTPUT_DIR) / f"{slug}_all_companies_search{SLICE_SUFFIX}.xlsx"
    out = Path(OUTPUT_DIR) / f"{slug}_final{SLICE_SUFFIX}.xlsx"
    if not src.exists():
        raise SystemExit(f"missing slice workbook: {src}")

    filings = _load_filings(src, state)
    rr = sum(1 for f in filings if f.filing_type == "Rate/Rule")
    print(f"[{state}] loaded {len(filings)} slice filings ({rr} Rate/Rule -> PDF download)", flush=True)

    groups: dict[tuple[str, str], list[Filing]] = defaultdict(list)
    for f in filings:
        groups[(f.state, f.target_company)].append(f)

    t0 = time.time()
    # Resilient per-group enrichment (mirrors run_{state}_full.py): retry
    # _submit_search with context refresh, isolate per-group and per-filing
    # failures, and CHECKPOINT-write after every group so a transient SERFF
    # navigation timeout can't crash the run or lose completed work. (The bare
    # enrich_filings orchestrator had none of this and a single Page.goto
    # timeout aborted the whole CO run.)
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=HEADLESS)
        ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
        page = ctx.new_page()

        def _refresh_context():
            nonlocal ctx, page
            try:
                ctx.close()
            except Exception:
                pass
            ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
            page = ctx.new_page()

        def _submit_with_retry(st: str, company: str, attempts: int = 3) -> bool:
            for attempt in range(1, attempts + 1):
                try:
                    if _submit_search(page, st, company):
                        return True
                    print(f"  [submit {attempt}/{attempts}] returned False", flush=True)
                except Exception as e:
                    print(f"  [submit {attempt}/{attempts}] {type(e).__name__}: {e}", flush=True)
                _refresh_context()
            return False

        order = sorted(groups.keys(), key=lambda k: len(groups[k]))
        for gi, (st, company) in enumerate(order, 1):
            group = groups[(st, company)]
            print(f"\n[group {gi}/{len(order)}] {st} / {company}: {len(group)} filings", flush=True)
            try:
                if not _submit_with_retry(st, company):
                    print(f"  ! search failed after retries, skipping {company}", flush=True)
                    continue
                _set_rows_per_page_100(page)
                for i, f in enumerate(group, 1):
                    try:
                        enrich_filing(page, f, download_pdfs=True)
                    except Exception as e:
                        print(f"    [warn] {f.serff_tracking_number}: {type(e).__name__}: {e}", flush=True)
                        _refresh_context()
                        _submit_with_retry(st, company)
                        _set_rows_per_page_100(page)
                    time.sleep(REQUEST_DELAY)
                    if i % 10 == 0:
                        print(f"    {i}/{len(group)} enriched", flush=True)
            except Exception as e:
                print(f"  ! group {company} crashed: {type(e).__name__}: {e} — checkpointing and continuing", flush=True)
            write_excel(filings, out)  # checkpoint after each group
            print(f"  [checkpoint] wrote {out.name} ({len(filings)} filings)", flush=True)
        browser.close()

    write_excel(filings, out)
    print(f"[{state}] wrote {out} ({len(filings)} filings) in {(time.time()-t0)/60:.1f} min", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
