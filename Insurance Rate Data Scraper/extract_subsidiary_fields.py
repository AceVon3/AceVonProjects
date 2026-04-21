"""Extract (overall_indicated_change, overall_rate_impact, policyholders_affected)
per subsidiary from downloaded SERFF PDFs.

Three extraction strategies, in priority order:
  1. PV Form A layout — per-company section with labeled Section 10 / Section 12.
  2. Table-based extraction (pdfplumber) — headered tables with Company / Rate Impact /
     Policyholders columns.
  3. Free-text regex — company-anchored paragraphs with labeled phrases.

Usage:
  python extract_subsidiary_fields.py --serff SFMA-134676753           # single filing
  python extract_subsidiary_fields.py --all                            # all 23 filings
  python extract_subsidiary_fields.py --state CO --filing-id 134702926 # explicit

Output: JSON dict keyed by serff -> list of subsidiary records + raw audit info.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import OUTPUT_DIR

ALL_STATES = OUTPUT_DIR / "all_states_final.xlsx"
PRIOR_XLSX = OUTPUT_DIR / "rate_changes.xlsx"
PDF_ROOT = OUTPUT_DIR / "pdfs"
OUT_JSON = OUTPUT_DIR / "subsidiary_fields.json"

# PDFs bigger than this get skipped — they're rate tables, not summaries.
MAX_PDF_MB = 3.5

# Filenames matching these substrings are skipped regardless of size —
# they're rate manuals / tier tables / historical exhibits, not summaries.
SKIP_NAME_PATTERNS = [
    "complete manual",
    "rate manual",
    "manual pages",
    "tier exhibit",
    "insurance history",
    "tracked changes",
    "compare",
    "auto cost containment",
    "credit factors",
    "rate capping",
    "rate stability",
    "filing certification",
    "pcrcklst",
    "policy premium breakdown",
    "objection response",
]

# Files matching these get PRIORITIZED (scanned first)
PRIORITY_NAME_PATTERNS = [
    "pv form a",
    "form a",
    "form ra",
    "actuarial memorandum",
    "actuarial memo",
    "filing memo",
    "filing packet",
    "filing exhibits",
    "cover letter",
    "filing summary",
]

# Company-name anchors for per-subsidiary slicing
COMPANY_HINTS = [
    "State Farm Mutual Automobile Insurance Company",
    "State Farm Fire and Casualty Company",
    "State Farm Mutual",
    "State Farm Fire",
    "Allstate Fire and Casualty Insurance Company",
    "Allstate Indemnity Company",
    "Allstate Insurance Company",
    "Allstate Northbrook Indemnity Company",
    "Allstate Property and Casualty Insurance Company",
    "Allstate Fire",
    "Allstate Indemnity",
    "Allstate Insurance",
    "Allstate Northbrook",
    "Allstate Property",
    "GEICO General Insurance Company",
    "GEICO Indemnity Company",
    "GEICO Casualty Company",
    "GEICO General",
    "GEICO Indemnity",
    "GEICO Casualty",
    "Government Employees Insurance Company",
    "Progressive Direct Insurance Company",
    "Progressive Northern Insurance Company",
    "Progressive Specialty Insurance Company",
    "Progressive Preferred Insurance Company",
    "Progressive Classic Insurance Company",
    "Progressive Direct",
    "Progressive Northern",
    "Progressive Specialty",
    "Progressive Preferred",
    "Liberty Mutual Fire Insurance Company",
    "Liberty Insurance Corporation",
    "LM General Insurance Company",
    "Liberty Mutual",
    "Liberty Insurance",
    "LM General",
    "Travelers Home and Marine Insurance Company",
    "Travelers Indemnity Company",
    "Travelers Casualty Insurance Company",
    "Standard Fire Insurance Company",
    "Travelers Home",
    "Travelers Indemnity",
    "Travelers Casualty",
    "Standard Fire",
]


# ---------- text helpers ----------

def normalize(text: str) -> str:
    t = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    t = re.sub(r"([a-zA-Z])(\d)", r"\1 \2", t)
    t = re.sub(r"(\d)([a-zA-Z])", r"\1 \2", t)
    t = re.sub(r"[ \t]+", " ", t)
    return t


def pdf_text(path: Path) -> str:
    try:
        with pdfplumber.open(str(path)) as pdf:
            return "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception:
        return ""


# ---------- pattern library ----------

PCT = r"([+-]?\d+(?:\.\d+)?)\s*%"

# Section 12 = current filing impact; Section 17 = LAST rate change (historical). Must reject 17.
# Pattern 12 variants from SERFF Form A PV:
#   12. OVERALL % RATE IMPACT/ CHANGE (Provide Justification): -7.822%
#   12. OVERALL % RATE IMPACT/ CHANGE: X%
#   OVERALL % RATE IMPACT / CHANGE
FORM_A_IMPACT_RE = re.compile(
    r"12\.\s*OVERALL\s*%\s*RATE\s*IMPACT\s*/\s*CHANGE[^:]*:\s*" + PCT,
    re.IGNORECASE,
)
FORM_A_POLICYHOLDERS_RE = re.compile(
    r"10\.\s*NUMBER\s*OF\s*POLICYHOLDERS\s*AFFECTED[^:]*:\s*([\d,]+)",
    re.IGNORECASE,
)
# Section 1 = COMPANY NAME (capture up to newline)
FORM_A_COMPANY_RE = re.compile(
    r"1\.\s*COMPANY\s*NAME[^:]*:\s*(.{3,120}?)(?:\r?\n|$)",
    re.IGNORECASE,
)

# Free-text patterns — match labeled phrases with optional colon/dash
IMPACT_TEXT_PATTERNS = [
    re.compile(rf"overall\s*%?\s*rate\s*impact[^.\n]{{0,60}}?{PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*(?:proposed\s*)?rate\s*(?:level\s*)?(?:change|increase|decrease)\s*(?:of|is)?\s*{PCT}", re.IGNORECASE),
    re.compile(rf"statewide\s*(?:average|avg)\s*{PCT}\s*(?:change|impact|increase|decrease)", re.IGNORECASE),
    re.compile(rf"proposed\s*(?:overall\s*)?rate\s*(?:level\s*)?change\s*(?:of|is)?\s*{PCT}", re.IGNORECASE),
]

INDICATED_TEXT_PATTERNS = [
    re.compile(rf"overall\s*%?\s*indicated\s*(?:rate\s*)?(?:change|impact|level\s*change)[^.\n]{{0,60}}?{PCT}", re.IGNORECASE),
    re.compile(rf"indicated\s*rate\s*level\s*change[^.\n]{{0,60}}?{PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*rate\s*indication[^.\n]{{0,60}}?{PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*indication\s*(?:of|is)?\s*{PCT}", re.IGNORECASE),
    re.compile(rf"indication[^.\n]{{0,40}}?{PCT}", re.IGNORECASE),  # last resort
]

POLICYHOLDERS_TEXT_PATTERNS = [
    re.compile(r"policyholders?\s*affected[^.\n]{0,40}?([\d,]+)", re.IGNORECASE),
    re.compile(r"policies?\s*affected[^.\n]{0,40}?([\d,]+)", re.IGNORECASE),
    re.compile(r"number\s*of\s*policyholders?[^.\n]{0,40}?([\d,]+)", re.IGNORECASE),
    re.compile(r"([\d,]+)\s*policyholders?\s*(?:affected|impacted|in\s*force)", re.IGNORECASE),
]

# Reject these labeled historical variants
HISTORICAL_REJECT_RES = [
    re.compile(r"17\.\s*OVERALL\s*%\s*RATE\s*IMPACT\s*OF\s*LAST\s*RATE\s*CHANGE", re.IGNORECASE),
    re.compile(r"rate\s*impact\s*of\s*last\s*rate\s*change", re.IGNORECASE),
    re.compile(r"previous\s*rate\s*change", re.IGNORECASE),
]

# Filing-level patterns (no company anchor required — used as fallback for
# single-company filings where the PDF doesn't mention the subsidiary by name
# in sentences like "the proposed overall rate level change is 11.9%").
FILING_LEVEL_IMPACT_PATTERNS = [
    re.compile(rf"(?:proposed\s*)?overall\s*rate\s*level\s*change(?:\s*of)?\s*{PCT}", re.IGNORECASE),
    re.compile(rf"proposed\s*rate\s*level\s*change(?:\s*of)?\s*{PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*statewide\s*(?:rate\s*)?(?:average\s*)?change(?:\s*of)?\s*{PCT}", re.IGNORECASE),
    re.compile(rf"statewide\s*average\s*{PCT}\s*(?:change|impact)", re.IGNORECASE),
    re.compile(rf"overall\s*rate\s*impact(?:\s*of)?\s*{PCT}", re.IGNORECASE),
    re.compile(rf"12\.\s*OVERALL\s*%\s*RATE\s*IMPACT\s*/\s*CHANGE[^:]*:\s*{PCT}", re.IGNORECASE),
]
FILING_LEVEL_INDICATED_PATTERNS = [
    re.compile(rf"indicated\s*rate\s*level\s*change[^.\n]{{0,50}}?{PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*indicated\s*(?:rate\s*)?(?:change|level\s*change)\s*{PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*rate\s*indication(?:\s*of|\s*is)?\s*{PCT}", re.IGNORECASE),
    re.compile(rf"overall\s*indication(?:\s*of|\s*is)?\s*{PCT}", re.IGNORECASE),
]
FILING_LEVEL_POLICYHOLDERS_PATTERNS = [
    re.compile(r"10\.\s*NUMBER\s*OF\s*POLICYHOLDERS\s*AFFECTED[^:]*:\s*([\d,]+)", re.IGNORECASE),
    re.compile(r"policyholders?\s*affected[^.\n]{0,40}?([\d,]+)", re.IGNORECASE),
    re.compile(r"number\s*of\s*policyholders?[^.\n]{0,40}?([\d,]+)", re.IGNORECASE),
]


def is_in_historical_context(text: str, pos: int, window: int = 80) -> bool:
    start = max(0, pos - window)
    end = min(len(text), pos + window)
    ctx = text[start:end]
    return any(r.search(ctx) for r in HISTORICAL_REJECT_RES)


# ---------- Form A section-based extraction ----------

def extract_form_a_blocks(text: str) -> list[dict]:
    """Find 'Section 1 COMPANY NAME' blocks and extract sect 10 / sect 12 values within.

    Returns list of {company, policyholders, impact, source:'form_a'}.
    A valid block needs at minimum an impact value.
    """
    # Find all anchor positions
    company_matches = list(FORM_A_COMPANY_RE.finditer(text))
    impact_matches = list(FORM_A_IMPACT_RE.finditer(text))
    policyholder_matches = list(FORM_A_POLICYHOLDERS_RE.finditer(text))

    # Reject historical section-17 matches (pattern already skips "OF LAST", but defense-in-depth)
    impact_matches = [m for m in impact_matches if not is_in_historical_context(text, m.start())]

    out = []
    # Strategy: iterate through impact matches in doc order; for each, find the nearest
    # preceding company name and nearest preceding/adjacent policyholders count.
    for imp_m in impact_matches:
        imp_pos = imp_m.start()
        imp_val = float(imp_m.group(1))

        # Nearest preceding company name
        company = None
        for cm in company_matches:
            if cm.start() <= imp_pos:
                company = cm.group(1).strip()
            else:
                break

        # Nearest preceding policyholders (same block)
        policyholders = None
        for pm in policyholder_matches:
            if pm.start() <= imp_pos:
                # require within 3000 chars (typical Form A section length)
                if imp_pos - pm.start() < 3000:
                    try:
                        policyholders = int(pm.group(1).replace(",", ""))
                    except ValueError:
                        pass
            else:
                break

        out.append({
            "company": company,
            "policyholders_affected": policyholders,
            "overall_rate_impact": imp_val,
            "overall_indicated_change": None,  # Form A doesn't label indicated
            "source": "form_a",
            "anchor_pos": imp_pos,
        })
    return out


# ---------- free-text per-company slicing ----------

def locate_company_hits(text: str) -> list[tuple[int, str]]:
    """Return sorted [(pos, normalized_name)] for every company mention in text.
    Deduplicates nearby hits of the same company (within 40 chars)."""
    hits: list[tuple[int, str]] = []
    for name in COMPANY_HINTS:
        for m in re.finditer(re.escape(name), text, re.IGNORECASE):
            hits.append((m.start(), name))
    hits.sort()
    # dedupe close adjacent hits
    deduped = []
    for pos, name in hits:
        if deduped and pos - deduped[-1][0] < 40:
            continue
        deduped.append((pos, name))
    return deduped


def canonicalize_company(name: str) -> str:
    """Collapse partial-name hits to full canonical carrier name."""
    low = name.lower()
    mapping = [
        ("state farm mutual", "State Farm Mutual Automobile Insurance Company"),
        ("state farm fire", "State Farm Fire and Casualty Company"),
        ("government employees", "Government Employees Insurance Company"),
        ("geico general", "GEICO General Insurance Company"),
        ("geico indemnity", "GEICO Indemnity Company"),
        ("geico casualty", "GEICO Casualty Company"),
        ("allstate fire", "Allstate Fire and Casualty Insurance Company"),
        ("allstate indemnity", "Allstate Indemnity Company"),
        ("allstate insurance", "Allstate Insurance Company"),
        ("allstate northbrook", "Allstate Northbrook Indemnity Company"),
        ("allstate property", "Allstate Property and Casualty Insurance Company"),
        ("progressive direct", "Progressive Direct Insurance Company"),
        ("progressive northern", "Progressive Northern Insurance Company"),
        ("progressive specialty", "Progressive Specialty Insurance Company"),
        ("progressive preferred", "Progressive Preferred Insurance Company"),
        ("progressive classic", "Progressive Classic Insurance Company"),
        ("liberty mutual", "Liberty Mutual Fire Insurance Company"),
        ("liberty insurance", "Liberty Insurance Corporation"),
        ("lm general", "LM General Insurance Company"),
        ("travelers home", "Travelers Home and Marine Insurance Company"),
        ("travelers indemnity", "Travelers Indemnity Company"),
        ("travelers casualty", "Travelers Casualty Insurance Company"),
        ("standard fire", "Standard Fire Insurance Company"),
    ]
    for needle, canon in mapping:
        if needle in low:
            return canon
    return name


_COMPANY_ALT = r"(?:State\s*Farm\s*(?:Mutual(?:\s*Automobile\s*Insurance\s*Company)?|Fire(?:\s*and\s*Casualty(?:\s*Company)?)?)" \
               r"|Allstate\s*(?:Fire(?:\s*and\s*Casualty\s*Insurance\s*Company)?|Indemnity(?:\s*Company)?|Insurance\s*Company|Northbrook(?:\s*Indemnity\s*Company)?|Vehicle\s*and\s*Property(?:\s*Insurance\s*Company)?|Property(?:\s*and\s*Casualty\s*Insurance\s*Company)?)" \
               r"|Al?lstate\s*New\s*Jersey(?:\s*Property\s*and\s*Casualty)?(?:\s*Insurance\s*Company)?" \
               r"|GEICO\s*(?:General(?:\s*Insurance\s*Company)?|Indemnity(?:\s*Company)?|Casualty(?:\s*Company)?|Secure(?:\s*Insurance\s*Company)?|Advantage(?:\s*Insurance\s*Company)?|Choice(?:\s*Insurance\s*Company)?|Marine\s*Insurance\s*Company)" \
               r"|Government\s*Employees\s*Insurance\s*Company" \
               r"|Progressive\s*(?:Direct(?:\s*Insurance\s*Company)?|Northern(?:\s*Insurance\s*Company)?|Specialty(?:\s*Insurance\s*Company)?|Preferred(?:\s*Insurance\s*Company)?|Classic(?:\s*Insurance\s*Company)?|Casualty(?:\s*Insurance\s*Company)?|Universal(?:\s*Insurance\s*Company)?)" \
               r"|Liberty\s*(?:Mutual(?:\s*(?:Fire|Personal)\s*Insurance\s*Company)?|Insurance\s*Corporation)" \
               r"|LM\s*(?:General|Personal)(?:\s*Insurance\s*Company)?" \
               r"|Safeco\s*Insurance\s*Company(?:\s*of\s*\w+)?" \
               r"|Travelers\s*(?:Home(?:\s*and\s*Marine(?:\s*Insurance\s*Company)?)?|Indemnity(?:\s*Company)?(?:\s*of\s*\w+)?|Casualty(?:\s*Insurance\s*Company)?|Personal(?:\s*Insurance\s*Company)?|Property\s*Casualty(?:\s*Insurance\s*Company)?)" \
               r"|Standard\s*Fire(?:\s*Insurance\s*Company)?" \
               r"|First\s*Floridian(?:\s*Auto\s*and\s*Home)?" \
               r"|TravCo(?:\s*Insurance\s*Company)?)"

# Signed percentage followed by "for|to" + company name, OR company + "... <value>%"
VALUE_TO_COMPANY_RE = re.compile(
    rf"({PCT})\s*(?:change\s*)?(?:for|to|in)\s+({_COMPANY_ALT})",
    re.IGNORECASE,
)
COMPANY_TO_VALUE_RE = re.compile(
    rf"({_COMPANY_ALT})[^%.\n]{{0,60}}?is\s*{PCT}",
    re.IGNORECASE,
)


def _leading_context(text: str, pos: int, window: int = 150) -> str:
    """Return lowercase context preceding position `pos`."""
    start = max(0, pos - window)
    return text[start:pos].lower()


def classify_leading_phrase(ctx: str) -> str | None:
    """Given the text preceding a <value>% <company> hit, classify which field it is."""
    # "indicated rate change", "rate indication", "indication" → indicated
    if any(kw in ctx for kw in (
        "indicated rate level change", "indicated rate change", "rate indication",
        "overall indication", "indicated change", "indicated impact",
    )):
        return "overall_indicated_change"
    # Historical / previous — reject
    if any(kw in ctx for kw in (
        "last rate change", "previous rate change",
    )):
        return None
    # "statewide average change", "rate impact", "rate change", "change of X%", "overall"
    if any(kw in ctx for kw in (
        "statewide average change", "overall statewide average",
        "rate impact", "overall rate impact",
        "overall rate level change", "proposed rate level change",
        "overall rate change", "overall percent change",
        "overall change of", "rate change of", "change of",
        "impact of", "proposed change",
    )):
        return "overall_rate_impact"
    # Default — ambiguous, skip
    return None


def extract_free_text(text: str) -> list[dict]:
    """Scan for <value>% [for|to] <company> patterns; classify by preceding context.

    Also extracts policyholder counts by matching "NUMBER OF POLICYHOLDERS AFFECTED"
    labels near company mentions.
    """
    results: list[dict] = []

    # Pass 1: value-to-company pattern ("X% to SFM and Y% to SFFC")
    for m in VALUE_TO_COMPANY_RE.finditer(text):
        try:
            val = float(m.group(2))  # group 2 = numeric part of PCT
        except (ValueError, IndexError):
            continue
        comp_raw = m.group(3)  # group 3 = company
        if is_in_historical_context(text, m.start()):
            continue
        ctx = _leading_context(text, m.start())
        field = classify_leading_phrase(ctx)
        if not field:
            continue
        results.append({
            "company": canonicalize_company(comp_raw),
            "field": field,
            "value": val,
            "pos": m.start(),
            "context": ctx[-80:],
        })

    # Pass 2: company-to-value pattern ("for SFM is X%")
    for m in COMPANY_TO_VALUE_RE.finditer(text):
        comp_raw = m.group(1)
        try:
            val = float(m.group(2))
        except (ValueError, IndexError):
            continue
        if is_in_historical_context(text, m.start()):
            continue
        ctx = _leading_context(text, m.start())
        field = classify_leading_phrase(ctx)
        if not field:
            continue
        results.append({
            "company": canonicalize_company(comp_raw),
            "field": field,
            "value": val,
            "pos": m.start(),
            "context": ctx[-80:],
        })

    # Group by company, deduplicating (keep first seen per field).
    by_company: dict[str, dict] = {}
    for r in results:
        slot = by_company.setdefault(r["company"], {
            "company": r["company"],
            "overall_rate_impact": None,
            "overall_indicated_change": None,
            "policyholders_affected": None,
            "source": "free_text",
            "impact_candidates": [],
            "indicated_candidates": [],
            "policyholders_candidates": [],
            "impact_contexts": [],
            "indicated_contexts": [],
        })
        if r["field"] == "overall_rate_impact":
            slot["impact_candidates"].append(r["value"])
            slot["impact_contexts"].append(r["context"])
            if slot["overall_rate_impact"] is None:
                slot["overall_rate_impact"] = r["value"]
        elif r["field"] == "overall_indicated_change":
            slot["indicated_candidates"].append(r["value"])
            slot["indicated_contexts"].append(r["context"])
            if slot["overall_indicated_change"] is None:
                slot["overall_indicated_change"] = r["value"]

    return list(by_company.values())


# ---------- table extraction ----------

TABLE_HEADER_KEYWORDS = {
    "company": "company",
    "naic": "naic",
    "rate impact": "overall_rate_impact",
    "overall rate": "overall_rate_impact",
    "overall impact": "overall_rate_impact",
    "indicated": "overall_indicated_change",
    "indication": "overall_indicated_change",
    "policyholders": "policyholders_affected",
    "policies": "policyholders_affected",
    "# policies": "policyholders_affected",
}


def classify_header(cell: str) -> str | None:
    if not cell:
        return None
    low = cell.lower().strip()
    # order matters — check indicated before rate so "indicated rate level" → indicated
    for kw, tag in TABLE_HEADER_KEYWORDS.items():
        if kw in low:
            return tag
    return None


def _clean_table_cell(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _parse_pct(cell: str) -> float | None:
    if not cell:
        return None
    m = re.search(r"([+-]?\d+(?:\.\d+)?)", cell)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _parse_int(cell: str) -> int | None:
    if not cell:
        return None
    cleaned = re.sub(r"[^\d]", "", cell)
    if not cleaned:
        return None
    try:
        n = int(cleaned)
        if 1 <= n <= 100_000_000:
            return n
    except ValueError:
        return None
    return None


def extract_tables(path: Path) -> list[dict]:
    """Return list of subsidiary records from pdfplumber tables.

    Requires an explicit 'company' header column — prevents treating coverage-type
    row labels (BI / PD / etc) as company names.
    """
    out: list[dict] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                tables = page.extract_tables() or []
                for tbl_idx, table in enumerate(tables):
                    if not table or len(table) < 2:
                        continue
                    header_row = None
                    header_idx = -1
                    for hi in range(min(3, len(table))):
                        row = [_clean_table_cell(c) for c in table[hi]]
                        tags = [classify_header(c) for c in row]
                        # require both "company" header AND at least one of our 3 target tags
                        if ("company" in tags and any(t in (
                            "overall_rate_impact", "overall_indicated_change",
                            "policyholders_affected"
                        ) for t in tags)):
                            header_row = row
                            header_idx = hi
                            break
                    if header_row is None:
                        continue
                    tags = [classify_header(c) for c in header_row]
                    for data_row in table[header_idx + 1:]:
                        cells = [_clean_table_cell(c) for c in data_row]
                        if not any(cells):
                            continue
                        rec: dict[str, Any] = {"source": "table",
                                               "page": page_idx + 1,
                                               "table_idx": tbl_idx}
                        for col_idx, (cell, tag) in enumerate(zip(cells, tags)):
                            if tag == "company":
                                rec["company"] = cell
                            elif tag == "overall_rate_impact":
                                rec["overall_rate_impact"] = _parse_pct(cell)
                            elif tag == "overall_indicated_change":
                                rec["overall_indicated_change"] = _parse_pct(cell)
                            elif tag == "policyholders_affected":
                                rec["policyholders_affected"] = _parse_int(cell)
                        # Skip totals/summary rows without a real company name
                        comp = rec.get("company", "").strip()
                        if not comp or comp.lower() in ("total", "totals", "grand total", "weighted", "weighted average"):
                            continue
                        if not any(rec.get(k) is not None for k in
                                   ("overall_rate_impact", "overall_indicated_change",
                                    "policyholders_affected")):
                            continue
                        rec["company"] = canonicalize_company(comp)
                        out.append(rec)
    except Exception as e:
        print(f"    [table error] {path.name}: {e}", file=sys.stderr)
    return out


# ---------- filing driver ----------

def should_skip(path: Path) -> str | None:
    low = path.name.lower()
    for pat in SKIP_NAME_PATTERNS:
        if pat in low:
            return f"name contains {pat!r}"
    try:
        mb = path.stat().st_size / 1024 / 1024
    except OSError:
        return "cannot stat"
    if mb > MAX_PDF_MB:
        return f"{mb:.1f} MB > {MAX_PDF_MB}"
    return None


def prioritize(paths: list[Path]) -> list[Path]:
    def score(p: Path) -> int:
        low = p.name.lower()
        for i, pat in enumerate(PRIORITY_NAME_PATTERNS):
            if pat in low:
                return i
        return len(PRIORITY_NAME_PATTERNS)
    return sorted(paths, key=score)


def extract_filing_level(text: str) -> dict:
    """Run company-agnostic patterns; return first non-historical match per field."""
    out = {
        "overall_rate_impact": None,
        "overall_indicated_change": None,
        "policyholders_affected": None,
        "impact_context": "",
        "indicated_context": "",
        "policyholders_context": "",
    }
    for p in FILING_LEVEL_IMPACT_PATTERNS:
        for m in p.finditer(text):
            if is_in_historical_context(text, m.start()):
                continue
            try:
                out["overall_rate_impact"] = float(m.group(1))
                s = max(0, m.start() - 60); e = min(len(text), m.end() + 60)
                out["impact_context"] = text[s:e]
            except (ValueError, IndexError):
                continue
            break
        if out["overall_rate_impact"] is not None:
            break

    for p in FILING_LEVEL_INDICATED_PATTERNS:
        for m in p.finditer(text):
            if is_in_historical_context(text, m.start()):
                continue
            try:
                out["overall_indicated_change"] = float(m.group(1))
                s = max(0, m.start() - 60); e = min(len(text), m.end() + 60)
                out["indicated_context"] = text[s:e]
            except (ValueError, IndexError):
                continue
            break
        if out["overall_indicated_change"] is not None:
            break

    for p in FILING_LEVEL_POLICYHOLDERS_PATTERNS:
        for m in p.finditer(text):
            raw = m.group(1).replace(",", "")
            try:
                n = int(raw)
            except ValueError:
                continue
            if 1 <= n <= 100_000_000:
                out["policyholders_affected"] = n
                s = max(0, m.start() - 60); e = min(len(text), m.end() + 60)
                out["policyholders_context"] = text[s:e]
                break
        if out["policyholders_affected"] is not None:
            break

    return out


def process_filing(state: str, fid: str, serff: str, target_company: str | None = None) -> dict:
    pdf_dir = PDF_ROOT / state / fid
    if not pdf_dir.exists():
        return {"error": "pdf dir missing", "pdf_dir": str(pdf_dir)}

    all_pdfs = prioritize(sorted(pdf_dir.glob("*.pdf")))
    scanned: list[str] = []
    skipped: list[tuple[str, str]] = []
    form_a_records: list[dict] = []
    table_records: list[dict] = []
    free_text_records: list[dict] = []
    filing_level_records: list[dict] = []
    source_by_pdf: dict[str, dict] = {}

    for pdf_path in all_pdfs:
        reason = should_skip(pdf_path)
        if reason:
            skipped.append((pdf_path.name, reason))
            continue
        raw = pdf_text(pdf_path)
        if not raw:
            skipped.append((pdf_path.name, "no text extracted"))
            continue
        text = normalize(raw)
        fa = extract_form_a_blocks(text)
        for r in fa:
            r["pdf"] = pdf_path.name
        form_a_records.extend(fa)

        tb = extract_tables(pdf_path)
        for r in tb:
            r["pdf"] = pdf_path.name
        table_records.extend(tb)

        ft = extract_free_text(text)
        for r in ft:
            r["pdf"] = pdf_path.name
        free_text_records.extend(ft)

        fl = extract_filing_level(text)
        if any(fl.get(k) is not None for k in
               ("overall_rate_impact", "overall_indicated_change", "policyholders_affected")):
            fl["pdf"] = pdf_path.name
            filing_level_records.append(fl)

        scanned.append(pdf_path.name)
        source_by_pdf[pdf_path.name] = {
            "form_a": len(fa),
            "tables": len(tb),
            "free_text": len(ft),
            "filing_level": 1 if any(fl.get(k) is not None for k in
                                     ("overall_rate_impact", "overall_indicated_change",
                                      "policyholders_affected")) else 0,
        }

    # Reconcile per subsidiary: group all hits by canonical company, prefer form_a > table > free_text
    by_company: dict[str, dict] = {}
    # Form A first — has authoritative labeled values
    for r in form_a_records:
        comp = canonicalize_company(r.get("company") or "(unknown)")
        slot = by_company.setdefault(comp, {
            "company": comp,
            "overall_rate_impact": None,
            "overall_indicated_change": None,
            "policyholders_affected": None,
            "sources": [],
        })
        if slot["overall_rate_impact"] is None and r.get("overall_rate_impact") is not None:
            slot["overall_rate_impact"] = r["overall_rate_impact"]
            slot["sources"].append(f"form_a:{r.get('pdf','')}")
        if slot["policyholders_affected"] is None and r.get("policyholders_affected") is not None:
            slot["policyholders_affected"] = r["policyholders_affected"]
    # Tables next — may have indicated / per-company rows
    for r in table_records:
        comp = canonicalize_company(r.get("company") or "(unknown)")
        slot = by_company.setdefault(comp, {
            "company": comp,
            "overall_rate_impact": None,
            "overall_indicated_change": None,
            "policyholders_affected": None,
            "sources": [],
        })
        if slot["overall_rate_impact"] is None and r.get("overall_rate_impact") is not None:
            slot["overall_rate_impact"] = r["overall_rate_impact"]
            slot["sources"].append(f"table:{r.get('pdf','')}:p{r.get('page')}")
        if slot["overall_indicated_change"] is None and r.get("overall_indicated_change") is not None:
            slot["overall_indicated_change"] = r["overall_indicated_change"]
            slot["sources"].append(f"table_ind:{r.get('pdf','')}:p{r.get('page')}")
        if slot["policyholders_affected"] is None and r.get("policyholders_affected") is not None:
            slot["policyholders_affected"] = r["policyholders_affected"]
    # Free text fills remaining gaps
    for r in free_text_records:
        comp = canonicalize_company(r.get("company") or "(unknown)")
        slot = by_company.setdefault(comp, {
            "company": comp,
            "overall_rate_impact": None,
            "overall_indicated_change": None,
            "policyholders_affected": None,
            "sources": [],
        })
        if slot["overall_rate_impact"] is None and r.get("overall_rate_impact") is not None:
            slot["overall_rate_impact"] = r["overall_rate_impact"]
            slot["sources"].append(f"text:{r.get('pdf','')}")
        if slot["overall_indicated_change"] is None and r.get("overall_indicated_change") is not None:
            slot["overall_indicated_change"] = r["overall_indicated_change"]
            slot["sources"].append(f"text_ind:{r.get('pdf','')}")
        if slot["policyholders_affected"] is None and r.get("policyholders_affected") is not None:
            slot["policyholders_affected"] = r["policyholders_affected"]

    # Filing-level fallback: when no per-subsidiary data was found AND the filing
    # has a single named target company (not "Multiple"), attribute filing-level
    # values to the target. Skip if any per-company data already exists.
    if not by_company and target_company and target_company.lower() != "multiple":
        merged = {
            "company": target_company,
            "overall_rate_impact": None,
            "overall_indicated_change": None,
            "policyholders_affected": None,
            "sources": [],
        }
        for r in filing_level_records:
            if merged["overall_rate_impact"] is None and r.get("overall_rate_impact") is not None:
                merged["overall_rate_impact"] = r["overall_rate_impact"]
                merged["sources"].append(f"filing_level:{r.get('pdf','')}")
            if merged["overall_indicated_change"] is None and r.get("overall_indicated_change") is not None:
                merged["overall_indicated_change"] = r["overall_indicated_change"]
                merged["sources"].append(f"filing_level_ind:{r.get('pdf','')}")
            if merged["policyholders_affected"] is None and r.get("policyholders_affected") is not None:
                merged["policyholders_affected"] = r["policyholders_affected"]
                merged["sources"].append(f"filing_level_pol:{r.get('pdf','')}")
        if any(merged[k] is not None for k in ("overall_rate_impact",
                                                "overall_indicated_change",
                                                "policyholders_affected")):
            by_company[target_company] = merged

    subsidiaries = list(by_company.values())

    return {
        "serff": serff,
        "state": state,
        "filing_id": fid,
        "target_company": target_company,
        "pdf_dir": str(pdf_dir),
        "scanned": scanned,
        "skipped": skipped,
        "subsidiaries": subsidiaries,
        "raw": {
            "form_a": form_a_records,
            "tables": table_records,
            "free_text": free_text_records,
            "filing_level": filing_level_records,
            "source_by_pdf": source_by_pdf,
        },
    }


# ---------- entry points ----------

def load_filing_index() -> dict[str, tuple[str, str]]:
    """Map serff -> (state, filing_id) from all_states_final.xlsx."""
    wb = openpyxl.load_workbook(ALL_STATES, read_only=True, data_only=True)
    ws = wb["Filings"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out = {}
    for r in rows[1:]:
        d = dict(zip(header, r))
        serff = d.get("serff_tracking_number")
        fid = d.get("filing_id")
        state = d.get("state")
        if serff and fid and state:
            out[serff] = (str(state), str(fid))
    return out


def load_target_serffs() -> list[str]:
    wb = openpyxl.load_workbook(PRIOR_XLSX, read_only=True, data_only=True)
    ws = wb["Rate Changes"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    seen = set()
    out = []
    for r in rows[1:]:
        d = dict(zip(header, r))
        s = d.get("serff_tracking_number")
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return out


def load_target_companies() -> dict[str, str]:
    """Map serff -> first-seen company_name from rate_changes.xlsx."""
    wb = openpyxl.load_workbook(PRIOR_XLSX, read_only=True, data_only=True)
    ws = wb["Rate Changes"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out: dict[str, str] = {}
    for r in rows[1:]:
        d = dict(zip(header, r))
        s = d.get("serff_tracking_number")
        c = d.get("company_name")
        if s and c and s not in out:
            out[s] = c
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--serff", help="Single SERFF tracking number")
    ap.add_argument("--state", help="State (with --filing-id)")
    ap.add_argument("--filing-id", help="Filing ID (with --state)")
    ap.add_argument("--all", action="store_true", help="Process all 23 target filings")
    ap.add_argument("--out", default=str(OUT_JSON), help="Output JSON path")
    args = ap.parse_args()

    results: dict[str, dict] = {}
    idx = load_filing_index()
    companies = load_target_companies()

    if args.serff:
        state, fid = idx.get(args.serff, (None, None))
        if not state:
            print(f"! {args.serff} not found in all_states_final")
            return 1
        tc = companies.get(args.serff)
        print(f"[{args.serff}] {state}/{fid} target={tc}")
        results[args.serff] = process_filing(state, fid, args.serff, tc)
    elif args.state and args.filing_id:
        serff = f"manual-{args.state}-{args.filing_id}"
        results[serff] = process_filing(args.state, args.filing_id, serff, None)
    elif args.all:
        target = load_target_serffs()
        print(f"[all] {len(target)} target serffs")
        for i, s in enumerate(target, 1):
            state, fid = idx.get(s, (None, None))
            if not state:
                print(f"  [{i:2d}/{len(target)}] {s:24s}  SKIP (not in index)")
                continue
            tc = companies.get(s)
            print(f"  [{i:2d}/{len(target)}] {s:24s}  state={state} fid={fid} target={tc}")
            results[s] = process_filing(state, fid, s, tc)
    else:
        ap.print_help()
        return 1

    out_path = Path(args.out)
    out_path.write_text(json.dumps(results, indent=2, default=str), encoding="utf-8")
    print(f"\n[write] {out_path}")

    # Summary print
    print("\n" + "=" * 70)
    print("SUBSIDIARY EXTRACTION SUMMARY")
    print("=" * 70)
    for serff, data in results.items():
        subs = data.get("subsidiaries", [])
        print(f"\n{serff} ({data.get('state')}/{data.get('filing_id')})")
        print(f"  Scanned: {len(data.get('scanned', []))}  Skipped: {len(data.get('skipped', []))}")
        if not subs:
            print(f"  No subsidiaries extracted.")
            continue
        for s in subs:
            comp = s.get("company", "?")
            impact = s.get("overall_rate_impact")
            ind = s.get("overall_indicated_change")
            pol = s.get("policyholders_affected")
            print(f"  {comp}")
            print(f"    overall_rate_impact:        {impact}")
            print(f"    overall_indicated_change:   {ind}")
            print(f"    policyholders_affected:     {pol}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
