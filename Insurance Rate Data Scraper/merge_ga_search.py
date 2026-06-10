"""Merge the GA per-carrier search retries into ga_all_companies_search.xlsx.

Run history:
- Initial GA scrape (2026-06-08): the sequential all-companies sweep hit heavy
  SERFF rate-limiting mid-run; Allstate, Travelers, Liberty Mutual, and Encompass
  all failed with "Begin Search" Locator.click timeouts -> 0 rows in the main
  workbook (the OR/NV/NM silent-fail pattern). Verify-retry in isolation (one
  brand at a time) recovered all four — Allstate 153, Travelers 300, Liberty
  Mutual 138, Encompass 55 — confirming the original 0s were false zeros.
"""
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook, Workbook

OUTPUT_DIR = Path("output")
MAIN = OUTPUT_DIR / "ga_all_companies_search.xlsx"
RETRY_FILES = [
    OUTPUT_DIR / "ga_allstate_search.xlsx",
    OUTPUT_DIR / "ga_travelers_search.xlsx",
    OUTPUT_DIR / "ga_liberty_mutual_search.xlsx",
    OUTPUT_DIR / "ga_encompass_search.xlsx",
]
RETRY_CARRIERS = {"Allstate", "Travelers", "Liberty Mutual", "Encompass"}


def _read_rows(path: Path) -> tuple[list, list[list]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    header, rows = _read_rows(MAIN)
    print(f"[main] {MAIN.name}: {len(rows)} rows")
    idx_target = header.index("target_company")
    before = len(rows)
    rows = [r for r in rows if r[idx_target] not in RETRY_CARRIERS]
    print(f"[main] dropped {before - len(rows)} pre-existing rows for retry carriers")
    for path in RETRY_FILES:
        h2, r2 = _read_rows(path)
        if h2 != header:
            raise SystemExit(f"header mismatch in {path}")
        print(f"[merge] {path.name}: {len(r2)} rows")
        rows.extend(r2)
    idx_serff = header.index("serff_tracking_number")
    rows.sort(key=lambda r: (r[idx_target] or "", r[idx_serff] or ""))
    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(MAIN))
    print(f"\n[write] {MAIN}: {len(rows)} total rows")
    per = Counter(r[idx_target] for r in rows)
    for c in sorted(per):
        print(f"  {c:16s} {per[c]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
