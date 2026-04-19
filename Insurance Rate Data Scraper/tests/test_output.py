"""Smoke test for the Excel output pipeline."""
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.models import AttachedPdf, Filing
from src.output import FILINGS_COLUMNS, UNPARSEABLE_COLUMNS, write_excel


def _sample_filings():
    rate_change = Filing(
        state="WA",
        serff_tracking_number="SFMA-134559519",
        filing_id="134559519",
        company_name="State Farm Mutual Automobile Insurance Company",
        target_company="State Farm",
        naic_codes=["176"],
        product_name="Commercial Automobile",
        type_of_insurance="Commercial Auto",
        filing_type="Rate",
        filing_status="Filed",
        submission_date=date(2025, 9, 15),
        disposition_date=date(2025, 10, 1),
        disposition_status="Filed",
        overall_rate_effect=19.1,
        requested_rate_effect=19.1,
        current_avg_premium=1200.00,
        proposed_avg_premium=1429.20,
        pdfs=[AttachedPdf(
            category="Filing Memo",
            display_name="WA Commercial Auto Filing Memo & Exhibits 2025.pdf",
            url="https://filingaccess.serff.com/sfa/.../memo.pdf",
            local_path="output/pdfs/WA/134559519/memo.pdf",
        )],
        pdf_parse_status="parsed",
        pdf_parse_fields_found=["overall_rate_effect", "requested_rate_effect"],
        detail_url="https://filingaccess.serff.com/sfa/filingSummary.xhtml?filingId=134559519",
    )
    rate_change.compute_premium_change()

    neutral = Filing(
        state="WA",
        serff_tracking_number="PROG-134458809",
        filing_id="134458809",
        company_name="Progressive Direct Insurance Company",
        target_company="Progressive",
        filing_type="Rate/Rule",
        submission_date=date(2025, 3, 4),
        overall_rate_effect=0.0,
        pdfs=[AttachedPdf(category="Addendum", display_name="WA 2025 Symbols Addendum Letter.pdf", url="https://example/a.pdf")],
        pdf_parse_status="parsed",
        pdf_parse_fields_found=["overall_rate_effect"],
        detail_url="https://filingaccess.serff.com/sfa/filingSummary.xhtml?filingId=134458809",
    )

    unparseable = Filing(
        state="WA",
        serff_tracking_number="SFMA-134076828",
        filing_id="134076828",
        company_name="State Farm Mutual Automobile Insurance Company",
        target_company="State Farm",
        filing_type="Rate",
        submission_date=date(2024, 5, 10),
        pdfs=[
            AttachedPdf(category="Exhibit", display_name="2024 CW MCY Exhibit A.pdf", url="https://example/e.pdf"),
            AttachedPdf(category="Manual", display_name="Complete CW Motorcycle Manual.pdf", url="https://example/m.pdf"),
        ],
        pdf_parse_status="no_fields_matched",
        pdf_parse_fields_found=[],
        detail_url="https://filingaccess.serff.com/sfa/filingSummary.xhtml?filingId=134076828",
    )

    no_pdfs = Filing(
        state="ID",
        serff_tracking_number="GECO-9999",
        filing_id="9999",
        company_name="GEICO Casualty Company",
        target_company="GEICO",
        filing_type="Rate",
        submission_date=date(2024, 8, 1),
        pdfs=[],
        pdf_parse_status="no_pdfs_attached",
    )

    return [rate_change, neutral, unparseable, no_pdfs]


def test_write_excel_smoke(tmp_path):
    out = tmp_path / "filings.xlsx"
    filings = _sample_filings()
    write_excel(filings, out)

    assert out.exists() and out.stat().st_size > 0

    wb = load_workbook(out)
    assert wb.sheetnames == ["Filings", "Unparseable PDFs"]

    filings_ws = wb["Filings"]
    header = [c.value for c in filings_ws[1]]
    assert header == FILINGS_COLUMNS

    # 1 header + 4 filings
    assert filings_ws.max_row == 5

    # Verify the State Farm 19.1% rate change row
    state_col = FILINGS_COLUMNS.index("overall_rate_effect") + 1
    tid_col = FILINGS_COLUMNS.index("serff_tracking_number") + 1
    prem_change_col = FILINGS_COLUMNS.index("premium_change_dollars") + 1

    row_by_tid = {
        filings_ws.cell(r, tid_col).value: r
        for r in range(2, filings_ws.max_row + 1)
    }
    r = row_by_tid["SFMA-134559519"]
    assert filings_ws.cell(r, state_col).value == 19.1
    assert filings_ws.cell(r, prem_change_col).value == 229.20

    # Unparseable sheet should contain only SFMA-134076828 (has PDFs, no rate fields).
    un_ws = wb["Unparseable PDFs"]
    assert [c.value for c in un_ws[1]] == UNPARSEABLE_COLUMNS
    un_tids = [un_ws.cell(r, 2).value for r in range(2, un_ws.max_row + 1)]
    assert un_tids == ["SFMA-134076828"], f"Expected only the unparseable filing, got: {un_tids}"


def test_unparseable_excludes_neutral_and_no_pdfs(tmp_path):
    """0.0 rate is parsed, not unparseable; filings with no PDFs aren't review candidates."""
    filings = _sample_filings()
    write_excel(filings, tmp_path / "out.xlsx")

    wb = load_workbook(tmp_path / "out.xlsx")
    un_tids = {
        wb["Unparseable PDFs"].cell(r, 2).value
        for r in range(2, wb["Unparseable PDFs"].max_row + 1)
    }
    # Neutral filing (0.0) and no-PDFs filing must NOT appear in manual review queue.
    assert "PROG-134458809" not in un_tids
    assert "GECO-9999" not in un_tids
