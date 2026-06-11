"""Fold the 10 new-carrier GA search workbooks into ga_all_companies_search.xlsx.

13-brand expansion (SCOPE.md). Unlike the retry merges (which replace a
carrier's rows), this is a pure APPEND: new-carrier filing_ids are disjoint
from the existing 9-keyword universe; any overlap keeps the EXISTING row
(dedup by filing_id, first wins) so the 8-brand baseline cannot be disturbed.
verify_search_universe requires this merge before run_final_rates will run.
"""
from pathlib import Path

from openpyxl import Workbook, load_workbook

OUTPUT_DIR = Path("output")
MAIN = OUTPUT_DIR / "ga_all_companies_search.xlsx"
SLUGS = (
    "usaa", "united_services", "garrison",
    "farmers", "mid_century", "fire_insurance_exchange", "truck_insurance_exchange",
    "nationwide", "american_family", "country",
)
# Union ALL generations of each keyword's workbook (base + _retry2 ...) —
# WAF-failed retries write 0-row workbooks and a 2026-06-10 retry clobbered
# the 223-row nationwide base; suffixed retries + union-dedup make every
# generation additive. Rows WITH a submission_date win over undated
# duplicates (the first sweep's date fetches were WAF-challenged).
NEW_FILES = [
    p for slug in SLUGS
    for p in sorted(OUTPUT_DIR.glob(f"ga_{slug}_search*.xlsx"))
]


def _read(path: Path):
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    header, rows = _read(MAIN)
    fi = header.index("filing_id")
    di = header.index("submission_date")
    baseline_ids = {str(r[fi]) for r in rows if r[fi] is not None}
    by_fid: dict[str, list] = {}  # new-carrier rows, dated generation preferred
    print(f"[main] {MAIN.name}: {len(rows)} rows")
    for p in NEW_FILES:
        h2, r2 = _read(p)
        if h2 != header:
            raise SystemExit(f"header mismatch in {p}")
        added = upgraded = dup = 0
        for r in r2:
            fid = str(r[fi]) if r[fi] is not None else ""
            if not fid or fid in baseline_ids:
                dup += 1
                continue
            prev = by_fid.get(fid)
            if prev is None:
                by_fid[fid] = r
                added += 1
            elif prev[di] in (None, "") and r[di] not in (None, ""):
                by_fid[fid] = r  # dated copy beats undated copy
                upgraded += 1
            else:
                dup += 1
        print(f"[merge] {p.name}: +{added} (date-upgraded {upgraded}, dup/in-universe {dup})")
    rows.extend(by_fid.values())
    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(MAIN))
    print(f"\n[write] {MAIN}: {len(rows)} total rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
