"""Build {state}_final_rates.xlsx from cached/downloaded SERFF system PDFs.

Usage:
    python run_final_rates.py ID
    python run_final_rates.py WA
    python run_final_rates.py CO

B1 pipeline (standard since 2026-06-10, validated 0-cell-diff on NM/UT/ID —
see followup_B_enrichment_redundancy.md): reads the state's SEARCH-workbook
universe (union of output/{state}_all_companies_search.xlsx + side workbooks,
deduped by filing_id — see _search_universe_paths), filters to target-TOI
target-carrier filings, downloads each filing's system Filing Summary PDF
(cached), parses it via `utils.parse_filing_summary_pdf`, and emits one row
per per-company rate row in AM Best Disposition Page Data format. The
enrichment phase (run_{state}_full.py) is NO LONGER required: its only unique
contribution to the deliverable was submission_date for search-phase fetch
failures, now supplied by backfill_submission_dates.py (sidecar CSV) or, for
legacy states, the existing enriched workbook.

Non-rate filings (Form, Rule, new-product Rate/Rule) are excluded.
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pdfplumber
from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import (
    AMBEST_VALIDATED_FROM,
    EFFECTIVE_DATE_FROM,
    EFFECTIVE_DATE_TO,
    HEADLESS,
    USER_AGENT,
)

_EFF_WINDOW_FROM = datetime.strptime(EFFECTIVE_DATE_FROM, "%m/%d/%Y").date()
_EFF_WINDOW_TO = datetime.strptime(EFFECTIVE_DATE_TO, "%m/%d/%Y").date()
_AMBEST_VALIDATED_FROM = datetime.strptime(AMBEST_VALIDATED_FROM, "%m/%d/%Y").date()


def _parse_pdf_eff_date(s):
    """Parse the PDF-extracted effective date string (MM/DD/YY or MM/DD/YYYY).
    Returns None on blank or unparseable input."""
    if not s:
        return None
    s = str(s).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _in_effective_window(eff_str) -> bool:
    """Effective-date emit filter. Blank effective_date is KEPT (filer omission).
    Parseable dates outside [EFFECTIVE_DATE_FROM, EFFECTIVE_DATE_TO] are dropped."""
    d = _parse_pdf_eff_date(eff_str)
    if d is None:
        return True
    return _EFF_WINDOW_FROM <= d <= _EFF_WINDOW_TO


def _load_backfill_ids(state: str) -> set[str]:
    """filing_ids in the state's 2026-06-02 back-extension slice
    (output/{state}_all_companies_search_backfill2024.xlsx). Membership in this
    set is the robust provenance signal for "recovered by the back-extension"
    (equivalent to submitted before AMBEST_CROSSCHECK_SUBMISSION_FROM, but
    reliable even when the per-row submission date failed to scrape). Empty set
    if the slice workbook is absent (e.g. a state collected entirely fresh)."""
    src = Path(f"output/{state.lower()}_all_companies_search_backfill2024.xlsx")
    if not src.exists():
        return set()
    import openpyxl as _ox
    wb = _ox.load_workbook(src, read_only=True)
    ws = wb["Filings"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    fi = hdr.index("filing_id")
    ids = {str(r[fi]) for r in ws.iter_rows(min_row=2, values_only=True) if r[fi] is not None}
    wb.close()
    return ids


def _validation_tier(eff_str, is_backfill: bool) -> str:
    """Per-row validation provenance — three tiers (2026-06-02 refinement):

    "ambest_validated"        in the original cross-check dataset (NOT a
                              back-extension recovery) and effective_date in the
                              AM Best window (>= AMBEST_VALIDATED_FROM) or blank
                              (blank-effective rows were kept in the original
                              deliverable). The documented match rates apply.
    "ambest_window_unmatched" recovered by the back-extension (in the
                              2023-07-01 -> 2024-06-30 submission slice) with an
                              in-window effective_date — postdates the
                              cross-checks, never individually verified.
    "pipeline_only"           effective_date below the AM Best window (the 2024
                              back-extension), or a back-extension recovery with
                              a blank effective_date.

    Provenance is decided by SLICE MEMBERSHIP (`is_backfill`), not the per-row
    submission date, because the original finals include rows whose submission
    date failed to scrape (e.g. 5 in OR) and some back-extension TRVD rows have
    blank submission dates. Slice membership is exact either way. Consequence:
    filtering `validation_tier == "ambest_validated"` reproduces exactly the
    original cross-checked set (including its blank-effective rows)."""
    d = _parse_pdf_eff_date(eff_str)
    # eff below the window is decisively the 2024 back-extension.
    if d is not None and d < _AMBEST_VALIDATED_FROM:
        return "pipeline_only"
    # eff is in-window (>= 2025) or blank.
    if is_backfill:
        return "ambest_window_unmatched" if d is not None else "pipeline_only"
    return "ambest_validated"


# The ONLY filing with a documented, all-field AM Best Disposition Page Data
# match (SFMA-134676753, 14/14 — see verify_anchor + dataset_summary
# "Validation"). The per-state cross-checks elsewhere were AGGREGATE COVERAGE
# analyses (directional AM-Best->ours, largely date-relaxed, ID presence-only,
# CO none) and do NOT substantiate per-row validation. So this is the only row
# set that can honestly carry an external "validated" claim. (2026-06-04
# tier-honesty pass.)
_VALIDATED_ANCHORS = {"SFMA-134676753"}


def _relabel(provenance_tier: str, tracking: str) -> tuple[str, str, str, str]:
    """Emit-time BASE labels (provenance != validation). Returns
    (source, external_validation, match_strength, validation_tier).

    The four-tier external_validation is EARNED at ingestion: the emitter sets a
    base, then a cross-check + apply_validation_tiers.py step UPGRADES matched
    rows to "ambest_cross_checked" (+ match_strength) and DOWNGRADES eff>=2025
    rows in a non-cross-checked state (CO, ID non-anchor) to "pipeline_extracted".
    See CROSS_CHECK_STANDARD.md.

      source: "original" (pre-extension 468) | "extension" (2024 back-extension).
      external_validation base:
        "field_validated"  documented all-field per-row AM Best match (anchor).
        "pipeline_extracted_in_validated_window"  eff>=2025 (provisional).
        "pipeline_extracted"  eff<2025 / blank-effective back-extension.
      match_strength: "field" for the anchor, "" otherwise.
      validation_tier: coarse app-compat alias (ambest_validated for
        field_validated/ambest_cross_checked, else pipeline_only)."""
    if tracking in _VALIDATED_ANCHORS:
        source = "original" if provenance_tier == "ambest_validated" else "extension"
        return source, "field_validated", "field", "ambest_validated"
    if provenance_tier == "ambest_validated":           # eff>=2025, original set
        return "original", "pipeline_extracted_in_validated_window", "", "pipeline_only"
    if provenance_tier == "ambest_window_unmatched":    # eff>=2025, back-extension
        return "extension", "pipeline_extracted_in_validated_window", "", "pipeline_only"
    return "extension", "pipeline_extracted", "", "pipeline_only"  # eff<2025 / blank
import src.search as _serff_search
from src.quiet_period import guard as _quiet_guard
from src.detail import download_system_summary_pdf
from src.search import (
    _back_to_results,
    _click_row_to_detail,
    _set_rows_per_page_100,
    _submit_search,
)
from src.utils import parse_filing_summary_pdf

TARGET_TOI = ("19.0", "04.0")  # Personal Auto + Homeowners (Farmowners explicitly out of scope)
GROUP_SEARCH = {  # group -> list of SERFF search terms (each term = a separate SERFF query)
    # MGA Insurance Company is a State Farm subsidiary that files under its
    # own name on SERFF and is NOT returned by a "state farm" keyword search
    # (Item #3a, 2026-05-15).
    "State Farm":     ["state farm", "mga insurance"],
    "GEICO":          ["geico", "government employees"],
    # Encompass files under its own brand on SERFF and is NOT returned by an
    # "allstate" keyword search; we search both names under the Allstate group.
    "Allstate":       ["allstate", "encompass"],
    "Travelers":      ["travelers"],
    # Safeco is Liberty Mutual's independent-agent brand and files under its
    # own name; it does NOT surface under a "liberty mutual" search. Liberty
    # Insurance Corporation (a Liberty Mutual legacy rating company) likewise
    # files under its own name and is NOT returned by "liberty mutual" (no
    # "mutual" in the name) — the Allstate/Encompass search-term-gap family
    # (WV 2026-06-26: surfaced 2 in-target HO filings the group search missed).
    "Liberty Mutual": ["liberty mutual", "safeco", "liberty insurance", "general insurance", "american states", "first national insurance"],
    "Progressive":    ["progressive", "artisan and truckers"],
    # 13-brand expansion (2026-06-10, SCOPE.md). USAA needs three keywords —
    # the parent files as "United Services Automobile Association" and
    # Garrison carries neither string (GA portal check: three disjoint NAIC
    # buckets). Farmers' exchanges/Mid-Century likewise lack "farmers".
    "USAA":             ["usaa", "united services", "garrison"],
    "Farmers":          ["farmers", "mid-century", "fire insurance exchange", "truck insurance exchange"],
    "Nationwide":       ["nationwide"],
    "American Family":  ["american family"],
    "Country Financial": ["country"],
}
GROUP_KW = {  # subsidiary-name keywords used to assign a filing to its parent group
    "State Farm":     ["state farm", "mga insurance"],
    "GEICO":          ["geico", "government employees"],
    "Allstate":       ["allstate", "encompass", "integon", "north american insurance"],
    "Travelers":      ["travelers", "standard fire"],
    "Liberty Mutual": ["liberty mutual", "liberty insurance corporation", "safeco", "first national insurance company of america", "general insurance company of america", "american states"],
    "Progressive":    ["progressive", "artisan and truckers"],
    # 13-brand expansion (2026-06-10, SCOPE.md). Anchors are tighter than the
    # search keywords where bare keywords risk stray matches (Farm-Bureau-
    # style names, Country-Wide); the bare forms remain at the END of each
    # list because the target_company label fallback for "Multiple"-company
    # filings carries the short label ("Farmers", "Country"). Known stray /
    # collision entities are suppressed per-row via
    # EXCLUDED_SUBSIDIARY_PATTERNS (esp. the Munich Re "American Family Home"
    # collision, NAIC 23450).
    "USAA":             ["usaa", "united services", "garrison"],
    "Farmers":          ["farmers insurance exchange", "fire insurance exchange",
                         "truck insurance exchange", "mid-century",
                         "farmers insurance company", "farmers casualty",
                         "farmers property and casualty", "farmers direct",
                         "farmers group property", "farmers"],
    "Nationwide":       ["nationwide"],
    "American Family":  ["american family"],
    "Country Financial": ["country mutual", "country preferred", "country casualty", "country"],
}
# Out-of-scope subsidiaries (do NOT classify as one of our groups):
#   - Esurance (Allstate, wound down 2020)
#   - Drive Insurance (Progressive, retired)
#   - United Financial (Progressive specialty)
# Identifies filings that launch a new product (vs. modifying an existing one).
# A bare "Introduction of" / "Initial Submission" / "Initial Filing" keyword can
# false-positive on body text describing rating-factor additions, deductible
# tweaks, or references to prior filings. This regex anchors those keywords to
# header fields (Project Name/Number, Company Tracking #) and requires body-text
# "introduction of" to be followed by a product-launch noun (Program, line of
# business). Standalone "New Program" / "new product" remain catch-alls because
# audit found no false-positives for those phrasings in our corpus.
NEW_PRODUCT_RE = re.compile(
    r"("
    r"Project Name/Number:[^\n]*\b(?:Initial Filing|Initial Submission|Introduction of)\b"
    r"|Company Tracking #:[^\n]*\bINTRODUCTION OF\b"
    # `\bNew Program\b` excludes the "Rate Transition Modification - New Program
    # Table" SERFF supporting-document boilerplate via negative lookahead
    # (Issue #3, 2026-05-26). Without the lookahead, 5 Travelers HO rate filings
    # in AZ were false-positively excluded as new-product launches.
    # GA (2026-06-10): GA filing summaries embed a DOI questionnaire with the
    # literal field "New Program (Type Yes or No): No" — the bare keyword
    # matched the QUESTION on 225/237 GA rate filings whose answer is "No"
    # (0-row deliverable). The "...: No" lookahead skips the answered-No
    # questionnaire; the explicit "...: Yes" alternative below trusts the
    # answered-Yes declaration outright.
    r"|New Program\s*\(Type Yes or No\)\s*:\s*Yes"
    r"|\bNew Program\b(?!\s+Table)(?!\s*\(Type Yes or No\)\s*:\s*No)"
    # `\bnew product\b` now requires introduction/rollout context within 40
    # chars (Issue #5, 2026-05-27 MT/WY/NV expansion). Without anchoring, the
    # bare \bnew product\b matched narrative/regulator-commentary references
    # like "discussing this new product with the Division" (NV ALSE-134362303,
    # a 47%-rate-change filing) and false-positively excluded 7 legitimate
    # rate filings across CO/AZ/MT/NV.
    r"|\b(?:introduc\w+|launch\w+)\b[\s\S]{0,40}\bnew product\b"
    r"|\bnew product\b[\s\S]{0,40}\b(?:will\s+be\s+(?:offered|available|launched|introduced)|to\s+new\s+business\s+only)\b"
    # The gap must not contain discount/rating-plan/factor terms: GA showed
    # "introduction of an IBHS Fortified Home Discount for its Homeowners line
    # of business" and "introduction of the Easy Pay Factor rating plan ...
    # Program" (2026-06-10) — those introduce a RATING ELEMENT to an existing
    # product (legitimate rate filings), not a product launch.
    r"|\bintroduction of\b(?:(?!\b(?:discount|rating\s+plan|factor)\b)[\s\S]){0,120}\b(?:Program|line of business|lines of business)\b"
    r")",
    re.IGNORECASE,
)
RATE_FILING_TYPES = {"Rate", "Rate/Rule"}


def _is_rate_filing_type(ft) -> bool:
    """State filing-type vocabularies differ: WA..NM use bare "Rate"/"Rate/Rule",
    but GA labels its rate filings "Rate/Rule PPA-Prior Approval", "Rate/Rule
    PPA- File and Use" and "Rate/Rule other than PPA" (found 2026-06-10 when the
    exact-membership check excluded ALL 247 GA rate filings as form_or_rule ->
    0-row deliverable). Exact "Rate" or a "Rate/Rule" prefix counts; nothing
    else ("Form", "Reporting", ...) does."""
    ft = (ft or "").strip()
    return ft == "Rate" or ft.startswith("Rate/Rule")
PDF_FILING_TYPE_RE = re.compile(r"Filing Type:\s*([A-Za-z/ \-]+)\s*$", re.MULTILINE)

# Per-company rate-table rows for these subsidiary names are dropped at emission
# time. They appear in customer-facing-brand filings (kept at filing level via
# parent classification) but are themselves filing vehicles or out-of-scope
# specialty acquisitions, not customer-facing brands. Match is case-insensitive
# substring against the per-company name.
EXCLUDED_SUBSIDIARY_PATTERNS = (
    "lm general insurance company",
    "lm insurance corporation",
    # LM Property and Casualty Insurance Company — the 3rd LM-prefixed Liberty
    # filing vehicle (MA LBPM-133878161, 2026-07-06). Same doctrine as LM
    # General / LM Insurance Corp: no consumer brand, no own agent channel.
    # Decisive: LM General AND Peerless were already excluded at parse ON THE
    # SAME MA FILING -> including LM P&C would split the family arbitrarily.
    "lm property and casualty",
    # Liberty-Wausau family (Decision 2, 2026-07-06): brand retired ~2009 into
    # Liberty Commercial Markets — same profile as Peerless (no consumer
    # website, no own agent channel, AM Best consolidated under LM, legal/
    # filing vehicle). Identical profiles get identical treatment; Peerless is
    # excluded ON THE SAME MA FILING (LBPM-133878161). Patterns are specific
    # so "Mutual of Wausau Insurance Corporation" (an INDEPENDENT WI mutual,
    # no Liberty affiliation) can never match.
    "wausau underwriters",
    "wausau general",
    "wausau business",
    "employers insurance of wausau",
    "standard fire insurance",
    "integon ",                  # Integon Indemnity / Integon National (Allstate-acquired specialty)
    "national general",
    "esurance",
    "drive insurance",
    "united financial",
    # Liberty Mutual-owned filing entity (Item #3b, 2026-05-15). No consumer
    # website, no own agent channel, AM Best rating consolidated under LM,
    # sold under Safeco's umbrella. Matches LM General / Standard Fire pattern.
    "american economy",
    # Peerless Insurance Company / Peerless Indemnity Insurance Company.
    # Liberty Mutual phased out the consumer-facing Peerless brand in 2013;
    # both entities now exist only as legal/filing vehicles within Liberty
    # Mutual Holding Company. No consumer website, no own agent channel, AM
    # Best rating consolidated under LM. Same filing-vehicle pattern as LM
    # General / American Economy (Thread 1, 2026-05-26).
    "peerless",
    # --- 13-brand expansion exclusions (2026-06-10, SCOPE.md) -------------
    # USAA: Noblr is a distinct telematics brand.
    "noblr",
    # Farmers: distinct sub-brands + legacy-MetLife vehicles.
    "foremost",
    "bristol west",
    "coast national",
    "toggle insurance",
    "economy fire",
    "economy premier",
    "economy preferred",
    # Nationwide: Allied brand retired ~2020 (Peerless precedent) + non-standard.
    "amco insurance",
    "allied property and casualty",
    "depositors insurance",
    "titan indemnity",
    "victoria fire",
    # American Family: distinct brands/vehicles, and the NAME COLLISION —
    # "American Family Home Insurance Company" (NAIC 23450) is Munich Re /
    # American Modern, NOT AmFam (verified 2026-06-10). The sibling "American
    # Modern *" entities ride the same filings' rate tables (caught by the
    # Phase 3 guardrail differ as 3 unclassified rows, 2026-06-10).
    "american family home",
    "american modern",
    "american family connect",
    "homesite",
    "midvale",
    "main street america",
    "permanent general",
    "general automobile",
)


def _is_excluded_subsidiary(name: str | None) -> bool:
    n = (name or "").lower()
    return any(p in n for p in EXCLUDED_SUBSIDIARY_PATTERNS)


def carrier_group(*names: Optional[str]) -> Optional[str]:
    """Match the first non-empty name against carrier-group keywords.
    Multi-company filings list company_name as 'Multiple' — pass target_company
    as a fallback so they're not dropped."""
    for name in names:
        n = (name or "").lower()
        if not n:
            continue
        for g, kws in GROUP_KW.items():
            if any(k in n for k in kws):
                return g
    return None


@dataclass
class Target:
    tracking: str
    filing_id: str
    company: str
    toi: str
    sub_toi: str
    filing_type_xlsx: str
    submission_date: object
    disposition_date: object
    disposition_status_xlsx: str
    group: str


def load_targets_enriched(state: str) -> list[Target]:
    """LEGACY loader (pre-B1): targets from the enriched {state}_final.xlsx.
    Kept as the reference path for validate_b1.py's A-vs-B diff; production
    now uses load_targets (search-universe)."""
    src = Path(f"output/{state.lower()}_final.xlsx")
    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        toi = r[ix["type_of_insurance"]] or ""
        sub_toi = r[ix["sub_type_of_insurance"]] or ""
        # Fallback: parent TOI is populated by detail-page enrichment, but
        # when enrichment is skipped (row-not-found during a large search),
        # type_of_insurance can be blank while sub_type_of_insurance (captured
        # at search time) is still reliable. Derive parent TOI from the
        # sub-type prefix so target filings aren't silently dropped.
        if not toi and sub_toi:
            if sub_toi.startswith("19."):
                toi = "19.0 Personal Auto"
            elif sub_toi.startswith("04."):
                toi = "04.0 Homeowners"
        if not any(toi.startswith(p) for p in TARGET_TOI):
            grp = carrier_group(r[ix["company_name"]], r[ix["target_company"]])
            if grp and toi:
                tk = r[ix["serff_tracking_number"]] or ""
                print(f"  excluded — out of scope TOI: {tk} ({grp}, {toi})", flush=True)
            continue
        grp = carrier_group(r[ix["company_name"]], r[ix["target_company"]])
        if not grp:
            continue
        out.append(Target(
            tracking=r[ix["serff_tracking_number"]] or "",
            filing_id=str(r[ix["filing_id"]] or ""),
            company=r[ix["company_name"]] or "",
            toi=toi,
            sub_toi=r[ix["sub_type_of_insurance"]] or "",
            filing_type_xlsx=r[ix["filing_type"]] or "",
            submission_date=r[ix["submission_date"]],
            disposition_date=r[ix["disposition_date"]],
            disposition_status_xlsx=r[ix["disposition_status"]] or "",
            group=grp,
        ))
    return out


# ============================================================
# B1 loaders — search-workbook universe (standard since 2026-06-10)
# ============================================================

# Workbook-name patterns that are archives/inputs already folded into the
# universe, never universe members themselves.
_ARCHIVE_WORKBOOK_RE = re.compile(r"(prebackfill|pre_lookback|backfill2024|\.bak)")


def _search_universe_paths(state: str) -> list[Path]:
    """The search workbooks that together form the state's filing universe.

    States collected after MGA Insurance became a search keyword (Item #3a,
    2026-05-15 — e.g. NM, GA) carry all 9 keywords in the single all_companies
    sweep. Legacy states (UT, ID, ...) keep the MGA scrape in a side workbook
    that was merged into the enriched final but never folded back into
    all_companies — the 2026-06-09 B1 validation caught path B missing
    7 UT + 2 ID GNSC targets because of exactly this."""
    paths = [Path(f"output/{state.lower()}_all_companies_search.xlsx")]
    mga = Path(f"output/{state.lower()}_mga_insurance_search.xlsx")
    if mga.exists():
        paths.append(mga)
    return paths


def _workbook_filing_ids(path: Path) -> set[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Filings"] if "Filings" in wb.sheetnames else wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    fi = hdr.index("filing_id")
    out = {str(r[fi]) for r in ws.iter_rows(min_row=2, values_only=True) if r[fi] is not None}
    wb.close()
    return out


def verify_search_universe(state: str, universe_ids: set[str]) -> None:
    """HARD GATE: every non-archive {state}_*_search*.xlsx side workbook must be
    a subset of the universe. A side workbook with filing_ids the universe
    lacks means a legacy-style unmerged scrape exists and B1 would silently
    drop its rows — refuse to run until it is merged into all_companies (or
    added to _search_universe_paths) or renamed with an archive suffix.
    Audited clean across all 11 states on 2026-06-10
    (tools/audit_search_universe.py)."""
    universe_names = {p.name for p in _search_universe_paths(state)}
    problems = []
    for p in sorted(Path("output").glob(f"{state.lower()}_*_search*.xlsx")):
        if p.name in universe_names or _ARCHIVE_WORKBOOK_RE.search(p.name):
            continue
        extra = _workbook_filing_ids(p) - universe_ids
        if extra:
            problems.append(f"  {p.name}: {len(extra)} filing_id(s) not in universe, e.g. {sorted(extra)[:5]}")
    if problems:
        raise SystemExit(
            f"SEARCH-UNIVERSE VIOLATION for {state} — side workbook(s) contain filings "
            f"missing from the universe (would be silently dropped):\n" + "\n".join(problems)
        )


def _load_sidecar_dates(state: str) -> dict[str, object]:
    """submission_date backfill sidecar written by backfill_submission_dates.py
    ({filing_id: datetime.date}). Rows whose fetch failed (blank date) are
    skipped — the deliverable keeps a blank filing_date for those (precedented:
    OR ships 5 such rows; tiering uses slice membership, not dates)."""
    import csv
    p = Path(f"output/{state.lower()}_submission_date_backfill.csv")
    if not p.exists():
        return {}
    out: dict[str, object] = {}
    with open(p, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            d = (row.get("submission_date") or "").strip()
            if d:
                out[str(row["filing_id"])] = datetime.strptime(d, "%Y-%m-%d").date()
    return out


def _load_enriched_dates(state: str) -> dict[str, object]:
    """submission_date by filing_id from the legacy enriched workbook, if one
    exists (states collected before B1). New B1 states have none."""
    src = Path(f"output/{state.lower()}_final.xlsx")
    if not src.exists():
        return {}
    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    out: dict[str, object] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        fid = str(r[ix["filing_id"]] or "")
        if fid and r[ix["submission_date"]] not in (None, ""):
            out[fid] = r[ix["submission_date"]]
    wb.close()
    return out


def load_targets_search(state: str) -> tuple[list[Target], dict]:
    """B1 target loader: search-workbook universe + layered submission_date
    resolution (search value -> legacy enriched workbook -> mini-pass sidecar
    -> blank). Returns (targets, report)."""
    enriched_dates = _load_enriched_dates(state)
    sidecar_dates = _load_sidecar_dates(state)

    universe_rows: list[tuple] = []
    seen_fids: set[str] = set()
    hdr: list | None = None
    universe_files: list[tuple[str, int]] = []
    for src in _search_universe_paths(state):
        wb = openpyxl.load_workbook(src, read_only=True)
        ws = wb["Filings"] if "Filings" in wb.sheetnames else wb.active
        h = [c.value for c in next(ws.iter_rows(max_row=1))]
        if hdr is None:
            hdr = h
        elif h != hdr:
            raise SystemExit(f"header mismatch in {src}")
        n_new = 0
        fi = h.index("filing_id")
        for r in ws.iter_rows(min_row=2, values_only=True):
            fid = str(r[fi] or "")
            if fid and fid in seen_fids:
                continue
            seen_fids.add(fid)
            universe_rows.append(r)
            n_new += 1
        wb.close()
        universe_files.append((src.name, n_new))
    verify_search_universe(state, seen_fids)

    ix = {h: i for i, h in enumerate(hdr)}
    out: list[Target] = []
    search_blank: list[tuple[str, str, str]] = []  # (tracking, filing_id, group)
    filled_enriched: list[tuple[str, object]] = []
    filled_sidecar: list[tuple[str, object]] = []
    still_blank: list[str] = []
    for r in universe_rows:
        toi = r[ix["type_of_insurance"]] or ""
        sub_toi = r[ix["sub_type_of_insurance"]] or ""
        # Parent TOI is blank in search workbooks; derive from the sub-TOI
        # prefix (same rule the enriched path used as fallback).
        if not toi and sub_toi:
            if sub_toi.startswith("19."):
                toi = "19.0 Personal Auto"
            elif sub_toi.startswith("04."):
                toi = "04.0 Homeowners"
        if not any(toi.startswith(p) for p in TARGET_TOI):
            continue
        grp = carrier_group(r[ix["company_name"]], r[ix["target_company"]])
        if not grp:
            continue
        fid = str(r[ix["filing_id"]] or "")
        tracking = r[ix["serff_tracking_number"]] or ""
        sub_date = r[ix["submission_date"]]
        if sub_date in (None, ""):
            search_blank.append((tracking, fid, grp))
            if fid in enriched_dates:
                sub_date = enriched_dates[fid]
                filled_enriched.append((tracking, sub_date))
            elif fid in sidecar_dates:
                sub_date = sidecar_dates[fid]
                filled_sidecar.append((tracking, sub_date))
            else:
                still_blank.append(tracking)
        out.append(Target(
            tracking=tracking,
            filing_id=fid,
            company=r[ix["company_name"]] or "",
            toi=toi,
            sub_toi=sub_toi,
            filing_type_xlsx=r[ix["filing_type"]] or "",
            submission_date=sub_date,
            disposition_date=r[ix["disposition_date"]],
            disposition_status_xlsx=r[ix["disposition_status"]] or "",
            group=grp,
        ))
    report = {
        "universe_files": universe_files,
        "universe_size": len(universe_rows),
        "search_blank": search_blank,
        "filled_enriched": filled_enriched,
        "filled_sidecar": filled_sidecar,
        "still_blank": still_blank,
    }
    return out, report


def load_targets(state: str) -> list[Target]:
    """Standard (B1) entry point used by main()."""
    targets, rep = load_targets_search(state)
    for name, n in rep["universe_files"]:
        print(f"  search universe: {name} -> +{n} filings", flush=True)
    print(
        f"  submission_date: {len(rep['filled_enriched'])} from legacy enriched workbook, "
        f"{len(rep['filled_sidecar'])} from backfill sidecar, "
        f"{len(rep['still_blank'])} blank",
        flush=True,
    )
    if rep["still_blank"]:
        print(
            f"  blank submission_date targets (run backfill_submission_dates.py {state}): "
            f"{rep['still_blank']}",
            flush=True,
        )
    return targets


# Batched download mode (2026-06-10, Step 2 of the rest/batch/resume plan):
# one fresh "Begin Search" session per batch of DOWNLOAD_BATCH_SIZE uncached
# targets instead of one per target. Fresh-per-target was the deliberate fix
# for JSF results-page degradation (OR Travelers: 9/9 misses over a long
# reused session, all recovered under fresh contexts) — but the SERFF throttle
# rations fresh Begin-Search submissions (observed burst capacity 85 -> 14 ->
# ~12 across consecutive heavy days), so one-search-per-target makes
# final-rates cost ~1 burst-unit per filing and dominates every state's
# throttle bill. Batching cuts that ~batch_size-fold.
#
# Why this is miss-safe (not corruption-risky): download_system_summary_pdf
# clicks rows by stable filing_id (data-rk, pagination-aware) and the PDF is
# server-generated per filing — a degraded session can only FAIL TO FIND rows,
# never fetch the wrong filing's data. Misses fall through to the proven
# fresh-per-target FALLBACK pass below, and convergence re-runs retry whatever
# remains. Validate in vivo with validate_batch_download.py (Step 3) before a
# full batched state run.
DOWNLOAD_BATCH_SIZE = 8  # tune against JSF degradation; 1 = legacy fresh-per-target only
# End a batch session early after this many consecutive misses — the OR
# degradation signature is misses starting partway through a session. The
# batch's unattempted targets go back to the remaining pool (fallback /
# convergence), so aborting early costs nothing but a fresh search.
BATCH_ABORT_CONSECUTIVE_MISSES = 2
# FRONT-OF-BATCH GRACE: the consecutive-miss abort above is the *mid-session*
# degradation signature (a session that WAS landing rows then degraded). Two
# transient misses at the FRONT of a batch — before any row has landed — are NOT
# that signature; aborting on them kills an otherwise-healthy group before its
# bulk/later filings run (OH State Farm burst 6: 2 GNSC front-misses aborted the
# batch before the SFMA bulk; it recovered fine on a fresh run). So the degrade
# abort only arms AFTER a success; until then a larger front grace must pass
# before a "cold abort" (a dead session — wrong term / wall — that never landed a
# row). front_grace > BATCH_ABORT_CONSECUTIVE_MISSES.
BATCH_FRONT_GRACE = 4
# NOT-FOUND STOP: in the per-target fallback, a coverage-gap tail (rows not
# present under any of the group's search terms) returns clean Begin-Search 200s
# with row-not-found downloads — neither a 405 wall (so wall-stop never fires) nor
# a yield collapse the harvest guard reliably catches — so it grinds a fresh
# walling Begin-Search per straggler with nothing to halt it (VA Travelers
# coverage gap, ~40 min wasted). Stop the fallback after this many CONSECUTIVE
# genuine not-founds. None disables. Distinct from wall-stop (405) and from a
# transient row-miss (which still gets its retry across the group's terms).
NOTFOUND_STOP_THRESHOLD = 8

# Re-batched recovery (Q2): batch misses are mostly JSF-degradation misses (the
# row WAS findable; the reused session was contaminated), not genuine
# not-founds. Before paying the per-target fresh-search price (1 walling
# Begin-Search per straggler), re-pool the misses into NEW fresh-context batches
# and retry at batch rate (batch_size:1). Only what survives the recovery rounds
# drops to the per-target fallback. Each recovery round re-pools ALL current
# misses across the group's search terms, so a miss from chunk 1 can ride a
# fresh batch with a miss from chunk 9. A round that recovers nothing means the
# rest are true not-founds / a hard wall -> stop recovering (don't loop).
# VALIDATE-IN-VIVO: on the first rested burst, run ONE batch and confirm the
# recovery pass actually lands degradation-misses at batch rate before trusting
# it on a full run. Set to 0 for legacy behavior (straight to per-target).
REBATCH_RECOVERY_ROUNDS = 1

# Harvest-early (1b): the first few batches of a rested session deliver full
# batch_size:1; yield then decays as the JSF session ages and the WAF penalty
# deepens, collapsing toward the per-target 1:1 regime. Bank the high-leverage
# early batches and STOP before grinding the wall. Download-SCHEDULING only —
# the deliverable is re-derived from whatever PDFs are cached on disk, so
# stopping early never changes a parsed row (convergence re-runs collect the
# rest next burst). A fully-cached state downloads nothing -> guard never fires.
HARVEST_EARLY = True
HARVEST_EARLY_MIN_BATCHES = 3        # warm-up: never stop before this many batches
HARVEST_EARLY_WINDOW = 3             # judge collapse on a rolling window of recent batches
HARVEST_EARLY_COLLAPSE_RATIO = 0.34  # stop if recent landed/attempted < this (~collapsed toward 1:1)


class _HarvestEarlyGuard:
    """Run-level (not per-group) yield watchdog. Fed (landed, attempted) per
    batch; once past the warm-up it trips when recent yield collapses toward the
    per-target 1:1 regime, so the caller can stop before grinding the wall.
    Pure / no I/O -> unit-testable in isolation."""

    def __init__(self, *, enabled: bool = HARVEST_EARLY,
                 min_batches: int = HARVEST_EARLY_MIN_BATCHES,
                 window: int = HARVEST_EARLY_WINDOW,
                 collapse_ratio: float = HARVEST_EARLY_COLLAPSE_RATIO):
        self.enabled = enabled
        self.min_batches = min_batches
        self.window = window
        self.collapse_ratio = collapse_ratio
        self._batches: list[tuple[int, int]] = []  # (landed, attempted) per batch
        self._stopped = False

    def record(self, landed: int, attempted: int) -> None:
        if attempted > 0:
            self._batches.append((landed, attempted))

    def should_stop(self) -> bool:
        if not self.enabled or self._stopped:
            return self._stopped
        if len(self._batches) < self.min_batches:
            return False
        window = self._batches[-self.window:]
        landed = sum(l for l, _a in window)
        attempted = sum(a for _l, a in window)
        if attempted and landed / attempted < self.collapse_ratio:
            self._stopped = True
        return self._stopped


class _BeginSearchBudget:
    """Optional hard cap on Begin-Searches per run — a 'burst'. None = unlimited
    (default; preserves prior behavior). Counts every Begin-Search ATTEMPT (the
    WAF-rationed unit — incl. retries and re-batched-recovery searches). When the
    cap is reached the run stops SCHEDULING new searches and defers the rest, so a
    burst stops UNDER the wall (~17) instead of grinding into the 405. This is the
    proactive complement to _HarvestEarlyGuard, which only trips AFTER yield
    collapses (i.e. post-wall) — the cap lets us bank near-not-at the ceiling and
    leave headroom, which is gentler on the slow (days-scale) penalty component.

    Miss-safe, same guarantee as harvest-early: a capped run is a PARTIAL run, not
    a corrupt one — the deliverable is re-derived from whatever PDFs are cached on
    disk, so a deferred target is missing, never wrong, and the next burst (or a
    convergence re-run) collects it. Pairs with _HarvestEarlyGuard via an OR'd
    stop predicate (see download_all_pdfs._stop)."""

    def __init__(self, limit: int | None = None):
        self.limit = limit
        self.count = 0

    def spend(self) -> None:
        self.count += 1

    def exhausted(self) -> bool:
        return self.limit is not None and self.count >= self.limit


class _SustainedWallStop:
    """Stop the run after `threshold` CONSECUTIVE walled Begin-Searches — the REAL
    WAF wall signal (a Begin-Search that fails all its retries: the
    begin_search_link_timeout / HTTP 405 signature). threshold=None disables it.

    This is the correct stop signal for a no-cap harvest, and fixes the
    harvest-early ORDERING BUG: harvest-early keys on per-batch DOWNLOAD yield, so a
    wrong-search-term miss (the row is under a later term, e.g. Encompass filings
    under `encompass` not `allstate`) reads as a yield collapse and bails BEFORE the
    later terms + Q2 recovery run. A wrong-term search still returns 200 — only its
    row-downloads miss — so keying the stop on Begin-Search SUCCESS/FAILURE never
    mis-fires on wrong-term misses, lets every term + recovery have its chance, and
    still halts the 405-grind promptly (no fixed cap needed)."""

    def __init__(self, threshold: int | None = 2):
        self.threshold = threshold
        self.consecutive = 0
        self.tripped = False

    def record(self, search_ok: bool) -> None:
        if self.threshold is None or self.tripped:
            return
        if search_ok:
            self.consecutive = 0
        else:
            self.consecutive += 1
            if self.consecutive >= self.threshold:
                self.tripped = True

    def should_stop(self) -> bool:
        return self.tripped


class _NotFoundStop:
    """Stop the per-target fallback after `threshold` CONSECUTIVE genuine
    not-founds — a clean Begin-Search (HTTP 200) whose row download still misses,
    i.e. the row is not present under the search term. That is the coverage-gap
    signature (e.g. VA's TRVD-G series absent under bare `travelers`): it is NOT a
    405 wall (wall-stop owns that) and NOT a transient row-miss (which still gets
    its retry across the group's other terms). Without this stop the fallback
    fires one fresh walling Begin-Search per straggler with nothing to halt it
    (VA Travelers, ~40 min). threshold=None disables (default-off at the class
    level; the caller passes the configured threshold).

    record(search_ok, found): a wall (search_ok=False) is ignored here (wall-stop
    owns it); a found row resets the streak; a clean-but-not-found increments it.
    Pure / no I/O -> unit-testable in isolation."""

    def __init__(self, threshold: int | None = None):
        self.threshold = threshold
        self.consecutive = 0
        self.tripped = False

    def record(self, search_ok: bool, found: bool) -> None:
        if self.threshold is None or self.tripped:
            return
        if not search_ok:
            return  # a walled Begin-Search is not a not-found — wall-stop owns it
        if found:
            self.consecutive = 0
        else:
            self.consecutive += 1
            if self.consecutive >= self.threshold:
                self.tripped = True

    def should_stop(self) -> bool:
        return self.tripped


def _should_abort_batch(consecutive_misses: int, seen_success: bool, *,
                        degradation_threshold: int = BATCH_ABORT_CONSECUTIVE_MISSES,
                        front_grace: int = BATCH_FRONT_GRACE) -> bool:
    """Decide whether to end a reused batch session early (pure -> testable).

    Two distinct signatures:
      - DEGRADATION abort: >= degradation_threshold consecutive misses AFTER at
        least one row has landed this session (`seen_success`). This is the
        documented OR-degradation signature — a working session that decayed —
        and is the ORIGINAL behavior, unchanged.
      - COLD abort: >= front_grace consecutive misses with NO success yet — a dead
        session (wrong term / wall) that never produced a row.

    front_grace > degradation_threshold, so a couple of transient misses at the
    FRONT of a batch (before any success) are tolerated and the bulk + later
    filings still run (the OH State Farm burst-6 fix). When a success has been
    seen, behavior is identical to the prior `consecutive >= degradation_threshold`
    rule — defaults preserved for the normal mid-session-degradation case."""
    if seen_success:
        return consecutive_misses >= degradation_threshold
    return consecutive_misses >= front_grace


def _batched_with_recovery(search_terms, remaining, *, batch_size, run_batch_fn,
                           recovery_rounds: int = REBATCH_RECOVERY_ROUNDS,
                           on_batch=None, should_stop=None, post_pass_stop=None,
                           log=print):
    """Primary batched pass + up to `recovery_rounds` re-batched passes over the
    misses, returning (still_remaining, stopped_early). `run_batch_fn(term, chunk)
    -> list[miss]` runs ONE fresh-context batched session (real impl drives
    Playwright; tests inject a mock). `on_batch(landed, attempted)` feeds the
    harvest guard.

    TWO stop hooks, deliberately ordered (the harvest-early ORDERING fix):
      - `should_stop()` — checked MID-CHUNK, for the IMMEDIATE signals only
        (sustained 405 wall, burst-cap). These are WAF/budget facts that should
        halt scheduling at once.
      - `post_pass_stop()` — checked ONLY AFTER all search terms + every recovery
        round have run. This is where the harvest-early yield guard belongs: a
        wrong-term round-0 search returns 200 with zero landed rows (e.g. Encompass
        filings miss under `allstate`, land under `encompass`), which looks like a
        yield collapse mid-round. Judging yield only after the LATER terms +
        recovery have had their chance prevents that false bail (VA Allstate). If
        it trips, bank progress and skip the per-target fallback.

    Pure orchestration otherwise — the I/O is entirely behind the injected
    callables, so this is unit-testable offline. Stable-filing_id keying and the
    per-batch abort/grace live inside run_batch_fn (real _run_batch) and are
    untouched — this only schedules which targets get re-batched vs dropped to
    per-target. Backward-compatible: post_pass_stop=None preserves prior behavior."""
    remaining = list(remaining)
    rounds = 1 + max(0, recovery_rounds)
    for round_idx in range(rounds):
        if not remaining:
            break
        before = len(remaining)
        label = "BATCH" if round_idx == 0 else f"RECOVERY-{round_idx}"
        for term in search_terms:
            if not remaining:
                break
            log(f"[{label}] search={term!r}, {len(remaining)} filing(s) in chunks of {batch_size}")
            still: list = []
            for i in range(0, len(remaining), batch_size):
                chunk = remaining[i:i + batch_size]
                misses = run_batch_fn(term, chunk)
                if on_batch is not None:
                    on_batch(len(chunk) - len(misses), len(chunk))
                still.extend(misses)
                if should_stop is not None and should_stop():
                    still.extend(remaining[i + batch_size:])  # unattempted in this term
                    return still, True
            remaining = still
        after = len(remaining)
        if round_idx > 0 and after >= before:
            log(f"[RECOVERY-{round_idx}] no progress ({after} remain) -> stop recovery")
            break
    # All terms + recovery have run — NOW judge harvest-early yield (ordering fix).
    if remaining and post_pass_stop is not None and post_pass_stop():
        return remaining, True
    return remaining, False


def download_all_pdfs(state: str, targets: list[Target], *,
                      batch_size: int = DOWNLOAD_BATCH_SIZE,
                      begin_search_budget: int | None = None,
                      harvest_early: bool = True,
                      wall_stop: int | None = 2,
                      notfound_stop: int | None = NOTFOUND_STOP_THRESHOLD) -> dict[str, str]:
    """Download system PDF for every target. Returns {filing_id: download_status}.
    Primary path: batched (batch_size downloads per fresh search session).
    Stragglers then get the legacy fresh-search-per-target fallback."""
    # Failure-signature capture for every fresh search this run makes
    # (ledger row per search + snapshot on failure — see src/search.DIAG_DIR).
    _serff_search.DIAG_DIR = Path("output/serff_diagnostics")
    by_group: dict[str, list[Target]] = {}
    for t in targets:
        by_group.setdefault(t.group, []).append(t)
    statuses: dict[str, str] = {}
    harvest = _HarvestEarlyGuard(enabled=harvest_early)  # run-level yield watchdog (1b); inert when fully cached
    # ORDERING FIX (2026-06-25): harvest-early is now evaluated only AFTER all
    # search terms + the recovery round run (passed as post_pass_stop, NOT in the
    # mid-chunk _stop_hard), so a wrong-term round-0 search (Encompass under
    # `allstate`) can no longer read as a yield collapse and bail before the later
    # terms + recovery land their rows. The mid-chunk stop is the WAF/budget facts
    # only (sustained wall + burst cap). harvest_early ON is now safe to default.
    budget = _BeginSearchBudget(begin_search_budget)  # optional burst cap (Begin-Searches); None = unlimited
    wall = _SustainedWallStop(wall_stop)  # stop on the REAL wall (consecutive Begin-Search 405s)
    notfound = _NotFoundStop(notfound_stop)  # stop the per-target fallback on a coverage-gap not-found tail

    def _stop_hard() -> bool:
        """IMMEDIATE stop (checked mid-chunk): the sustained Begin-Search wall (real
        WAF wall) or the burst cap. WAF/budget facts that must halt scheduling at
        once — NOT harvest-early (that is judged only after all terms + recovery)."""
        return wall.should_stop() or budget.exhausted()

    def _stop() -> bool:
        """Full stop predicate (post-pass / fallback): adds the harvest-yield
        collapse and the coverage-gap not-found tail to the hard signals. Either
        way the orchestration defers the rest cleanly (miss-safe)."""
        return _stop_hard() or harvest.should_stop() or notfound.should_stop()

    def _stop_reason() -> str:
        if wall.should_stop():
            return "sustained wall"
        if budget.exhausted():
            return "burst cap"
        if notfound.should_stop():
            return "not-found tail"
        return "harvest-early"

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
        page = ctx.new_page()

        def _refresh_context():
            nonlocal ctx, page
            try:
                ctx.close()
            except Exception:
                pass
            ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
            page = ctx.new_page()

        def _submit_with_retry(st: str, term: str) -> bool:
            attempted = False
            for attempt in range(3):
                if budget.exhausted():  # never fire a Begin-Search past the burst cap
                    break
                budget.spend()          # count every attempt (the WAF-rationed unit)
                attempted = True
                try:
                    if _submit_search(page, st, term):
                        wall.record(True)   # clean search resets the consecutive-wall counter
                        return True
                except Exception as e:
                    print(f"    [retry {attempt+1}/3] submit_search {term!r}: {e}", flush=True)
                _refresh_context()
            if attempted:               # all retries failed -> one walled Begin-Search
                wall.record(False)      # (budget-pre-empted searches are not counted as walls)
            return False

        def _run_batch(term: str, batch: list[Target]) -> list[Target]:
            """One fresh search session, then sequential downloads for the
            whole batch. Returns the targets still missing (misses + any
            unattempted after an early abort)."""
            _refresh_context()
            if not _submit_with_retry(state, term):
                return list(batch)
            _set_rows_per_page_100(page)
            misses: list[Target] = []
            consecutive = 0
            seen_success = False
            for j, t in enumerate(batch, 1):
                dest_dir = Path(f"output/pdfs/{state}/{t.filing_id}")
                try:
                    pdf = download_system_summary_pdf(page, t.filing_id, t.tracking, dest_dir)
                except Exception as e:
                    print(f"    [batch {j}/{len(batch)}] {t.tracking}: error {type(e).__name__} -> miss", flush=True)
                    pdf = None
                if pdf:
                    statuses[t.filing_id] = "ok"
                    consecutive = 0
                    seen_success = True
                    print(f"    [batch {j}/{len(batch)}] {t.tracking}: ok", flush=True)
                else:
                    misses.append(t)
                    consecutive += 1
                    print(f"    [batch {j}/{len(batch)}] {t.tracking}: miss", flush=True)
                    # Front-of-batch grace: a couple of transient FRONT misses (no
                    # success yet) don't abort — only mid-session degradation
                    # (after a success) or a dead session (front_grace misses) does.
                    if _should_abort_batch(consecutive, seen_success) and j < len(batch):
                        rest = batch[j:]
                        print(f"    [batch] {consecutive} consecutive misses — ending session early "
                              f"({len(rest)} unattempted -> retried later)", flush=True)
                        misses.extend(rest)
                        break
            return misses

        for grp, items in by_group.items():
            # The not-found tail is a PER-GROUP coverage-gap signal (a gap in
            # Travelers says nothing about Liberty), so reset it each group — unlike
            # the wall/budget/harvest signals, which are run-level. _stop() reads
            # this rebind at call time.
            notfound = _NotFoundStop(notfound_stop)
            uncached = []
            for t in items:
                pdf = Path(f"output/pdfs/{state}/{t.filing_id}/filing_summary.pdf")
                if pdf.exists() and pdf.stat().st_size > 5000:
                    statuses[t.filing_id] = "cached"
                else:
                    uncached.append(t)
            if not uncached:
                print(f"[{grp}] all {len(items)} cached", flush=True); continue
            if _stop():
                print(f"[{grp}] {_stop_reason()} reached — deferring {len(uncached)} "
                      f"uncached to next burst (no grind)", flush=True)
                continue
            search_terms = GROUP_SEARCH[grp]
            print(f"[{grp}] searches={search_terms!r}, downloading {len(uncached)}/{len(items)}", flush=True)
            remaining = list(uncached)

            # --- primary batched pass + re-batched recovery (Q2) ---
            # Recovery re-pools batch misses into fresh batches (batch_size:1)
            # before the per-target fallback, because most misses are
            # JSF-degradation (row findable, session contaminated) not true
            # not-founds. Stable-filing_id keying + the per-batch degradation
            # guard live in _run_batch and are untouched.
            stopped_early = False
            if batch_size > 1:
                remaining, stopped_early = _batched_with_recovery(
                    search_terms, remaining, batch_size=batch_size,
                    run_batch_fn=_run_batch, on_batch=harvest.record,
                    should_stop=_stop_hard,            # mid-chunk: WAF wall + burst cap only
                    post_pass_stop=harvest.should_stop,  # harvest judged AFTER all terms + recovery
                    log=lambda m, g=grp: print(f"  [{g}] {m}", flush=True),
                )
            if stopped_early:
                print(f"  [{grp}] {_stop_reason()} stop — banked {len(uncached) - len(remaining)} landed, "
                      f"{len(remaining)} deferred to next burst (skipping per-target grind)", flush=True)
                continue  # do NOT enter the per-target fallback for the deferred misses

            # --- fallback: proven fresh context + fresh search PER target ---
            # Only TRUE not-founds (survived recovery) should reach here now.
            # Reusing one context across many navigate->download->go_back cycles
            # degrades the JSF results page (the 9/9 OR Travelers signature);
            # fresh-per-target is the proven recovery for those. Harvest guard
            # still watches so a collapse here also stops the run cleanly.
            for search_term in search_terms:
                if not remaining:
                    break
                print(f"  [{grp}] FALLBACK search={search_term!r}, attempting {len(remaining)} filing(s)", flush=True)
                still_remaining = []
                for idx, t in enumerate(remaining, 1):
                    _refresh_context()
                    if not _submit_with_retry(state, search_term):
                        still_remaining.append(t)
                        notfound.record(search_ok=False, found=False)  # wall, not a not-found
                        continue
                    _set_rows_per_page_100(page)
                    dest_dir = Path(f"output/pdfs/{state}/{t.filing_id}")
                    try:
                        pdf = download_system_summary_pdf(page, t.filing_id, t.tracking, dest_dir)
                    except Exception as e:
                        print(f"    [{idx}/{len(remaining)}] {t.tracking}: error {type(e).__name__} -> miss", flush=True)
                        pdf = None
                    if pdf:
                        statuses[t.filing_id] = "ok"
                        harvest.record(1, 1)
                        notfound.record(search_ok=True, found=True)
                        print(f"    [{idx}/{len(remaining)}] {t.tracking}: ok", flush=True)
                    else:
                        still_remaining.append(t)
                        harvest.record(0, 1)
                        # clean Begin-Search but row not found -> coverage-gap signal
                        notfound.record(search_ok=True, found=False)
                    if _stop():
                        still_remaining.extend(t2 for t2 in remaining[idx:])
                        print(f"  [{grp}] {_stop_reason()} stop in fallback — "
                              f"{len(still_remaining)} deferred to next burst", flush=True)
                        break
                remaining = still_remaining
                if _stop():
                    break
            if not _stop():
                for t in remaining:
                    statuses[t.filing_id] = "fail:row_not_found"
                    print(f"  {t.tracking}: not found in any of {search_terms!r}", flush=True)
        browser.close()
    if begin_search_budget is not None:
        print(f"\n[burst] Begin-Searches spent this run: {budget.count}/{begin_search_budget} "
              f"(cap {'REACHED — stopped under the wall' if budget.exhausted() else 'not reached'}). "
              f"Per-search clean/wall outcomes: output/serff_diagnostics/search_ledger.csv", flush=True)
    if wall.should_stop():
        print(f"\n[wall] sustained-wall stop tripped ({wall.threshold} consecutive walled "
              f"Begin-Searches) — run halted at the WAF wall; remaining deferred (miss-safe). "
              f"Ledger: output/serff_diagnostics/search_ledger.csv", flush=True)
    return statuses


def _read_pdf_text(pdf_path: Path) -> str:
    """Extract full text from a PDF once (pdfplumber, all pages)."""
    with pdfplumber.open(str(pdf_path)) as pdf:
        return "\n".join((pg.extract_text() or "") for pg in pdf.pages)


def detect_filing_type_and_new_product(pdf_path: Path, text: str | None = None) -> tuple[Optional[str], bool]:
    """Return (filing_type, is_new_product) from the PDF text. `text`: optional
    pre-extracted full text so the caller can read the PDF once and share it with
    parse_filing_summary_pdf (audit decision C, 2026-06-08); when None the PDF is
    opened here (back-compat for other callers, e.g. audit_no_pdf.py)."""
    if text is None:
        text = _read_pdf_text(pdf_path)
    ft = None
    if m := PDF_FILING_TYPE_RE.search(text):
        ft = m.group(1).strip()
    return ft, bool(NEW_PRODUCT_RE.search(text))


# AM Best Disposition Page Data column order (per user spec)
COLUMNS = [
    "state",
    "effective_date",
    "company_name",
    "line_of_business",
    "sub_type_of_insurance",
    "overall_indicated_change",
    "overall_rate_impact",
    "written_premium_change",
    "policyholders_affected",
    "written_premium_for_program",
    "maximum_percent_change",
    "minimum_percent_change",
    "rate_activity",
    "serff_tracking_number",
    "disposition_status",
    "filing_date",
    "source_pdf",
    # Appended after the 17 AM Best Disposition Page Data columns so their
    # column order is preserved. validation_tier is now RE-DERIVED from
    # external_validation (2026-06-04 tier-honesty pass); source +
    # external_validation separate provenance from external validation.
    "validation_tier",
    "source",
    "external_validation",
    "match_strength",
]


def build_rows(state: str, targets: list[Target], backfill_ids: set[str] | None = None) -> tuple[list[dict], dict]:
    """Parse each cached PDF and emit one row per per-company rate row.
    `backfill_ids` = filing_ids recovered by the 2026-06-02 back-extension
    (drives validation_tier). Returns (rows, stats)."""
    backfill_ids = backfill_ids or set()
    rows: list[dict] = []
    stats = {
        "filings_total": len(targets),
        "filings_excluded_form_or_rule": 0,
        "filings_excluded_new_product": 0,
        "filings_excluded_no_pdf": 0,
        "filings_excluded_rate_data_does_not_apply": 0,
        "filings_excluded_out_of_effective_window": 0,
        "filings_emitted": 0,
        "rows_emitted": 0,
        "anchor_match_count": 0,
    }
    for t in targets:
        pdf = Path(f"output/pdfs/{state}/{t.filing_id}/filing_summary.pdf")
        if not pdf.exists() or pdf.stat().st_size < 5000:
            stats["filings_excluded_no_pdf"] += 1; continue
        # Read the PDF text ONCE and share it with both consumers (audit decision
        # C, 2026-06-08) — previously detect_* and parse_* each opened + extracted
        # the full PDF independently (two reads per filing per pass).
        pdf_text = _read_pdf_text(pdf)
        ft_pdf, is_new = detect_filing_type_and_new_product(pdf, text=pdf_text)
        ft = ft_pdf or t.filing_type_xlsx
        if not _is_rate_filing_type(ft):
            stats["filings_excluded_form_or_rule"] += 1; continue
        if is_new:
            stats["filings_excluded_new_product"] += 1; continue
        fs = parse_filing_summary_pdf(pdf, t.tracking, text=pdf_text)
        if not fs.rate_data_applies:
            stats["filings_excluded_rate_data_does_not_apply"] += 1; continue
        if not fs.company_rates:
            # rate_data_applies=True but no rows extracted — record as zero-row anomaly
            print(f"  ! {t.tracking}: rate_data_applies=True but 0 rows extracted")
            stats["filings_excluded_no_pdf"] += 1
            continue

        # Determine rate_activity from disposition status. UT uses REJECTED for
        # disapproved filings; other states use Disapproved/DISAPPROVED. NV uses
        # "Open" for undisposed/in-review filings (added 2026-05-27 in MT/WY/NV
        # expansion — equivalent to PENDING). GA (2026-06-10) adds "Received"
        # (with the state, no disposition yet) and "Exam" (under examination) —
        # both in-review, classified pending like NV's Open; GA's terminal
        # accepted vocabulary ("Acknowledged", "Filed", "Approved") falls
        # through to rate_change.
        ds = (fs.disposition_status or "").upper()
        if "WITHDRAWN" in ds:
            activity = "rate_change_withdrawn"
        elif "DISAPPROV" in ds or "REJECT" in ds:
            activity = "rate_change_disapproved"
        elif "PENDING" in ds or ds in ("OPEN", "RECEIVED", "EXAM"):
            activity = "rate_change_pending"
        else:
            activity = "rate_change"

        eff = fs.effective_date_new or fs.effective_date_renewal
        if not _in_effective_window(eff):
            stats["filings_excluded_out_of_effective_window"] += 1
            continue
        rel_pdf = pdf.relative_to(Path(".")).as_posix() if pdf.is_absolute() is False else pdf.as_posix()
        _src, _ext, _ms, _vt = _relabel(_validation_tier(eff, t.filing_id in backfill_ids), t.tracking)
        for r in fs.company_rates:
            if _is_excluded_subsidiary(r.company_name):
                stats["rows_excluded_filing_vehicle"] = stats.get("rows_excluded_filing_vehicle", 0) + 1
                continue
            rows.append({
                "state": state,
                "effective_date": eff,
                "company_name": r.company_name,
                "line_of_business": t.toi,
                "sub_type_of_insurance": t.sub_toi,
                "overall_indicated_change": r.overall_indicated_change,
                "overall_rate_impact": r.overall_rate_impact,
                "written_premium_change": r.written_premium_change,
                "policyholders_affected": r.policyholders_affected,
                "written_premium_for_program": r.written_premium_for_program,
                "maximum_percent_change": r.maximum_pct_change,
                "minimum_percent_change": r.minimum_pct_change,
                "rate_activity": activity,
                "serff_tracking_number": t.tracking,
                "disposition_status": fs.disposition_status,
                "filing_date": (t.submission_date.isoformat() if hasattr(t.submission_date, "isoformat") else t.submission_date),
                "source_pdf": rel_pdf,
                "validation_tier": _vt,
                "source": _src,
                "external_validation": _ext,
                "match_strength": _ms,
            })
        stats["filings_emitted"] += 1
        stats["rows_emitted"] += len(fs.company_rates)
    return rows, stats


def write_xlsx(rows: list[dict], state: str) -> Path:
    out = Path(f"output/{state.lower()}_final_rates.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "rates"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    wb.save(out)
    return out


# ============================================================
# Idaho-specific anchor verification (SFMA-134676753 vs AM Best)
# ============================================================
ANCHOR_EXPECTED = {
    ("SFMA-134676753", "State Farm Fire and Casualty Company"): dict(
        ind="15.900%", imp="-2.100%", prem_chg="-554469", ph=20679,
        prem_for="26357498", maxp="388.400%", minp="-41.500%",
    ),
    ("SFMA-134676753", "State Farm Mutual Automobile Insurance Company"): dict(
        ind="-2.600%", imp="-9.700%", prem_chg="-25716996", ph=360274,
        prem_for="263832752", maxp="847.900%", minp="-52.200%",
    ),
}


def verify_anchor(rows: list[dict]) -> tuple[int, list[str]]:
    """Return (matches_count, mismatch_messages) for anchor SFMA-134676753."""
    mismatches: list[str] = []
    matched = 0
    by_key = {(r["serff_tracking_number"], r["company_name"]): r for r in rows}
    for key, exp in ANCHOR_EXPECTED.items():
        r = by_key.get(key)
        if not r:
            mismatches.append(f"  MISSING: {key}"); continue
        actual = dict(
            ind=r["overall_indicated_change"], imp=r["overall_rate_impact"],
            prem_chg=r["written_premium_change"], ph=r["policyholders_affected"],
            prem_for=r["written_premium_for_program"],
            maxp=r["maximum_percent_change"], minp=r["minimum_percent_change"],
        )
        ok = True
        for k, ev in exp.items():
            if actual[k] != ev:
                mismatches.append(f"  {key} {k}: got={actual[k]!r} expected={ev!r}"); ok = False
        if ok: matched += 7  # 7 fields per row
    return matched, mismatches


def main():
    ap = argparse.ArgumentParser(
        description="Build {state}_final_rates.xlsx from cached/downloaded SERFF system PDFs.")
    ap.add_argument("state", nargs="?", default="ID", help="two-letter state code (default ID)")
    ap.add_argument("--burst", type=int, default=None, metavar="N",
                    help="Begin-Search budget for THIS run (a 'burst'): stop scheduling new "
                         "searches after N Begin-Searches and defer the rest, so the run stops "
                         "UNDER the WAF wall (~17) instead of grinding into it. Omit = unlimited "
                         "(prior behavior). Miss-safe: a capped run is partial, not corrupt "
                         "(deliverable re-derived from cached PDFs; missing != wrong).")
    ap.add_argument("--no-harvest-early", action="store_true",
                    help="Disable the harvest-early yield watchdog for this run (default: ON). "
                         "Use with --burst when validating Q2 re-batched recovery on a "
                         "degradation-prone state: harvest-early can trip on round-0 misses "
                         "before the recovery round (and later search terms) run, so disabling it "
                         "lets recovery execute, bounded by the --burst Begin-Search cap.")
    ap.add_argument("--wall-stop", type=int, default=2, metavar="K",
                    help="Stop the run after K CONSECUTIVE walled Begin-Searches (the real WAF "
                         "wall: begin_search_link_timeout/405). Default 2; 0 disables. This is the "
                         "no-cap harvest's natural stop and avoids the harvest-early ordering bug — "
                         "it keys on Begin-Search success, never on wrong-term download misses, so "
                         "all search terms + recovery run before the wall halts it.")
    args = ap.parse_args()
    state = args.state.upper()
    _quiet_guard("run_final_rates")  # refuse during a declared SERFF rest window
    t0 = time.time()
    wall_stop = args.wall_stop or None  # 0 -> disabled
    cap = f" (BURST cap: {args.burst} Begin-Searches)" if args.burst is not None else ""
    he = " | harvest-early OFF" if args.no_harvest_early else ""
    ws = f" | wall-stop {wall_stop}" if wall_stop else " | wall-stop OFF"
    print(f"=== {state} final-rates pipeline{cap}{he}{ws} ===")
    targets = load_targets(state)
    print(f"loaded {len(targets)} target-TOI target-carrier filings")
    download_all_pdfs(state, targets, begin_search_budget=args.burst,
                      harvest_early=not args.no_harvest_early, wall_stop=wall_stop)
    backfill_ids = _load_backfill_ids(state)
    if backfill_ids:
        print(f"loaded {len(backfill_ids)} back-extension filing_ids for tiering", flush=True)
    rows, stats = build_rows(state, targets, backfill_ids)
    out = write_xlsx(rows, state)
    elapsed = time.time() - t0

    # storage delta — sum size of cached system PDFs
    pdf_dir = Path(f"output/pdfs/{state}")
    total_kb = sum(p.stat().st_size for p in pdf_dir.rglob("filing_summary.pdf")) / 1024

    # field-completion rates
    completion = {c: 0 for c in COLUMNS}
    for r in rows:
        for c in COLUMNS:
            if r.get(c) not in (None, ""):
                completion[c] += 1

    print("\n=== STATS ===")
    for k, v in stats.items(): print(f"  {k}: {v}")
    print(f"\n=== FIELD COMPLETION ({len(rows)} rows) ===")
    for c in COLUMNS:
        pct = (100 * completion[c] / len(rows)) if rows else 0
        print(f"  {completion[c]:4d}/{len(rows)}  ({pct:5.1f}%)  {c}")
    print(f"\n=== STORAGE ===")
    print(f"  system PDFs cached: {total_kb:.1f} KB across {len(targets)} filings")
    print(f"  avg per filing:     {total_kb / max(1, len(targets)):.1f} KB")
    print(f"\n=== RUNTIME ===")
    print(f"  elapsed: {elapsed:.1f} s ({elapsed/60:.1f} min)")
    print(f"\n=== OUTPUT ===")
    print(f"  -> {out} ({len(rows)} data rows)")

    if state == "ID":
        matched, mismatches = verify_anchor(rows)
        print(f"\n=== ANCHOR (SFMA-134676753 vs AM Best) ===")
        print(f"  matched: {matched}/14 fields")
        if mismatches:
            print("  mismatches:")
            for m in mismatches: print(m)


if __name__ == "__main__":
    main()
