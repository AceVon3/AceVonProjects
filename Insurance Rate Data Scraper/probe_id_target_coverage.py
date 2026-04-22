"""Probe rate_data_applies coverage on all 49 Idaho target-TOI target-carrier filings.

For each filing produce:
    tracking, carrier_group, toi, filing_type_pdf, filing_type_xlsx,
    disposition_status, rate_data_applies, new_product_flag

Then categorize into 3 tables:
  TABLE 1 — should-apply: filing_type IN (Rate, Rate/Rule) AND not new product
            -> True count = real coverage; False count = real gap
  TABLE 2 — should NOT apply: Form / Rule / Scoring Model OR new product
  TABLE 3 — anomalies (off-target TOI; rate_data_applies=True with non-Rate filing_type;
            Table 1 False with rate-change language present in PDF text)

Reuses one playwright session, batches by carrier-group search term.
"""
from __future__ import annotations
import re, sys, zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import openpyxl
import pdfplumber
from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")
from src.config import USER_AGENT, HEADLESS
from src.search import _submit_search, _set_rows_per_page_100, _click_row_to_detail, _back_to_results
from validate_summary_pdf import parse_filing_summary_pdf

STATE = "ID"
TARGET_TOI = ("19.0", "04.0", "03.0")
GROUP_SEARCH = {  # carrier-group -> SERFF search term
    "State Farm":     "state farm",
    "GEICO":          "geico",
    "Allstate":       "allstate",
    "Travelers":      "travelers",
    "Liberty Mutual": "liberty mutual",
}
GROUP_KW = {
    "State Farm":     ["state farm", "mga insurance"],
    "GEICO":          ["geico"],
    "Allstate":       ["allstate", "encompass", "esurance", "integon", "north american insurance"],
    "Travelers":      ["travelers", "standard fire"],
    "Liberty Mutual": ["liberty mutual", "safeco", "american economy"],
}


def carrier_group(name: str) -> Optional[str]:
    n = (name or "").lower()
    for g, kws in GROUP_KW.items():
        if any(k in n for k in kws):
            return g
    return None


@dataclass
class Target:
    tracking: str
    filing_id: str
    company: str
    toi: str
    sub_toi: str
    filing_type_xlsx: str
    product_name: str
    disposition_date: object
    disposition_status_xlsx: str
    group: str
    # filled after probe
    pdf_path: Optional[Path] = None
    download_status: str = ""
    filing_type_pdf: Optional[str] = None
    disposition_status_pdf: Optional[str] = None
    rate_data_applies: Optional[bool] = None
    new_product: Optional[bool] = None
    rate_change_language: bool = False  # presence of "rate change" / "rate increase" / etc.
    company_rates_count: int = 0


# ============================================================
# Load targets from id_final.xlsx
# ============================================================
def load_targets() -> list[Target]:
    wb = openpyxl.load_workbook("output/id_final.xlsx", read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        toi = r[ix["type_of_insurance"]] or ""
        if not any(toi.startswith(p) for p in TARGET_TOI):
            continue
        grp = carrier_group(r[ix["company_name"]])
        if not grp:
            continue
        out.append(Target(
            tracking=r[ix["serff_tracking_number"]] or "",
            filing_id=str(r[ix["filing_id"]] or ""),
            company=r[ix["company_name"]] or "",
            toi=toi,
            sub_toi=r[ix["sub_type_of_insurance"]] or "",
            filing_type_xlsx=r[ix["filing_type"]] or "",
            product_name=r[ix["product_name"]] or "",
            disposition_date=r[ix["disposition_date"]],
            disposition_status_xlsx=r[ix["disposition_status"]] or "",
            group=grp,
        ))
    return out


# ============================================================
# Download minimal zip + extract {tracking}.pdf
# ============================================================
def _download_and_extract(page, t: Target) -> str:
    """Returns 'cached' | 'ok' | 'fail:<reason>'."""
    dest = Path("output/pdfs") / STATE / t.filing_id
    dest.mkdir(parents=True, exist_ok=True)
    out_pdf = dest / "filing_summary.pdf"
    if out_pdf.exists() and out_pdf.stat().st_size > 5000:
        t.pdf_path = out_pdf
        return "cached"
    if not _click_row_to_detail(page, t.filing_id):
        return "fail:row_click"
    page.wait_for_load_state("networkidle", timeout=20000)
    page.wait_for_timeout(500)
    tmp_zip = dest / f"{t.tracking or t.filing_id}.zip"
    try:
        with page.expect_download(timeout=90000) as dl_info:
            page.evaluate("document.getElementById('summaryForm:downloadLink').click();")
        dl_info.value.save_as(str(tmp_zip))
    except Exception as e:
        _back_to_results(page); return f"fail:download:{type(e).__name__}"
    inner = f"{t.tracking}.pdf"
    extracted = False
    try:
        with zipfile.ZipFile(tmp_zip) as zf:
            names = zf.namelist()
            if inner in names:
                with zf.open(inner) as src, open(out_pdf, "wb") as dst:
                    dst.write(src.read())
                extracted = True
            else:
                # try first .pdf at root
                pdfs = [n for n in names if n.endswith(".pdf") and "/" not in n.strip("/")]
                if pdfs:
                    with zf.open(pdfs[0]) as src, open(out_pdf, "wb") as dst:
                        dst.write(src.read())
                    extracted = True
    except Exception as e:
        tmp_zip.unlink(missing_ok=True); _back_to_results(page)
        return f"fail:zip:{type(e).__name__}"
    tmp_zip.unlink(missing_ok=True)
    _back_to_results(page)
    page.wait_for_load_state("domcontentloaded", timeout=15000)
    if not extracted:
        return "fail:no_pdf_in_zip"
    t.pdf_path = out_pdf
    return "ok"


def download_all(targets: list[Target]) -> None:
    by_group: dict[str, list[Target]] = {}
    for t in targets:
        by_group.setdefault(t.group, []).append(t)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
        page = ctx.new_page()
        for grp, items in by_group.items():
            search_term = GROUP_SEARCH[grp]
            # if all cached, skip the search entirely
            uncached = [t for t in items if not (Path("output/pdfs")/STATE/t.filing_id/"filing_summary.pdf").exists()
                        or (Path("output/pdfs")/STATE/t.filing_id/"filing_summary.pdf").stat().st_size <= 5000]
            for t in items:
                if t not in uncached:
                    t.pdf_path = Path("output/pdfs")/STATE/t.filing_id/"filing_summary.pdf"
                    t.download_status = "cached"
            if not uncached:
                print(f"[{grp}] all {len(items)} cached, skipping search")
                continue
            print(f"[{grp}] search={search_term!r}, need {len(uncached)} of {len(items)}")
            if not _submit_search(page, STATE, search_term):
                for t in uncached: t.download_status = "fail:search"
                continue
            _set_rows_per_page_100(page)
            for idx, t in enumerate(uncached, 1):
                # paginate to find row
                found = False
                # reset to first page is hard; rely on search starting on page 1
                for _ in range(10):
                    if page.locator(f'tr[data-rk="{t.filing_id}"]').count():
                        found = True; break
                    nxt = page.locator(".ui-paginator-next").first
                    if not nxt.count() or "ui-state-disabled" in (nxt.get_attribute("class") or ""):
                        break
                    nxt.click(); page.wait_for_load_state("networkidle", timeout=15000)
                if not found:
                    t.download_status = "fail:row_not_found"
                    print(f"  [{idx}/{len(uncached)}] {t.tracking}: row not found"); continue
                t.download_status = _download_and_extract(page, t)
                print(f"  [{idx}/{len(uncached)}] {t.tracking}: {t.download_status}")
                # after _back_to_results, search results should be back; ensure we are
                page.wait_for_timeout(300)
                # re-search if back didn't return us to results
                if not page.locator(".ui-paginator-next").count():
                    _submit_search(page, STATE, search_term)
                    _set_rows_per_page_100(page)
        browser.close()


# ============================================================
# Parse each PDF + extract filing_type / new-product flag
# ============================================================
_FT_RE = re.compile(r"Filing Type:\s*([A-Za-z/ \-]+)\s*$", re.MULTILINE)
_NEW_PRODUCT_RE = re.compile(
    r"\b(New Program|New Product|Initial Filing|Initial Submission|"
    r"Introduction of|Introduction Of)\b",
    re.IGNORECASE,
)
_RATE_LANG_RE = re.compile(r"\b(rate (change|increase|decrease|impact|revision|adjustment)|"
                           r"overall rate|premium impact)\b", re.IGNORECASE)


def parse_one(t: Target) -> None:
    if not t.pdf_path or not t.pdf_path.exists():
        return
    fs = parse_filing_summary_pdf(t.pdf_path, t.tracking)
    t.disposition_status_pdf = fs.disposition_status
    t.rate_data_applies = fs.rate_data_applies
    t.company_rates_count = len(fs.company_rates or [])
    with pdfplumber.open(str(t.pdf_path)) as pdf:
        text = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    if m := _FT_RE.search(text):
        t.filing_type_pdf = m.group(1).strip()
    t.new_product = bool(_NEW_PRODUCT_RE.search(text))
    t.rate_change_language = bool(_RATE_LANG_RE.search(text))


# ============================================================
# Report
# ============================================================
RATE_TYPES = {"Rate", "Rate/Rule"}
NORATE_TYPES = {"Form", "Rule", "Scoring Model", "Withdrawal"}


def render(targets: list[Target]) -> None:
    print("\n" + "=" * 100)
    print("PER-FILING REPORT (49 target-TOI target-carrier filings)")
    print("=" * 100)
    print(f"{'tracking':22s} {'group':14s} {'toi':6s} {'ftype':10s} {'disp':10s} {'rda':5s} {'np':3s} {'rows':4s} {'dl':10s} company")
    for t in sorted(targets, key=lambda x: (x.group, x.tracking)):
        ft = t.filing_type_pdf or t.filing_type_xlsx
        rda = "True" if t.rate_data_applies else ("False" if t.rate_data_applies is False else "?")
        np = "Y" if t.new_product else ("N" if t.new_product is False else "?")
        print(f"{t.tracking:22s} {t.group:14s} {t.toi[:6]:6s} {(ft or '')[:10]:10s} "
              f"{(t.disposition_status_pdf or '')[:10]:10s} {rda:5s} {np:3s} "
              f"{t.company_rates_count:4d} {t.download_status[:10]:10s} {t.company[:40]}")

    # Build categorization
    table1, table2, table3 = [], [], []
    for t in targets:
        ft = t.filing_type_pdf or t.filing_type_xlsx
        is_rate_type = ft in RATE_TYPES
        is_no_rate_type = ft in NORATE_TYPES
        new_prod = bool(t.new_product)
        # off-target TOI? (defensive — shouldn't happen with our load filter)
        if not any(t.toi.startswith(p) for p in TARGET_TOI):
            table3.append((t, "off_target_toi"))
            continue
        if is_rate_type and not new_prod:
            table1.append(t)
        elif is_no_rate_type or new_prod:
            table2.append(t)
        else:
            table3.append((t, f"unclassified_filing_type:{ft}"))
        # extra anomalies
        if t.rate_data_applies and not is_rate_type:
            table3.append((t, "rda_True_but_not_rate_filing"))
        if (t.rate_data_applies is False) and is_rate_type and not new_prod and t.rate_change_language:
            table3.append((t, "Table1_False_with_rate_language"))

    def rda_split(rows):
        return sum(1 for r in rows if r.rate_data_applies), sum(1 for r in rows if r.rate_data_applies is False), sum(1 for r in rows if r.rate_data_applies is None)

    print("\n" + "=" * 100)
    print(f"TABLE 1 — should-apply (Rate or Rate/Rule, NOT new product): {len(table1)} filings")
    print("=" * 100)
    t_, f_, u_ = rda_split(table1)
    print(f"  rate_data_applies=True : {t_}  (real coverage)")
    print(f"  rate_data_applies=False: {f_}  (REAL GAP — likely Form A filings)")
    print(f"  rate_data_applies=None : {u_}  (parse failed)")
    if f_:
        print("  Gap detail (False rows in Table 1):")
        for t in table1:
            if t.rate_data_applies is False:
                print(f"    {t.tracking}  {t.group:14s}  ftype={t.filing_type_pdf or t.filing_type_xlsx:10s}  {t.company[:50]}")

    print("\n" + "=" * 100)
    print(f"TABLE 2 — should NOT apply (Form/Rule/ScoringModel OR new product): {len(table2)} filings")
    print("=" * 100)
    t_, f_, u_ = rda_split(table2)
    print(f"  rate_data_applies=True : {t_}  (suspicious — should be False)")
    print(f"  rate_data_applies=False: {f_}  (correct, expected)")
    print(f"  rate_data_applies=None : {u_}  (parse failed)")
    if t_:
        print("  Suspicious True rows in Table 2:")
        for t in table2:
            if t.rate_data_applies:
                print(f"    {t.tracking}  ftype={t.filing_type_pdf or t.filing_type_xlsx}  np={t.new_product}  {t.company[:50]}")

    print("\n" + "=" * 100)
    print(f"TABLE 3 — anomalies: {len(table3)} entries")
    print("=" * 100)
    for t, reason in table3:
        print(f"  {reason:35s}  {t.tracking}  ftype={t.filing_type_pdf or t.filing_type_xlsx}  toi={t.toi[:15]}  {t.company[:40]}")

    # decision summary
    print("\n" + "=" * 100)
    print("DECISION SUMMARY")
    print("=" * 100)
    t1_T, t1_F, _ = rda_split(table1)
    print(f"  Real-rate-filing coverage today (path A only):  {t1_T}/{t1_T + t1_F}  ({100*t1_T/max(1,t1_T+t1_F):.0f}%)")
    print(f"  Real gap that path C (Form A) would close:     {t1_F} filings")
    print(f"  Out-of-scope (correctly excluded by filer):    {len(table2)} filings")
    print(f"  Anomalies to inspect:                          {len(table3)} entries")


def main():
    targets = load_targets()
    print(f"loaded {len(targets)} target filings")
    download_all(targets)
    print("\n[parse phase]")
    for t in targets:
        try:
            parse_one(t)
        except Exception as e:
            print(f"  parse failed for {t.tracking}: {e}")
    render(targets)


if __name__ == "__main__":
    main()
