"""Union AK search-workbook generations into ak_all_companies_search.xlsx.

Mirrors merge_hi_search.py (union + dedupe by filing_id, dated-rows-win), with an
EXPLICIT ordered input list:

  1. ak_all_companies_search.xlsx          (the cold-sweep base — the terms that
                                            got through; read first, wins dups)
  2. ak_universe_completion_retry.xlsx     (the completion sweep — USAA family,
                                            MGA, Farmers, Travelers + Allstate
                                            date-slice de-cap past the 100-cap)

Both generations are --no-dates (undated), so on a duplicate filing_id the
first-read (base) copy is kept; new filing_ids from the retry are added. The
Allstate date-slices in the retry recover Allstate filings hidden past the
first-sweep 100-cap.
"""
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

OUTPUT_DIR = Path("output")
MAIN = OUTPUT_DIR / "ak_all_companies_search.xlsx"
GENS = [
    MAIN,
    OUTPUT_DIR / "ak_universe_completion_retry.xlsx",
]


def _read(path: Path):
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    header = None
    by_fid: dict[str, list] = {}
    fi = di = None
    for p in GENS:
        if not p.exists():
            raise SystemExit(f"missing input: {p}")
        h2, r2 = _read(p)
        if header is None:
            header = h2
            fi = header.index("filing_id")
            di = header.index("submission_date")
        elif h2 != header:
            raise SystemExit(f"header mismatch in {p}")
        added = upgraded = dup = 0
        for r in r2:
            fid = str(r[fi]) if r[fi] is not None else ""
            if not fid:
                continue
            prev = by_fid.get(fid)
            if prev is None:
                by_fid[fid] = r
                added += 1
            elif prev[di] in (None, "") and r[di] not in (None, ""):
                by_fid[fid] = r  # dated copy beats undated copy
                upgraded += 1
            else:
                dup += 1
        print(f"[merge] {p.name}: +{added} (date-upgraded {upgraded}, dup {dup})")

    idx_target = header.index("target_company")
    idx_serff = header.index("serff_tracking_number")
    idx_intt = header.index("in_target_lines")
    rows = sorted(by_fid.values(), key=lambda r: (r[idx_target] or "", r[idx_serff] or ""))
    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(MAIN))
    print(f"\n[write] {MAIN}: {len(rows)} total rows")
    raw = Counter(r[idx_target] for r in rows)
    intt = Counter(r[idx_target] for r in rows if r[idx_intt])
    print(f"{'target_company':28s} {'raw':>5s} {'in-target':>9s}")
    for c in sorted(raw):
        print(f"  {c:26s} {raw[c]:5d} {intt[c]:9d}")
    print(f"\nIN-TARGET TOTAL (TOI-level; brand-exclusions applied later at parse): {sum(intt.values())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
