"""Merge per-carrier WA search retries into wa_all_companies_search.xlsx.

Carriers that time out on "Begin Search" link click during the combined
--all-companies run are retried in fresh processes. Mirrors merge_ut_search.py
and merge_id_search.py.

Run history:
- Lookback re-scrape (DATE_FROM=07/01/2024): Travelers + Liberty Mutual +
  Safeco timed out, retried.
"""
from pathlib import Path
from openpyxl import load_workbook, Workbook

OUTPUT_DIR = Path("output")

MAIN = OUTPUT_DIR / "wa_all_companies_search.xlsx"
RETRY_FILES = [
    OUTPUT_DIR / "wa_travelers_search.xlsx",
    OUTPUT_DIR / "wa_liberty_mutual_search.xlsx",
    OUTPUT_DIR / "wa_safeco_search.xlsx",
]


def _read_rows(path: Path) -> tuple[list, list[list]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    header, rows = _read_rows(MAIN)
    print(f"[main] {MAIN.name}: {len(rows)} rows")

    idx_target = header.index("target_company")
    retry_carriers = {"Travelers", "Liberty Mutual", "Safeco"}
    before = len(rows)
    rows = [r for r in rows if r[idx_target] not in retry_carriers]
    print(f"[main] dropped {before - len(rows)} pre-existing rows for retry carriers")

    for path in RETRY_FILES:
        h2, r2 = _read_rows(path)
        if h2 != header:
            raise SystemExit(f"header mismatch in {path}")
        print(f"[merge] {path.name}: {len(r2)} rows")
        rows.extend(r2)

    idx_serff = header.index("serff_tracking_number")
    rows.sort(key=lambda r: (r[idx_target] or "", r[idx_serff] or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(MAIN))

    print(f"\n[write] {MAIN}: {len(rows)} total rows")
    from collections import Counter
    per = Counter(r[idx_target] for r in rows)
    for c in sorted(per):
        print(f"  {c:16s} {per[c]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
