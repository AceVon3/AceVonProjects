"""Merge an incremental back-fill search slice into a state's canonical
search workbook, deduping by filing_id.

Used by the 2026-06-02 effective-date back-extension (2024-01-01 floor): the
new submission slice 2023-07-01 -> 2024-06-30 is searched into a separate
`{state}_all_companies_search_backfill2024.xlsx` (via run_search.py
--out-suffix) so it doesn't clobber the existing workbook; this script unions
it into `{state}_all_companies_search.xlsx`.

Only the "Filings" sheet is carried (matches merge_wy_search.py precedent and
what run_{state}_full.py::_load_filings consumes). The pre-merge canonical
workbook is backed up to `*.prebackfill.bak.xlsx` first.

    python merge_backfill_slice.py WY
"""
from __future__ import annotations

import shutil
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

OUTPUT_DIR = Path("output")
SLICE_SUFFIX = "_backfill2024"


def _read_filings(path: Path) -> tuple[list, list[list]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    if len(sys.argv) < 2:
        raise SystemExit("usage: python merge_backfill_slice.py <STATE>")
    state = sys.argv[1].upper()
    slug = state.lower()
    main_path = OUTPUT_DIR / f"{slug}_all_companies_search.xlsx"
    slice_path = OUTPUT_DIR / f"{slug}_all_companies_search{SLICE_SUFFIX}.xlsx"

    if not main_path.exists():
        raise SystemExit(f"missing canonical workbook: {main_path}")
    if not slice_path.exists():
        raise SystemExit(f"missing slice workbook: {slice_path}")

    header, main_rows = _read_filings(main_path)
    s_header, slice_rows = _read_filings(slice_path)
    if s_header != header:
        raise SystemExit(
            f"header mismatch:\n  main : {header}\n  slice: {s_header}"
        )

    idx_fid = header.index("filing_id")
    idx_target = header.index("target_company")
    idx_serff = header.index("serff_tracking_number")
    idx_sub = header.index("submission_date")

    print(f"[{state}] canonical : {len(main_rows)} rows")
    print(f"[{state}] slice     : {len(slice_rows)} rows  ({slice_path.name})")

    # Union, dedup by filing_id. Existing canonical rows win on collision
    # (they carry any enrichment that the search-only slice lacks).
    by_fid: dict[str, list] = {}
    for r in main_rows:
        by_fid[str(r[idx_fid])] = r
    collisions = 0
    added = 0
    for r in slice_rows:
        fid = str(r[idx_fid])
        if fid in by_fid:
            collisions += 1
            continue
        by_fid[fid] = r
        added += 1

    merged = list(by_fid.values())
    merged.sort(key=lambda r: (r[idx_target] or "", r[idx_serff] or "", str(r[idx_fid])))

    print(f"[{state}] new unique : {added}")
    print(f"[{state}] collisions : {collisions} (already present, slice row dropped)")
    print(f"[{state}] merged tot : {len(merged)}")

    # Backup before overwrite.
    bak = main_path.with_suffix(".prebackfill.bak.xlsx")
    if not bak.exists():
        shutil.copy2(main_path, bak)
        print(f"[{state}] backup     -> {bak.name}")
    else:
        print(f"[{state}] backup     exists ({bak.name}), not overwriting")

    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in merged:
        ws.append(r)
    wb.save(str(main_path))
    print(f"[{state}] wrote      -> {main_path}")

    # Submission-date span sanity (helps confirm the slice extended the window).
    dates = [str(r[idx_sub]) for r in merged if r[idx_sub]]
    if dates:
        print(f"[{state}] submission span: {min(dates)} .. {max(dates)}")
    per = Counter(r[idx_target] for r in merged)
    for c in sorted(per, key=lambda x: (x or "")):
        print(f"    {str(c):18s} {per[c]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
