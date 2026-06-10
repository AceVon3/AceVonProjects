"""Merge the Progressive NM search retry into nm_all_companies_search.xlsx.

Run history:
- Initial NM scrape (2026-06-08): Progressive failed during the sequential
  all-companies sweep with a "Begin Search" Locator.click Timeout (the OR/NV
  silent-fail / rate-limit pattern) -> 0 rows in the main workbook. Verify-retry
  in isolation recovered 25 Progressive filings (nm_progressive_search.xlsx),
  confirming the original 0 was a false zero, not a genuine absence.
"""
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook, Workbook

OUTPUT_DIR = Path("output")

MAIN = OUTPUT_DIR / "nm_all_companies_search.xlsx"
RETRY_FILES = [
    OUTPUT_DIR / "nm_progressive_search.xlsx",
]
RETRY_CARRIERS = {"Progressive"}


def _read_rows(path: Path) -> tuple[list, list[list]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    header, rows = _read_rows(MAIN)
    print(f"[main] {MAIN.name}: {len(rows)} rows")

    idx_target = header.index("target_company")
    before = len(rows)
    rows = [r for r in rows if r[idx_target] not in RETRY_CARRIERS]
    print(f"[main] dropped {before - len(rows)} pre-existing rows for retry carriers")

    for path in RETRY_FILES:
        h2, r2 = _read_rows(path)
        if h2 != header:
            raise SystemExit(f"header mismatch in {path}")
        print(f"[merge] {path.name}: {len(r2)} rows")
        rows.extend(r2)

    idx_serff = header.index("serff_tracking_number")
    rows.sort(key=lambda r: (r[idx_target] or "", r[idx_serff] or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(MAIN))

    print(f"\n[write] {MAIN}: {len(rows)} total rows")
    per = Counter(r[idx_target] for r in rows)
    for c in sorted(per):
        print(f"  {c:16s} {per[c]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
