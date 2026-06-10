"""B-validation probe: search-workbook schema + submission_date coverage per state."""
import openpyxl

for st in ("nm", "ut", "mt", "id", "or"):
    p = f"output/{st}_all_companies_search.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True)
    ws = wb["Filings"] if "Filings" in wb.sheetnames else wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    ix = {h: i for i, h in enumerate(hdr)}
    n = len(rows)

    def filled(col):
        if col not in ix:
            return -1
        return sum(1 for r in rows if r[ix[col]] not in (None, ""))

    sd, ft, stoi = filled("submission_date"), filled("filing_type"), filled("sub_type_of_insurance")
    print(f"{st.upper()}: {n} filings | submission_date {sd}/{n} (missing {n - sd}) | "
          f"filing_type {ft}/{n} | sub_toi {stoi}/{n}")
    print(f"   sheets={wb.sheetnames}")
    print(f"   cols={hdr}")
    wb.close()
