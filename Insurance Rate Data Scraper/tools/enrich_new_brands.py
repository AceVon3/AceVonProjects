"""Enrich newly-added Safeco/Encompass filings and append to {state}_final.xlsx.

Loads the brand-specific search workbooks (output/{state}_{brand}_search.xlsx),
filters to filing_ids not yet present in {state}_final.xlsx, runs detail
enrichment + PDF download on those, and rewrites the final xlsx with the
combined set.

Usage:
    python tools/enrich_new_brands.py WA
    python tools/enrich_new_brands.py ID
    python tools/enrich_new_brands.py CO
"""
from __future__ import annotations

import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import HEADLESS, OUTPUT_DIR, REQUEST_DELAY, USER_AGENT
from src.detail import enrich_filing
from src.models import AttachedPdf, Filing
from src.output import write_excel
from src.search import _parse_date, _set_rows_per_page_100, _submit_search


def _parse_any_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return _parse_date(s)


def _row_to_filing(row: dict) -> Filing:
    naic_raw = row.get("naic_codes") or ""
    naic = [c.strip() for c in str(naic_raw).split(";") if c.strip()]
    return Filing(
        state=row.get("state") or "",
        serff_tracking_number=row.get("serff_tracking_number") or "",
        filing_id=str(row.get("filing_id") or ""),
        company_name=row.get("company_name") or "",
        target_company=row.get("target_company") or "",
        naic_codes=naic,
        product_name=row.get("product_name"),
        type_of_insurance=row.get("type_of_insurance"),
        sub_type_of_insurance=row.get("sub_type_of_insurance"),
        filing_type=row.get("filing_type"),
        filing_status=row.get("filing_status"),
        submission_date=_parse_any_date(row.get("submission_date")),
        detail_url=row.get("detail_url"),
    )


def _load_xlsx_rows(path: Path, sheet: str | None = None) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    return [dict(zip(header, r)) for r in rows[1:]]


def _hydrate_filing_from_final_row(row: dict) -> Filing:
    """Rebuild a Filing from a {state}_final.xlsx row, including PDFs."""
    naic_raw = row.get("naic_codes") or ""
    naic = [c.strip() for c in str(naic_raw).split(";") if c.strip()]
    pdfs_raw = row.get("pdfs") or ""
    pdfs: list[AttachedPdf] = []
    if pdfs_raw:
        for entry in str(pdfs_raw).split("|"):
            entry = entry.strip()
            if not entry:
                continue
            parts = entry.split("::")
            name = parts[0] if parts else ""
            local = parts[1] if len(parts) > 1 else None
            pdfs.append(AttachedPdf(category="", display_name=name, url="", local_path=local))
    fields_raw = row.get("pdf_parse_fields_found") or ""
    fields = [c.strip() for c in str(fields_raw).split(";") if c.strip()]
    return Filing(
        state=row.get("state") or "",
        serff_tracking_number=row.get("serff_tracking_number") or "",
        filing_id=str(row.get("filing_id") or ""),
        company_name=row.get("company_name") or "",
        target_company=row.get("target_company") or "",
        naic_codes=naic,
        product_name=row.get("product_name"),
        type_of_insurance=row.get("type_of_insurance"),
        sub_type_of_insurance=row.get("sub_type_of_insurance"),
        filing_type=row.get("filing_type"),
        filing_status=row.get("filing_status"),
        submission_date=_parse_any_date(row.get("submission_date")),
        disposition_date=_parse_any_date(row.get("disposition_date")),
        disposition_status=row.get("disposition_status"),
        state_status=row.get("state_status"),
        requested_rate_effect=row.get("requested_rate_effect"),
        approved_rate_effect=row.get("approved_rate_effect"),
        overall_rate_effect=row.get("overall_rate_effect"),
        affected_policyholders=row.get("affected_policyholders"),
        written_premium_volume=row.get("written_premium_volume"),
        annual_premium_impact_dollars=row.get("annual_premium_impact_dollars"),
        current_avg_premium=row.get("current_avg_premium"),
        proposed_avg_premium=row.get("proposed_avg_premium"),
        premium_change_dollars=row.get("premium_change_dollars"),
        program_name=row.get("program_name"),
        filing_reason=row.get("filing_reason"),
        prior_approval=row.get("prior_approval"),
        pdfs=pdfs,
        pdf_parse_status=row.get("pdf_parse_status") or "not_attempted",
        pdf_parse_fields_found=fields,
        detail_url=row.get("detail_url"),
        in_target_lines=row.get("in_target_lines"),
        is_resubmission_of=row.get("is_resubmission_of"),
    )


def main(state: str) -> int:
    state_l = state.lower()
    final_path = OUTPUT_DIR / f"{state_l}_final.xlsx"

    existing_rows = _load_xlsx_rows(final_path, "Filings")
    existing_filings = [_hydrate_filing_from_final_row(r) for r in existing_rows]
    # Treat rows that lack type_of_insurance as un-enriched: previous run failed
    # to hit the SERFF search for them, so they came in with no detail data.
    enriched_ids = {f.filing_id for f in existing_filings if f.type_of_insurance}
    existing_filings = [f for f in existing_filings if f.filing_id in enriched_ids]
    existing_ids = enriched_ids
    print(f"[load] {final_path.name}: {len(existing_filings)} enriched filings (rows with no TOI dropped for retry)", flush=True)

    new_filings: list[Filing] = []
    for brand in ("safeco", "encompass"):
        brand_path = OUTPUT_DIR / f"{state_l}_{brand}_search.xlsx"
        if not brand_path.exists():
            print(f"[skip] {brand_path.name} not found", flush=True)
            continue
        rows = _load_xlsx_rows(brand_path)
        for r in rows:
            fid = str(r.get("filing_id") or "")
            if not fid or fid in existing_ids:
                continue
            new_filings.append(_row_to_filing(r))
        print(f"[load] {brand_path.name}: {len(rows)} rows -> {sum(1 for f in new_filings if f.target_company.lower()==brand)} new", flush=True)

    if not new_filings:
        print("[done] no new filings to enrich", flush=True)
        return 0

    print(f"\n[enrich] {len(new_filings)} new filings", flush=True)
    groups: dict[tuple[str, str], list[Filing]] = defaultdict(list)
    for f in new_filings:
        groups[(f.state, f.target_company)].append(f)

    started = time.time()
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=HEADLESS)
        try:
            for (st, company), group in sorted(groups.items(), key=lambda x: len(x[1])):
                print(f"\n[group] {st} / {company}: {len(group)} filings", flush=True)
                ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
                page = ctx.new_page()
                try:
                    if not _submit_search(page, st, company):
                        print(f"  ! search failed, skipping {company}", flush=True)
                        continue
                    _set_rows_per_page_100(page)
                    for i, f in enumerate(group, 1):
                        enrich_filing(page, f, download_pdfs=True)
                        if i % 5 == 0 or i == len(group):
                            elapsed = (time.time() - started) / 60
                            print(f"    [{i}/{len(group)}] elapsed {elapsed:.1f}m", flush=True)
                        time.sleep(REQUEST_DELAY)
                finally:
                    ctx.close()
        finally:
            browser.close()

    combined = existing_filings + new_filings
    write_excel(combined, final_path)
    print(f"\n[save] {final_path.name}: {len(combined)} total filings ({len(new_filings)} new)", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1] if len(sys.argv) > 1 else "WA"))
