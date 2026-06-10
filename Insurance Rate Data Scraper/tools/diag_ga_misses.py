"""Classify GA cross-check misses: structural download misses vs visibility gaps vs exclusions."""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

import run_final_rates as rf

# 1. The two structural download misses — what are they?
targets, _ = rf.load_targets_search("GA")
for t in targets:
    if t.tracking in ("TRVD-G134718262", "TRVD-G134818683"):
        print(f"download-miss {t.tracking}: company={t.company!r} sub_toi={t.sub_toi!r} "
              f"type={t.filing_type_xlsx!r} submitted={t.submission_date}")

# 2. Progressive 12/05/25 event — anything in the search universe?
print("\nProgressive filings in GA search universe (any TOI):")
wb = openpyxl.load_workbook("output/ga_all_companies_search.xlsx", read_only=True)
ws = wb["Filings"] if "Filings" in wb.sheetnames else wb.active
hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
ix = {h: i for i, h in enumerate(hdr)}
n = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    tc = (r[ix["target_company"]] or "")
    if tc != "Progressive":
        continue
    n += 1
    sub = r[ix["sub_type_of_insurance"]] or ""
    if sub.startswith("19."):
        print(f"  {r[ix['serff_tracking_number']]} sub={sub[:40]!r} type={r[ix['filing_type']]!r} submitted={r[ix['submission_date']]}")
print(f"  (total Progressive rows in universe: {n})")

# 3. Our GA Progressive rows in the deliverable
print("\nour GA Progressive deliverable rows (eff dates):")
wb2 = openpyxl.load_workbook("output/ga_final_rates.xlsx", read_only=True)
ws2 = wb2["rates"]
rows = list(ws2.iter_rows(values_only=True))
h2 = list(rows[0])
for r in rows[1:]:
    d = dict(zip(h2, r))
    if "progressive" in (d["company_name"] or "").lower():
        print(f"  {d['serff_tracking_number']} {d['company_name'][:45]} eff={d['effective_date']} imp={d['overall_rate_impact']}")
