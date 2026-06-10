"""One-off audit (2026-06-10, B deployment): for every state, does any
non-archive side search workbook contain filing_ids missing from the
universe (all_companies + mga_insurance)? Decides whether run_final_rates'
universe verification can be a hard error."""
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")


def ids(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Filings"] if "Filings" in wb.sheetnames else wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    fi = hdr.index("filing_id")
    out = {str(r[fi]) for r in ws.iter_rows(min_row=2, values_only=True) if r[fi] is not None}
    wb.close()
    return out


ARCHIVE = re.compile(r"(prebackfill|pre_lookback|backfill2024|\.bak)")
states = ["az", "co", "ga", "id", "mt", "nm", "nv", "or", "ut", "wa", "wy"]
for st in states:
    allc = Path(f"output/{st}_all_companies_search.xlsx")
    if not allc.exists():
        print(f"{st.upper()}: no all_companies workbook")
        continue
    universe = ids(allc)
    mga = Path(f"output/{st}_mga_insurance_search.xlsx")
    if mga.exists():
        universe |= ids(mga)
    extras_report = []
    for p in sorted(Path("output").glob(f"{st}_*_search*.xlsx")):
        if p in (allc, mga) or ARCHIVE.search(p.name):
            continue
        extra = ids(p) - universe
        if extra:
            extras_report.append(f"{p.name}: {len(extra)} extra ({sorted(extra)[:5]})")
    bf_extra = 0
    for p in sorted(Path("output").glob(f"{st}_*backfill2024.xlsx")):
        bf_extra += len(ids(p) - universe)
    status = "; ".join(extras_report) if extras_report else "all side workbooks subset"
    print(f"{st.upper()}: universe={len(universe)} | {status} | backfill2024 extras: {bf_extra}")
