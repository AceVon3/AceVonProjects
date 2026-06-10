"""Diagnostic: target_company values among GA Allstate-group targets, cached vs not."""
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.stdout.reconfigure(encoding="utf-8")

import run_final_rates as rf
import validate_batch_download as v

wb = openpyxl.load_workbook("output/ga_all_companies_search.xlsx", read_only=True)
ws = wb["Filings"] if "Filings" in wb.sheetnames else wb.active
hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
ix = {h: i for i, h in enumerate(hdr)}
tc_by_fid = {str(r[ix["filing_id"]]): (r[ix["target_company"]] or "")
             for r in ws.iter_rows(min_row=2, values_only=True) if r[ix["filing_id"]]}
wb.close()

targets, _ = rf.load_targets_search("GA")
cnt_c, cnt_u = Counter(), Counter()
for t in targets:
    if t.group != "Allstate":
        continue
    p = v.cache_pdf("GA", t.filing_id)
    cached = p.exists() and p.stat().st_size > 5000
    (cnt_c if cached else cnt_u)[tc_by_fid.get(t.filing_id, "?")] += 1
print("Allstate-group targets by target_company:")
print("  cached:  ", dict(cnt_c))
print("  uncached:", dict(cnt_u))
