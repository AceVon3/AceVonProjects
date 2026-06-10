"""B-validation probe: where do GNSC/MGA filings live — search workbook, final, or both?"""
import openpyxl


def trackings(path, sheet=None):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {h: i for i, h in enumerate(hdr)}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        tk = r[ix["serff_tracking_number"]]
        if tk:
            out[tk] = r[ix["company_name"]]
    wb.close()
    return out


for st in ("ut", "id", "nm", "mt", "or", "az", "nv", "wa", "co", "wy"):
    s = trackings(f"output/{st}_all_companies_search.xlsx", "Filings")
    f = trackings(f"output/{st}_final.xlsx")
    gnsc_s = sorted(t for t in s if str(t).startswith("GNSC"))
    gnsc_f = sorted(t for t in f if str(t).startswith("GNSC"))
    mga_s = sorted(t for t, c in s.items() if "mga insurance" in str(c).lower())
    mga_f = sorted(t for t, c in f.items() if "mga insurance" in str(c).lower())
    only_final = sorted(set(f) - set(s))
    print(f"{st.upper()}: search={len(s)} final={len(f)} | GNSC in search: {len(gnsc_s)}, "
          f"in final: {len(gnsc_f)} | MGA-co in search: {len(mga_s)}, in final: {len(mga_f)} | "
          f"rows only in final: {len(only_final)}")
    if only_final:
        non_gnsc = [t for t in only_final if not str(t).startswith("GNSC")]
        print(f"   only-in-final trackings: {only_final[:12]}{'...' if len(only_final) > 12 else ''}")
        if non_gnsc:
            print(f"   !! only-in-final NON-GNSC: {non_gnsc[:12]}{'...' if len(non_gnsc) > 12 else ''}")
