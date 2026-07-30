"""Overnight orchestrator — full collection arcs for IA then SC (detached-safe).

Per state, sequentially:
  1. SWEEP: run_search --all-companies (27 terms, ledger armed).
  2. RETRY ROUNDS: terms whose LATEST ledger outcome for this state is non-ok
     get single-term retries after a 20-min rest; up to 3 rounds (30/45-min
     rests between rounds); stop a round early on 2 consecutive walls.
  3. MERGE: union base + every {st}_*_search_retry.xlsx (dedupe by filing_id,
     dated-copy-wins) back into {st}_all_companies_search.xlsx (base backed up
     to .presweepmerge.bak.xlsx first).
  4. BURSTS: run_final_rates {ST} --burst 18 --wall-stop 2 --no-harvest-early
     every 15 min until all targets cached; zero-progress bursts back off
     (30 -> 60 min) and after 3 consecutive zero-progress bursts the state is
     BANKED for morning (miss-safe).
Cross-check + import are deliberately NOT run — morning judgment steps.
Every phase logs loudly; a kill anywhere resumes by re-running this script
(sweep skipped if the universe workbook exists, retries skip ok terms, bursts
cache-skip).
"""
from __future__ import annotations

import subprocess
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402

STATES = ["NY", "TX"]  # gap-state batch 2026-07-22: probes confirmed all three live on SERFF Filing Access with in-target rate filings (AL 8 / LA 14 / NC 3 State-Farm-only). WY probed structurally empty (0 rate filings, no date filter) — excluded. Prior chain KY..TN completed 07-22.
LEDGER = Path("output/serff_diagnostics/search_ledger.csv")
ALL_TERMS = ["state farm", "mga insurance", "geico", "government employees",
             "allstate", "encompass", "travelers", "travco", "liberty mutual",
             "safeco", "liberty insurance", "general insurance", "american states",
             "first national insurance", "progressive", "artisan and truckers",
             "usaa", "united services", "garrison", "farmers", "mid-century",
             "fire insurance exchange", "truck insurance exchange", "nationwide",
             "american family", "american standard insurance", "country"]


def log(msg: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)


def run(args: list[str], logfile: str) -> int:
    with open(logfile, "a", encoding="utf-8") as f:
        return subprocess.run([sys.executable, "-u"] + args,
                              stdout=f, stderr=subprocess.STDOUT, check=False).returncode


def latest_outcomes(st: str) -> dict[str, str]:
    out: dict[str, str] = {}
    if not LEDGER.exists():
        return out
    for line in LEDGER.read_text(encoding="utf-8", errors="replace").splitlines():
        parts = line.split(",")
        if len(parts) >= 4 and parts[1] == st:
            out[parts[2].lower()] = parts[3]
    return out


def failed_terms(st: str) -> list[str]:
    o = latest_outcomes(st)
    return [t for t in ALL_TERMS if o.get(t, "missing") != "ok"]


def merge(st: str) -> None:
    main = Path(f"output/{st.lower()}_all_companies_search.xlsx")
    bak = Path(f"output/{st.lower()}_all_companies_search.presweepmerge.bak.xlsx")
    if not bak.exists():
        bak.write_bytes(main.read_bytes())
    gens = [main] + sorted(Path("output").glob(f"{st.lower()}_*_search_retry.xlsx"))
    header = None
    by_fid: dict[str, list] = {}
    fi = di = None
    for p in gens:
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb["Filings"]
        it = ws.iter_rows(values_only=True)
        h = list(next(it))
        if header is None:
            header = h
            fi = h.index("filing_id"); di = h.index("submission_date")
        for r in it:
            r = list(r)
            fid = str(r[fi]) if r[fi] is not None else ""
            if not fid:
                continue
            prev = by_fid.get(fid)
            if prev is None:
                by_fid[fid] = r
            elif prev[di] in (None, "") and r[di] not in (None, ""):
                by_fid[fid] = r
        wb.close()
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Filings"
    ws.append(header)
    for r in by_fid.values():
        ws.append(r)
    out.save(str(main))
    log(f"{st}: merged {len(gens)} generations -> {len(by_fid)} filings")


def uncached(st: str) -> int:
    from run_final_rates import load_targets_search
    targets, _ = load_targets_search(st)
    return sum(1 for t in targets
               if not (Path(f"output/pdfs/{st}/{t.filing_id}/filing_summary.pdf").exists()
                       and Path(f"output/pdfs/{st}/{t.filing_id}/filing_summary.pdf").stat().st_size > 5000))


for st in STATES:
    lo = st.lower()
    log(f"===== {st} ARC START =====")
    # 1. sweep (skip if universe exists — resume support)
    uni = Path(f"output/{lo}_all_companies_search.xlsx")
    if uni.exists() and uncached(st) == 0:
        # Collection already complete — skip the whole arc (the ledger-based
        # retry heuristic reads burst-phase walls as failed sweep terms and
        # would sleep 20 min for terms whose filings are all cached).
        log(f"{st}: COLLECTION ALREADY COMPLETE — skipping arc")
        continue
    if uni.exists():
        log(f"{st}: universe exists, skipping sweep")
    else:
        log(f"{st}: SWEEP starting")
        run(["run_search.py", "--all-companies", "--state", st,
             "--diag", "output/serff_diagnostics"], f"output/{lo}_sweep.log")
        log(f"{st}: sweep done; failed terms: {failed_terms(st)}")
    # 2. retry rounds
    for rnd in range(1, 4):
        bad = failed_terms(st)
        if not bad:
            break
        rest = 300 if rnd == 1 else (900 if rnd == 2 else 1800)  # 5-min first retry per user 2026-07-10
        log(f"{st}: retry round {rnd} for {bad} after {rest}s rest")
        time.sleep(rest)
        consec = 0
        for t in bad:
            run(["run_search.py", "--state", st, "--company", t,
                 "--out-suffix", "_retry", "--no-dates",
                 "--diag", "output/serff_diagnostics"], f"output/{lo}_retry_console.log")
            ok = latest_outcomes(st).get(t.lower()) == "ok"
            consec = 0 if ok else consec + 1
            log(f"{st}: retry {t!r} -> {'ok' if ok else 'FAILED'}")
            if consec >= 2:
                log(f"{st}: 2 consecutive walls in retry — banking round")
                break
    # 3. merge
    merge(st)
    # 4. bursts
    zero_streak = 0
    for i in range(1, 41):  # cap raised 11->40 (2026-07-14): the zero-progress backoff is the real runaway guard; the old cap forced avoidable relaunch round-trips on big states (MD, MI)
        before = uncached(st)
        if before == 0:
            log(f"{st}: COLLECTION COMPLETE")
            break
        log(f"{st}: burst-{i} ({before} uncached)")
        run(["run_final_rates.py", st, "--burst", "18", "--wall-stop", "2",
             "--no-harvest-early"], f"output/{lo}_burst{i}.log")
        after = uncached(st)
        if after == 0:
            log(f"{st}: COLLECTION COMPLETE after burst-{i}")
            break
        if after >= before:
            zero_streak += 1
            rest = min(1800 * zero_streak, 3600)
            log(f"{st}: burst-{i} zero progress ({after} remain), streak {zero_streak}, backing off {rest}s")
            if zero_streak >= 3:
                log(f"{st}: BANKED for morning ({after} uncached)")
                break
        else:
            zero_streak = 0
            rest = 300  # 5-min cadence trial per user 2026-07-28 (was 900; zero-progress backoffs unchanged — the WAF guard)
            log(f"{st}: burst-{i} banked {before - after} ({after} remain), resting {rest}s")
        time.sleep(rest)
    log(f"===== {st} ARC END =====")
    time.sleep(900)  # inter-state rest (chain mode, 2026-07-12)

log("OVERNIGHT RUNNER DONE")
