"""Union all VA search-workbook generations into va_all_companies_search.xlsx.

VA collects fresh on the 13-brand keyword set (19 keywords), so unlike GA
there is no pre-existing 9-keyword baseline to protect: every generation —
the base all-companies sweep, suffixed sweep retries
(va_all_companies_search_retry*.xlsx), and any per-keyword isolation retries
(va_{keyword}_search*.xlsx) — is unioned and deduped by filing_id. Rows WITH
a submission_date win over undated duplicates (WAF-challenged date fetches
leave blanks; a later calmer generation upgrades them). Matches the
merge_ga_newcarriers.py union-dedup pattern; verify_search_universe requires
side workbooks be merged (or archive-suffixed) before run_final_rates runs.
"""
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

OUTPUT_DIR = Path("output")
MAIN = OUTPUT_DIR / "va_all_companies_search.xlsx"


def _read(path: Path):
    wb = load_workbook(path, read_only=True)
    ws = wb["Filings"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return header, rows


def main() -> int:
    gens = [p for p in sorted(OUTPUT_DIR.glob("va_*_search*.xlsx"))]
    if not gens:
        raise SystemExit("no va_*_search*.xlsx workbooks found")
    header = None
    by_fid: dict[str, list] = {}
    for p in gens:
        h2, r2 = _read(p)
        if header is None:
            header = h2
            fi = header.index("filing_id")
            di = header.index("submission_date")
        elif h2 != header:
            raise SystemExit(f"header mismatch in {p}")
        added = upgraded = dup = 0
        for r in r2:
            fid = str(r[fi]) if r[fi] is not None else ""
            if not fid:
                continue
            prev = by_fid.get(fid)
            if prev is None:
                by_fid[fid] = r
                added += 1
            elif prev[di] in (None, "") and r[di] not in (None, ""):
                by_fid[fid] = r  # dated copy beats undated copy
                upgraded += 1
            else:
                dup += 1
        print(f"[merge] {p.name}: +{added} (date-upgraded {upgraded}, dup {dup})")

    idx_target = header.index("target_company")
    idx_serff = header.index("serff_tracking_number")
    rows = sorted(by_fid.values(), key=lambda r: (r[idx_target] or "", r[idx_serff] or ""))
    wb = Workbook()
    ws = wb.active
    ws.title = "Filings"
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(MAIN))
    print(f"\n[write] {MAIN}: {len(rows)} total rows")
    per = Counter(r[idx_target] for r in rows)
    for c in sorted(per):
        print(f"  {c:24s} {per[c]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
