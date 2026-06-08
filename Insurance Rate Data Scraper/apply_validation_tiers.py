"""Apply the EARNED four-tier external_validation to per-state final_rates files,
from the per-row corroboration artifacts (CROSS_CHECK_STANDARD.md step 5).

Run AFTER run_final_rates.py (base labels) + the per-state compare_*_ambest.py
(which emit output/corroboration/*.csv). LABEL-ONLY: only external_validation,
match_strength, and validation_tier change; source + all rate-value columns are
copied verbatim.

  field_validated   - documented all-field anchor match (SFMA-134676753)
  ambest_cross_checked - matched an AM Best entry (match_strength: direct |
                      date_relaxed | reclassified) per the corroboration CSVs
  pipeline_extracted_in_validated_window - eff>=2025 in a cross-checked state,
                      not individually matched
  pipeline_extracted - CO / ID non-anchor / all eff-2024 extension

    python apply_validation_tiers.py            # all states
    python apply_validation_tiers.py AZ CO      # subset
"""
import csv
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ANCHOR = "SFMA-134676753"
# States with a per-row corroboration artifact OR a documented cross-check.
# Only states with artifacts can earn ambest_cross_checked; WA is cross-checked
# (documented 12/14) but has no reusable artifact, so its in-window rows stay
# pipeline_extracted_in_validated_window (see TIER_RELABEL.md / decision 3).
CROSS_CHECKED_STATES = {"AZ", "MT", "NV", "OR", "UT", "WA"}
RANK = {"direct": 3, "date_relaxed": 2, "reclassified": 1}
VT = {"field_validated": "ambest_validated", "ambest_cross_checked": "ambest_validated",
      "pipeline_extracted_in_validated_window": "pipeline_only", "pipeline_extracted": "pipeline_only"}
ALL = ["wy", "mt", "wa", "or", "nv", "id", "ut", "az", "co"]


def load_matched() -> dict:
    matched = {}
    for p in sorted(Path("output/corroboration").glob("*_corroboration.csv")):
        for r in csv.DictReader(open(p, encoding="utf-8")):
            if r["match_strength"] == "none":
                continue
            k = (r["state"], r["our_serff_tracking_number"], r["our_subsidiary"])
            if k not in matched or RANK[r["match_strength"]] > RANK[matched[k]]:
                matched[k] = r["match_strength"]
    return matched


def eff_year(v) -> int:
    e = str(v or "").replace("-", "/")
    for t in e.split("/"):
        if len(t) == 4 and t.isdigit():
            return int(t)
    p = e.split("/")
    return int("20" + p[2]) if len(p) == 3 and len(p[2]) == 2 else 0


def classify(matched, state, tk, company, eff):
    if tk == ANCHOR:
        return "field_validated", "field"
    k = (state, tk, company)
    if k in matched:
        return "ambest_cross_checked", matched[k]
    if eff_year(eff) >= 2025 and state in CROSS_CHECKED_STATES:
        return "pipeline_extracted_in_validated_window", ""
    return "pipeline_extracted", ""


def main(states) -> int:
    matched = load_matched()
    grand = Counter()
    for st in states:
        p = Path(f"output/{st.lower()}_final_rates.xlsx")
        if not p.exists():
            print(f"{st}: missing {p}"); continue
        wb = load_workbook(p); ws = wb.active
        hdr = [c.value for c in ws[1]]; idx = {h: i for i, h in enumerate(hdr)}
        ms_col = (idx["match_strength"] + 1) if "match_strength" in idx else len(hdr) + 1
        if "match_strength" not in idx:
            ws.cell(row=1, column=ms_col, value="match_strength")
        ev, vt = idx["external_validation"] + 1, idx["validation_tier"] + 1
        tkc, cnc, stc, efc = (idx["serff_tracking_number"] + 1, idx["company_name"] + 1,
                              idx["state"] + 1, idx["effective_date"] + 1)
        c = Counter()
        for r in range(2, ws.max_row + 1):
            tk = ws.cell(row=r, column=tkc).value
            if tk is None:
                continue
            ext, ms = classify(matched, ws.cell(row=r, column=stc).value, tk,
                               ws.cell(row=r, column=cnc).value, ws.cell(row=r, column=efc).value)
            ws.cell(row=r, column=ev, value=ext)
            ws.cell(row=r, column=ms_col, value=ms)
            ws.cell(row=r, column=vt, value=VT[ext])
            c[ext] += 1
        wb.save(p)
        grand.update(c)
        print(f"{st}: {dict(c)}")
    print(f"TOTAL: {dict(grand)} | sum {sum(grand.values())}")
    return 0


if __name__ == "__main__":
    args = [a.lower() for a in sys.argv[1:]] or ALL
    raise SystemExit(main(args))
