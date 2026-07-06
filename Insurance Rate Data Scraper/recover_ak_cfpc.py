"""Targeted recovery of the 2 material COUNTRY Country-Financial CFPC filings that
the wrapped-company-name parser bug zeroed in AK (the VT PRGS-134029613 family,
2nd occurrence — this time MATERIAL, so RECOVER per the WV Liberty / NH General
Insurance precedent; not drop like VT's immaterial 0%/0ph case).

  - CFPC-134283900: Authorized, eff 07/06/2025 (ACTIVE window), +9.591% / 9,644 ph
    -> Country Personal Auto Prospect signal on the AM-Best-undercounted brand.
  - CFPC-133947234: Approved, eff 08/28/2024 (pre-2025 back-extension), +11.414%
    / 4,486 ph.

Each co-files 3 COUNTRY entities (Mutual / Preferred / Casualty) -> 6 entity rows.
This RE-PARSES the two PDFs' Company Rate Information tables with explicit
wrapped-name handling (the data line carries the numbers; the next line is the
wrapped company-name continuation), then ASSERTS the parsed per-entity impacts
against the adjudication before appending. AK-ONLY: touches only
output/ak_final_rates.xlsx (94 -> 100 rows). No other state's deliverable, no
filings.db. The shared-parser fix is BACKLOG (deliberate, needs full
re-verification since it re-emits VT's 2 immaterial rows).
"""
from __future__ import annotations
import re
from pathlib import Path

import openpyxl
import pdfplumber

DELIVERABLE = Path("output/ak_final_rates.xlsx")

# Per-filing metadata (from the Disposition/Filing-at-a-Glance pages) + the
# adjudication cross-check values (per-entity Overall % Rate Impact).
FILINGS = {
    "CFPC-134283900": {
        "filing_id": "134283900", "effective_date": "07/06/2025",
        "disposition_status": "Authorized",
        "expect_impact": {"COUNTRY Mutual Insurance Company": "9.618%",
                          "COUNTRY Preferred Insurance Company": "9.494%",
                          "COUNTRY Casualty Insurance Company": "10.177%"},
    },
    "CFPC-133947234": {
        "filing_id": "133947234", "effective_date": "08/28/2024",
        "disposition_status": "Approved",
        "expect_impact": {"COUNTRY Mutual Insurance Company": "11.014%",
                          "COUNTRY Preferred Insurance Company": "9.105%",
                          "COUNTRY Casualty Insurance Company": "23.733%"},
    },
}

LINE = "19.0 Personal Auto"
SUBTOI = "19.0000 Personal Auto Combinations"

# A Company Rate Information data line: name-prefix then 8 numeric-ish columns.
# The company name's 2nd word ("Mutual"/"Preferred"/"Casualty") is on this line;
# "Insurance Company" wraps to the NEXT line (the bug the main parser trips on).
_DATA = re.compile(
    r"^(COUNTRY \w+)\s+"
    r"([\d.,]+%)\s+([\d.,]+%)\s+\$?([\d,]+)\s+([\d,]+)\s+\$?([\d,]+)\s+"
    r"([\d.,]+%)\s+(-?[\d.,]+%)\s*$")


def _num(s: str) -> str:
    return s.replace(",", "").replace("$", "")


def parse_rate_table(fid: str) -> list[dict]:
    with pdfplumber.open(f"output/pdfs/AK/{fid}/filing_summary.pdf") as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    # locate the FULL 8-column Company Rate Information table (has the header row)
    hdr = "Number of Policy"
    i = text.find("Company Rate Information")
    while i >= 0 and hdr not in text[i:i + 400]:
        i = text.find("Company Rate Information", i + 1)
    if i < 0:
        raise SystemExit(f"{fid}: full Company Rate Information table not found")
    lines = text[i:i + 1600].splitlines()
    out = []
    for n, ln in enumerate(lines):
        m = _DATA.match(ln.strip())
        if not m:
            continue
        cont = lines[n + 1].strip() if n + 1 < len(lines) else ""
        # wrapped-name handling: join the data-line name prefix with its continuation
        name = m.group(1)
        if cont.startswith("Insurance Company"):
            name = f"{name} Insurance Company"
        # Strip thousands-commas from % fields — the scraper/import format is
        # comma-free ('2162.300%' not '2,162.300%'); parse_percent does float()
        # after rstrip('%') and chokes on a comma.
        _pct = lambda s: s.replace(",", "")
        out.append({
            "company_name": name,
            "overall_indicated_change": _pct(m.group(2)),
            "overall_rate_impact": _pct(m.group(3)),
            "written_premium_change": _num(m.group(4)),
            "policyholders_affected": int(_num(m.group(5))),
            "written_premium_for_program": _num(m.group(6)),
            "maximum_percent_change": _pct(m.group(7)),
            "minimum_percent_change": _pct(m.group(8)),
        })
    return out


def build_rows() -> list[dict]:
    rows = []
    for tk, meta in FILINGS.items():
        parsed = parse_rate_table(meta["filing_id"])
        assert len(parsed) == 3, f"{tk}: expected 3 entities, got {len(parsed)}"
        eff_year = meta["effective_date"][-4:]
        # source/tier by eff-year (matches the AK deliverable's existing rows)
        if eff_year == "2024":
            source, ext = "extension", "pipeline_extracted"
        else:
            source, ext = "original", "pipeline_extracted_in_validated_window"
        for e in parsed:
            # cross-check the parsed impact against the adjudication
            exp = meta["expect_impact"][e["company_name"]]
            assert e["overall_rate_impact"] == exp, \
                f"{tk} {e['company_name']}: parsed {e['overall_rate_impact']} != adjudicated {exp}"
            rows.append({
                "state": "AK", "effective_date": meta["effective_date"],
                "company_name": e["company_name"], "line_of_business": LINE,
                "sub_type_of_insurance": SUBTOI,
                "overall_indicated_change": e["overall_indicated_change"],
                "overall_rate_impact": e["overall_rate_impact"],
                "written_premium_change": e["written_premium_change"],
                "policyholders_affected": e["policyholders_affected"],
                "written_premium_for_program": e["written_premium_for_program"],
                "maximum_percent_change": e["maximum_percent_change"],
                "minimum_percent_change": e["minimum_percent_change"],
                "rate_activity": "rate_change",
                "serff_tracking_number": tk,
                "disposition_status": meta["disposition_status"],
                "filing_date": None,
                "source_pdf": f"output/pdfs/AK/{meta['filing_id']}/filing_summary.pdf",
                "validation_tier": "pipeline_only", "source": source,
                "external_validation": ext, "match_strength": None,
            })
    return rows


def main() -> int:
    wb = openpyxl.load_workbook(DELIVERABLE)
    ws = wb["rates"]
    header = [c.value for c in ws[1]]
    before = ws.max_row - 1

    existing = {(ws.cell(r, header.index("serff_tracking_number") + 1).value,
                 ws.cell(r, header.index("company_name") + 1).value)
                for r in range(2, ws.max_row + 1)}

    rows = build_rows()
    added = 0
    for row in rows:
        key = (row["serff_tracking_number"], row["company_name"])
        if key in existing:
            print(f"  [skip] already present: {key}")
            continue
        ws.append([row.get(h) for h in header])
        added += 1
        print(f"  [add] {row['serff_tracking_number']:16s} {row['company_name']:38s} "
              f"eff={row['effective_date']} imp={row['overall_rate_impact']} "
              f"ph={row['policyholders_affected']} disp={row['disposition_status']} src={row['source']}")

    wb.save(DELIVERABLE)
    after = ws.max_row - 1
    print(f"\n[recover] AK deliverable: {before} -> {after} rows (+{added})")
    tot_ph = sum(r["policyholders_affected"] for r in rows)
    print(f"[recover] 6 recovered rows total policyholders: {tot_ph} "
          f"(134283900=9,644 active + 133947234=4,486 back-extension)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
