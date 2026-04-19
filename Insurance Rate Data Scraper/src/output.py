"""Excel output for scraped filings.

Two sheets:
  - "Filings":           one row per filing, all scraped fields
  - "Unparseable PDFs":  manual review queue for filings where the rate-effect
                         parser returned nothing despite PDFs being attached.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Filing

FILINGS_COLUMNS = [
    "state",
    "serff_tracking_number",
    "filing_id",
    "company_name",
    "target_company",
    "naic_codes",
    "product_name",
    "type_of_insurance",
    "sub_type_of_insurance",
    "filing_type",
    "filing_status",
    "submission_date",
    "disposition_date",
    "disposition_status",
    "state_status",
    "requested_rate_effect",
    "approved_rate_effect",
    "overall_rate_effect",
    "affected_policyholders",
    "written_premium_volume",
    "annual_premium_impact_dollars",
    "current_avg_premium",
    "proposed_avg_premium",
    "premium_change_dollars",
    "program_name",
    "filing_reason",
    "prior_approval",
    "pdfs",
    "pdf_parse_status",
    "pdf_parse_fields_found",
    "detail_url",
]

UNPARSEABLE_COLUMNS = [
    "state",
    "serff_tracking_number",
    "filing_id",
    "company_name",
    "filing_type",
    "submission_date",
    "pdf_parse_status",
    "pdf_count",
    "pdf_names",
    "detail_url",
]


def _is_unparseable(f: Filing) -> bool:
    """A filing needs manual review if PDFs were attached but no rate field extracted.

    `overall_rate_effect == 0.0` (premium-neutral sentinel) counts as parsed.
    """
    has_pdfs = bool(f.pdfs)
    no_rate_extracted = (
        f.overall_rate_effect is None
        and f.requested_rate_effect is None
        and f.approved_rate_effect is None
    )
    return has_pdfs and no_rate_extracted


def write_excel(filings: Iterable[Filing], output_path: Path) -> Path:
    """Write filings to xlsx. Returns the output path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_filings = wb.active
    ws_filings.title = "Filings"
    _write_header(ws_filings, FILINGS_COLUMNS)

    ws_unparseable = wb.create_sheet("Unparseable PDFs")
    _write_header(ws_unparseable, UNPARSEABLE_COLUMNS)

    filings_list = list(filings)
    for f in filings_list:
        row = f.to_row()
        ws_filings.append([row.get(c) for c in FILINGS_COLUMNS])

        if _is_unparseable(f):
            ws_unparseable.append([
                f.state,
                f.serff_tracking_number,
                f.filing_id,
                f.company_name,
                f.filing_type,
                f.submission_date.isoformat() if f.submission_date else None,
                f.pdf_parse_status,
                len(f.pdfs),
                ";".join(p.display_name for p in f.pdfs),
                f.detail_url,
            ])

    _autosize(ws_filings, FILINGS_COLUMNS)
    _autosize(ws_unparseable, UNPARSEABLE_COLUMNS)
    ws_filings.freeze_panes = "A2"
    ws_unparseable.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill


def _autosize(ws, columns: list[str]) -> None:
    for idx, col in enumerate(columns, start=1):
        max_len = len(col)
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)
