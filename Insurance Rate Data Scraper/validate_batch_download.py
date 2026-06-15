"""Step-3 in-vivo validation for batched downloads — ONE batch, ONE fresh search.

    python validate_batch_download.py GA --plan   # offline: show the selected batch
    python validate_batch_download.py GA          # LIVE: run it (USER APPROVAL REQUIRED
                                                  #   — only after the SERFF rest period)

Validates the batched download_all_pdfs path (run_final_rates.DOWNLOAD_BATCH_SIZE)
before any full batched state run, per the 2026-06-10 rest/batch/resume plan:

  - ONE fresh "Begin Search" session, then 8 sequential system-summary
    downloads in that session: 3 ALREADY-CACHED targets re-fetched into a
    scratch dir (never touching the known-good cache) interleaved with 5
    genuinely uncached targets written to the real cache (real GA progress).
  - Per-position success table -> detects JSF session degradation (the OR
    Travelers signature: misses starting partway through a reused session).
  - Cell-level parse comparison for the 3 re-fetches vs their known-good
    cached copies: filing type, new-product flag, rate_data_applies,
    disposition status, effective dates, and every per-company rate row
    (7 numeric fields each). Byte equality is NOT required (SERFF stamps
    generation time); parse equality is what the deliverable consumes.

Verdict CLEAN = 8/8 landed + 0 parse diffs -> proceed with the batched GA
resume (run_final_rates.py GA). Misses mid-batch or any parse diff -> fall
back to fresh-per-target (DOWNLOAD_BATCH_SIZE=1 or the built-in fallback pass)
and report before proceeding.

All 8 candidates are drawn from ONE (group, search term) pool, matched by the
search workbook's target_company column — the keyword that actually found the
filing — guaranteeing every row is findable in the single session's result set
(subsets of a group surface only under their own term, e.g. Encompass filings
never appear in an "allstate" search). If the pool has fewer than 5 uncached
targets, the batch tops up with extra cached re-fetch probes, keeping 8
downloads per session (the thing under test).
"""
from __future__ import annotations

import shutil
import sys
import time
from pathlib import Path

from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

import run_final_rates as rf
import src.search as _serff_search
from src.config import HEADLESS, REQUEST_DELAY, USER_AGENT
from src.detail import download_system_summary_pdf
from src.search import _set_rows_per_page_100, _submit_search
from src.utils import parse_filing_summary_pdf

DIAG = Path("output/serff_diagnostics")

BATCH_TOTAL = rf.DOWNLOAD_BATCH_SIZE  # 8 — the session length under test
N_CACHED_REFETCH_MIN = 3
N_UNCACHED_MAX = 5
SCRATCH = Path("output/batch_validation")


def cache_pdf(state: str, fid: str) -> Path:
    return Path(f"output/pdfs/{state}/{fid}/filing_summary.pdf")


def _target_company_by_fid(state: str) -> dict[str, str]:
    """filing_id -> target_company (the search keyword bucket that found it)
    from the state's search universe workbooks."""
    import openpyxl
    out: dict[str, str] = {}
    for src in rf._search_universe_paths(state):
        wb = openpyxl.load_workbook(src, read_only=True)
        ws = wb["Filings"] if "Filings" in wb.sheetnames else wb.active
        hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
        ix = {h: i for i, h in enumerate(hdr)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            fid = str(r[ix["filing_id"]] or "")
            if fid and fid not in out:
                out[fid] = r[ix["target_company"]] or ""
        wb.close()
    return out


def pick_batch(state: str, targets: list[rf.Target]) -> tuple[str, str, list[tuple[rf.Target, bool]]]:
    """Returns (group, search_term, [(target, is_cached_refetch)] ordered).
    Pools are keyed by (group, term) where term = target_company.lower() — the
    keyword whose search actually returned the filing, so every batch member
    is guaranteed findable in the one session. Prefers the pool with the most
    uncached targets (real progress); requires >= N_CACHED_REFETCH_MIN cached
    for the comparison probes; tops up with extra probes to BATCH_TOTAL."""
    tc = _target_company_by_fid(state)
    pools: dict[tuple[str, str], dict[str, list[rf.Target]]] = {}
    for t in targets:
        term = (tc.get(t.filing_id) or "").lower()
        if term not in rf.GROUP_SEARCH.get(t.group, []):
            continue
        p = cache_pdf(state, t.filing_id)
        bucket = "cached" if (p.exists() and p.stat().st_size > 5000) else "uncached"
        pools.setdefault((t.group, term), {"cached": [], "uncached": []})[bucket].append(t)
    candidates = [
        (g, term, d) for (g, term), d in pools.items()
        if len(d["cached"]) >= N_CACHED_REFETCH_MIN
        and len(d["cached"]) + len(d["uncached"]) >= BATCH_TOTAL
    ]
    if not candidates:
        raise SystemExit("no (group, term) pool has enough cached probes + total candidates")
    grp, term, d = max(candidates, key=lambda c: len(c[2]["uncached"]))
    uncached = d["uncached"][:N_UNCACHED_MAX]
    cached = d["cached"][:BATCH_TOTAL - len(uncached)]
    # Order so re-fetch probes sit at the START, spaced through the MIDDLE,
    # and at the END of the session — positional probes for degradation.
    middle_probes = cached[1:-1]
    middle: list[tuple[rf.Target, bool]] = []
    ui, ci = 0, 0
    while ui < len(uncached) or ci < len(middle_probes):
        if ui < len(uncached):
            middle.append((uncached[ui], False)); ui += 1
        if ci < len(middle_probes):
            middle.append((middle_probes[ci], True)); ci += 1
    return grp, term, [(cached[0], True)] + middle + [(cached[-1], True)]


def compare_parse(state: str, t: rf.Target, refetched: Path) -> list[str]:
    """Field-level diff of (known-good cached) vs (re-fetched) parse results."""
    diffs: list[str] = []
    good = cache_pdf(state, t.filing_id)
    sides = {}
    for label, path in (("cached", good), ("refetch", refetched)):
        text = rf._read_pdf_text(path)
        ft, is_new = rf.detect_filing_type_and_new_product(path, text=text)
        fs = parse_filing_summary_pdf(path, t.tracking, text=text)
        sides[label] = (ft, is_new, fs)
    (ft_a, new_a, fs_a), (ft_b, new_b, fs_b) = sides["cached"], sides["refetch"]
    for name, va, vb in (
        ("filing_type", ft_a, ft_b),
        ("is_new_product", new_a, new_b),
        ("rate_data_applies", fs_a.rate_data_applies, fs_b.rate_data_applies),
        ("disposition_status", fs_a.disposition_status, fs_b.disposition_status),
        ("effective_date_new", fs_a.effective_date_new, fs_b.effective_date_new),
        ("effective_date_renewal", fs_a.effective_date_renewal, fs_b.effective_date_renewal),
        ("n_company_rates", len(fs_a.company_rates), len(fs_b.company_rates)),
    ):
        if va != vb:
            diffs.append(f"{name}: cached={va!r} refetch={vb!r}")
    for ra, rb in zip(fs_a.company_rates, fs_b.company_rates):
        for f in ("company_name", "overall_indicated_change", "overall_rate_impact",
                  "written_premium_change", "policyholders_affected",
                  "written_premium_for_program", "maximum_pct_change", "minimum_pct_change"):
            va, vb = getattr(ra, f), getattr(rb, f)
            if va != vb:
                diffs.append(f"company_rates[{ra.company_name}].{f}: cached={va!r} refetch={vb!r}")
    return diffs


def main() -> int:
    args = sys.argv[1:]
    state = next((a for a in args if not a.startswith("--")), "GA").upper()
    plan_only = "--plan" in args

    targets, rep = rf.load_targets_search(state)
    n_cached = sum(1 for t in targets if cache_pdf(state, t.filing_id).exists()
                   and cache_pdf(state, t.filing_id).stat().st_size > 5000)
    print(f"=== {state} batched-download validation ===")
    print(f"targets: {len(targets)} | cached: {n_cached} | uncached: {len(targets) - n_cached}")

    grp, term, batch = pick_batch(state, targets)
    print(f"\nselected group={grp!r} search_term={term!r} — ONE session, {len(batch)} downloads:")
    for pos, (t, is_refetch) in enumerate(batch, 1):
        kind = "REFETCH (vs known-good cache)" if is_refetch else "uncached (real progress)"
        print(f"  [{pos}] {t.tracking} ({t.company}) — {kind}")
    if plan_only:
        print("\n--plan: no SERFF traffic. Run without --plan after the rest period (user approval).")
        return 0

    from src.quiet_period import guard
    guard("validate_batch_download")  # refuse during a declared SERFF rest window (live path only)

    # TRIPLE DUTY: (1) batch validation, (2) capacity probe, (3) failure-
    # signature capture. DIAG_DIR makes every fresh search write a timestamped
    # ledger row (output/serff_diagnostics/search_ledger.csv) and, on failure,
    # a snapshot dir with HTTP statuses, headers (Retry-After), page title/
    # body/HTML — the first-ever direct measurement of what SERFF serves when
    # it walls (everything before this was inferred from bare click timeouts).
    _serff_search.DIAG_DIR = DIAG
    print(f"\ndiagnostics: ledger + failure snapshots -> {DIAG}/")

    results: list[tuple[int, str, bool, bool]] = []  # (pos, tracking, is_refetch, ok)
    refetch_paths: dict[str, Path] = {}
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        ctx = browser.new_context(user_agent=USER_AGENT, accept_downloads=True)
        page = ctx.new_page()
        print(f"\n[session] ONE fresh search: {term!r}")
        try:
            search_ok = _submit_search(page, state, term)
        except Exception as e:
            print(f"  search submit RAISED: {type(e).__name__}: {e}")
            search_ok = False
        if not search_ok:
            print("  search submit FAILED — throttle not rested? aborting (0 downloads attempted)")
            snaps = sorted(DIAG.glob("fail_*"))
            if snaps:
                print(f"  FAILURE SIGNATURE CAPTURED: {snaps[-1]}/snapshot.json (+page.html)")
                print(f"  ledger: {DIAG / 'search_ledger.csv'}")
            browser.close()
            return 1
        _set_rows_per_page_100(page)
        for pos, (t, is_refetch) in enumerate(batch, 1):
            if is_refetch:
                dest = SCRATCH / state / t.filing_id
                shutil.rmtree(dest, ignore_errors=True)  # force a real re-download
            else:
                dest = Path(f"output/pdfs/{state}/{t.filing_id}")
            try:
                pdf = download_system_summary_pdf(page, t.filing_id, t.tracking, dest)
            except Exception as e:
                print(f"  [{pos}/{len(batch)}] {t.tracking}: error {type(e).__name__}: {e}", flush=True)
                pdf = None
            ok = pdf is not None
            results.append((pos, t.tracking, is_refetch, ok))
            if ok and is_refetch:
                refetch_paths[t.filing_id] = pdf
            print(f"  [{pos}/{len(batch)}] {t.tracking}: {'ok' if ok else 'MISS'}", flush=True)
            time.sleep(REQUEST_DELAY)
        browser.close()

    n_ok = sum(1 for *_x, ok in results if ok)
    print(f"\n=== POSITION TABLE (degradation check) ===")
    for pos, tracking, is_refetch, ok in results:
        print(f"  pos {pos}: {'ok  ' if ok else 'MISS'} {tracking}{' [refetch]' if is_refetch else ''}")
    first_miss = next((pos for pos, _t, _r, ok in results if not ok), None)
    degradation = first_miss is not None and any(ok for pos, _t, _r, ok in results if pos < first_miss)

    print(f"\n=== PARSE COMPARISON (refetch vs known-good cache) ===")
    total_diffs = 0
    for t, is_refetch in batch:
        if not is_refetch:
            continue
        rp = refetch_paths.get(t.filing_id)
        if rp is None:
            print(f"  {t.tracking}: refetch MISSED — no comparison possible")
            total_diffs += 1
            continue
        diffs = compare_parse(state, t, rp)
        if diffs:
            total_diffs += len(diffs)
            for d in diffs:
                print(f"  DIFF {t.tracking} {d}")
        else:
            print(f"  {t.tracking}: parse identical (all fields + all company rate rows)")

    clean = n_ok == len(batch) and total_diffs == 0
    print(f"\n=== VERDICT ===")
    print(f"  downloads: {n_ok}/{len(batch)} | parse diffs: {total_diffs} | "
          f"mid-session degradation: {'YES' if degradation else 'no'}")
    print(f"  {'CLEAN — proceed with batched resume (run_final_rates.py ' + state + ')' if clean else 'NOT CLEAN — use fresh-per-target (DOWNLOAD_BATCH_SIZE=1) and report before proceeding'}")
    return 0 if clean else 1


if __name__ == "__main__":
    raise SystemExit(main())
