"""Rebuild rate_changes.xlsx → rate_changes_v3.xlsx.

Adds 3 new columns (overall_indicated_change, overall_rate_impact,
policyholders_affected) and expands per-subsidiary breakdowns into
multiple rows where extraction found per-sub data.

Inputs:  output/rate_changes.xlsx  (existing 25-row deliverable)
         output/subsidiary_fields.json  (extraction output)
Output:  output/rate_changes_v3.xlsx
"""
from __future__ import annotations

import json
import sys
from copy import deepcopy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.stdout.reconfigure(encoding="utf-8")

from src.config import OUTPUT_DIR

SRC = OUTPUT_DIR / "rate_changes.xlsx"
SUBS = OUTPUT_DIR / "subsidiary_fields.json"
DST = OUTPUT_DIR / "rate_changes_v3.xlsx"

# Carrier classification — based on probe findings
FORM_A_FILERS = {"State Farm", "Allstate"}  # consistently include Form A or filing memo with extractable values
MINIMUM_FILERS = {"GEICO", "Progressive", "Travelers", "Liberty", "Liberty Mutual"}

NEW_COLS = ["overall_indicated_change", "overall_rate_impact", "policyholders_affected"]


def load_existing() -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Rate Changes"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    out = [dict(zip(headers, r)) for r in rows[1:]]
    return headers, out


def load_subs() -> dict[str, dict]:
    return json.loads(SUBS.read_text(encoding="utf-8"))


def fmt_pct(v) -> str | None:
    if v is None:
        return None
    return f"{v:+.3f}%"


def carrier_status(carrier: str) -> str:
    if carrier in MINIMUM_FILERS:
        return "minimum_filer"
    if carrier in FORM_A_FILERS:
        return "form_a_filer"
    return "unknown"


def build_rows(headers: list[str], existing: list[dict], subs_data: dict) -> list[dict]:
    """Build the new row list. Each output row is a dict keyed by column name.

    Strategy:
      - Look up extraction by serff number.
      - Subs count = 0: keep row as-is, append note about no Form A in PDFs.
      - Subs count = 1: enrich row in place with the 3 new fields.
      - Subs count >= 2: per-subsidiary expansion. If the existing deliverable
        already has multiple rows for this serff (SFMA-134522369 split case),
        match by company_name; else replicate the row per-subsidiary, set
        company_name to the subsidiary, and update rate_effect_value to the
        per-sub impact when available.
    """
    # Group existing rows by serff
    by_serff: dict[str, list[dict]] = {}
    for r in existing:
        by_serff.setdefault(r["serff_tracking_number"], []).append(r)

    out: list[dict] = []

    for serff, rows in by_serff.items():
        sub_record = subs_data.get(serff, {})
        subs = sub_record.get("subsidiaries", []) or []

        # Case 1: no per-sub extraction — preserve all existing rows, annotate
        if not subs:
            for r in rows:
                new = deepcopy(r)
                for c in NEW_COLS:
                    new[c] = None
                carrier = (new.get("carrier") or "").strip()
                status = carrier_status(carrier)
                if status == "minimum_filer":
                    note = ("No Form A in public PDFs — "
                            f"{carrier} files minimum-filer format (filing memo + manual pages). "
                            "Per-subsidiary impact / indicated change / policyholder counts not "
                            "publicly extractable; would require licensed NAIC statutory data.")
                elif carrier == "Liberty":
                    note = ("No Form A in public PDFs — Liberty Mutual filing did not include "
                            "extractable rate impact / policyholder fields.")
                else:  # Allstate or other Form A filers that still produced no subs
                    note = ("Form A not located in PDFs for this filing (rule/form revision); "
                            "no per-subsidiary impact / indicated change / policyholder counts "
                            "extractable.")
                _append_note(new, note)
                out.append(new)
            continue

        # Case 2: single subsidiary — enrich the existing row(s) in place
        if len(subs) == 1:
            sub = subs[0]
            for r in rows:
                new = deepcopy(r)
                new["overall_indicated_change"] = fmt_pct(sub.get("overall_indicated_change"))
                new["overall_rate_impact"] = fmt_pct(sub.get("overall_rate_impact"))
                new["policyholders_affected"] = sub.get("policyholders_affected")
                # Special-case: SFMA-134393639 had blank Section 12 in PDF
                if serff == "SFMA-134393639":
                    _append_note(new, ("Form A Section 10 populated (545,361 policyholders SFFC) "
                                       "but Section 12 (overall % rate impact) was blank in the "
                                       "PDF; impact field unextractable from this filing."))
                # Special-case: ALSE-134572006 0.0% confirmed
                if serff == "ALSE-134572006":
                    _append_note(new, ("AVPIC 0.0% overall rate impact is correct: filing memo "
                                       "states the Multiple Policy Discount – Auto revision was "
                                       "'calibrated to target an overall rate level change of 0.0%' "
                                       "(structural revenue-neutral revision)."))
                out.append(new)
            continue

        # Case 3: multi-subsidiary expansion
        # If existing deliverable has 1 "Multiple" row, expand into N per-sub rows.
        if len(rows) == 1 and (rows[0].get("company_name") or "").strip().lower() == "multiple":
            base = rows[0]
            for sub in subs:
                new = deepcopy(base)
                new["company_name"] = sub.get("company")
                impact = sub.get("overall_rate_impact")
                ind = sub.get("overall_indicated_change")
                pol = sub.get("policyholders_affected")
                new["overall_indicated_change"] = fmt_pct(ind)
                new["overall_rate_impact"] = fmt_pct(impact)
                new["policyholders_affected"] = pol
                # Update rate_effect_value with per-sub impact when present
                if impact is not None:
                    if base.get("rate_effect_value") is not None and base.get("original_value") is None:
                        new["original_value"] = base.get("rate_effect_value")
                    new["rate_effect_value"] = fmt_pct(impact).replace(".000%", ".00%")
                    if (new.get("rate_change_type") or "").lower() != "overall_impact":
                        if new.get("original_value") is None:
                            new["original_value"] = base.get("rate_effect_value")
                        new["rate_change_type"] = "overall_impact"
                    new["rate_effect_source"] = "pdf_form_a"
                _append_note(new, (f"Per-subsidiary split from {serff} (Multiple). "
                                   f"Values from PDF Form A Section 12 / Section 10 for "
                                   f"{sub.get('company')}."))
                out.append(new)
        else:
            # Existing has multi-row split (e.g., SFMA-134522369). Match by company name.
            for r in rows:
                new = deepcopy(r)
                # Find subsidiary matching this company_name (substring tolerant)
                row_co = (r.get("company_name") or "").strip().lower()
                match = None
                for sub in subs:
                    sc = (sub.get("company") or "").strip().lower()
                    if sc and (sc in row_co or row_co in sc):
                        match = sub
                        break
                if match is None and len(subs) == 1:
                    match = subs[0]
                if match:
                    new["overall_indicated_change"] = fmt_pct(match.get("overall_indicated_change"))
                    new["overall_rate_impact"] = fmt_pct(match.get("overall_rate_impact"))
                    new["policyholders_affected"] = match.get("policyholders_affected")
                else:
                    for c in NEW_COLS:
                        new[c] = None
                out.append(new)

    return out


def _append_note(row: dict, note: str) -> None:
    existing = (row.get("correction_note") or "").strip()
    if existing:
        if note in existing:
            return
        row["correction_note"] = f"{existing} | {note}"
    else:
        row["correction_note"] = note


def write_workbook(headers: list[str], rows: list[dict]) -> None:
    # Column order: insert NEW_COLS after rate_change_type
    out_headers = []
    inserted = False
    for h in headers:
        out_headers.append(h)
        if h == "rate_change_type" and not inserted:
            out_headers.extend(NEW_COLS)
            inserted = True
    if not inserted:
        out_headers.extend(NEW_COLS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rate Changes"

    # Header
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")
    new_col_fill = PatternFill("solid", fgColor="548235")
    for col_idx, h in enumerate(out_headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = bold
        c.fill = new_col_fill if h in NEW_COLS else fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(out_headers, 1):
            v = row.get(h)
            ws.cell(row=r_idx, column=c_idx, value=v)

    # Column widths
    widths = {
        "state": 6, "carrier": 11, "company_name": 32, "line_of_business": 38,
        "toi_code": 10, "sub_toi_code": 12, "filing_component": 16,
        "effective_date": 13, "effective_date_source": 16, "rate_effect_value": 14,
        "rate_effect_source": 16, "rate_change_type": 16,
        "overall_indicated_change": 18, "overall_rate_impact": 18,
        "policyholders_affected": 16, "original_value": 12,
        "correction_note": 60, "current_avg_premium": 14, "new_avg_premium": 14,
        "serff_tracking_number": 22, "filing_date": 12, "disposition_status": 14,
    }
    for col_idx, h in enumerate(out_headers, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(h, 14)
    ws.freeze_panes = "A2"

    # Legend & Notes
    legend = wb.create_sheet("Legend & Notes")
    legend_rows = _build_legend()
    bold_no_color = Font(bold=True)
    for r_idx, (a, b) in enumerate(legend_rows, 1):
        legend.cell(row=r_idx, column=1, value=a).font = bold_no_color if a and not b else Font()
        legend.cell(row=r_idx, column=2, value=b)
        legend.cell(row=r_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    legend.column_dimensions["A"].width = 32
    legend.column_dimensions["B"].width = 110

    # Manual review (effective_date) — copy through unchanged
    src_wb = openpyxl.load_workbook(SRC, data_only=True)
    src_review = src_wb["Manual Review (effective_date)"]
    review = wb.create_sheet("Manual Review (effective_date)")
    for r_idx, row in enumerate(src_review.iter_rows(values_only=True), 1):
        for c_idx, v in enumerate(row, 1):
            review.cell(row=r_idx, column=c_idx, value=v)
    review.freeze_panes = "A2"

    # Data Availability summary
    da = wb.create_sheet("Data Availability")
    _write_data_availability(da, rows)

    wb.save(DST)


def _build_legend() -> list[tuple[str | None, str | None]]:
    base = [
        ("Column", "Meaning"),
        ("state", "US state of the filing (WA / ID / CO)."),
        ("carrier", "Target carrier searched (Allstate, Geico, Liberty, Progressive, State Farm, Travelers)."),
        ("company_name", "Specific underwriting company that filed (e.g. Allstate Fire & Casualty). "
                         "For per-subsidiary expanded rows, this names the specific subsidiary "
                         "(e.g. State Farm Mutual Automobile Insurance Company vs State Farm Fire and Casualty Company)."),
        ("line_of_business", "Type Of Insurance — Sub Type Of Insurance string from SERFF."),
        ("toi_code", "NAIC Type Of Insurance code (19.0 Personal Auto, 04.0 Homeowners)."),
        ("sub_toi_code", "NAIC Sub-TOI code."),
        ("filing_component", "Per-form Homeowners component (Non-Tenant / Renters / Condominium)."),
        ("effective_date", "When the rate change takes / took effect."),
        ("effective_date_source", "How the effective_date was derived."),
        ("rate_effect_value", "Signed percentage change in rates. For per-subsidiary expanded rows, "
                              "this is the per-sub overall_rate_impact when extracted from Form A."),
        ("rate_effect_source", "Origin of rate_effect_value: serff field name, pdf_memo, or pdf_form_a."),
        ("rate_change_type", "overall_impact / indicated / ambiguous classification."),
        ("overall_indicated_change", "NEW. Actuarially indicated rate change (%). Sourced from PDF Form A "
                                     "tables or filing memo prose. Distinct from overall_rate_impact — the "
                                     "indicated value is what the actuarial analysis recommends; the impact "
                                     "is what the carrier proposes to file. Often the carrier files less "
                                     "than the indicated value."),
        ("overall_rate_impact", "NEW. Actual proposed weighted overall rate level change (%). Sourced from "
                                "PDF Form A Section 12 ('OVERALL % RATE IMPACT/CHANGE'), tables, or filing "
                                "memo prose ('overall rate level change of X%'). This is the headline "
                                "rate change number for the filing."),
        ("policyholders_affected", "NEW. Number of policyholders affected by this rate change. Sourced from "
                                   "PDF Form A Section 10 ('NUMBER OF POLICYHOLDERS AFFECTED FOR THIS PROGRAM'). "
                                   "Available primarily for State Farm filings; not publicly disclosed in "
                                   "GEICO/Progressive/Travelers minimum-filer PDFs."),
        ("original_value", "Pre-correction SERFF value when changed by PDF cross-check or per-sub split."),
        ("correction_note", "Explains corrections, splits, and data-availability status."),
        ("current_avg_premium", "Average annual premium BEFORE the change."),
        ("new_avg_premium", "Average annual premium AFTER the change."),
        ("serff_tracking_number", "Unique SERFF identifier."),
        ("filing_date", "Submission date the carrier filed with the state regulator."),
        ("disposition_status", "Regulator decision (Approved / Pending / etc)."),
        (None, None),
        ("METHODOLOGY", None),
        ("Line-of-business filter", "Strict NAIC Uniform P&C Product Coding Matrix code matching. "
                                    "Personal Auto (19.x) + Homeowners (04.x) sub-TOIs only."),
        ("Rate-effect classification", "Each rate_effect_value cross-checked against PDF using labeled "
                                       "phrase patterns to distinguish proposed-overall vs indicated "
                                       "rate change."),
        ("Form splits", "Where one SERFF filing covers multiple Homeowners forms with different rate "
                        "impacts, the row was split per form."),
        ("Per-subsidiary expansion (NEW)", "Where one SERFF filing covers multiple underwriting subsidiaries "
                                           "(e.g. State Farm Mutual Auto AND State Farm Fire & Casualty), "
                                           "the original 'Multiple' row was expanded into one row per "
                                           "subsidiary using PDF Form A extraction. Each expanded row "
                                           "carries its own overall_rate_impact, overall_indicated_change, "
                                           "and policyholders_affected. correction_note documents the split."),
        ("Form A extraction (NEW)", "Three strategies in priority order: (1) Form A section parsing — "
                                    "section 1 COMPANY NAME → section 10 NUMBER OF POLICYHOLDERS → section 12 "
                                    "OVERALL % RATE IMPACT/CHANGE; (2) pdfplumber table extraction with "
                                    "company-anchored rows; (3) free-text regex with company anchors. "
                                    "Section 17 (LAST rate change) is explicitly rejected to avoid "
                                    "historical contamination."),
        ("Validation (NEW)", "Parser validated against AM Best ground truth on SFMA-134676753 (ID): "
                             "extracted SFM impact -9.7% / indicated -2.6%, SFFC impact -2.1% / indicated "
                             "+15.9% — exact match to AM Best published values within rounding. "
                             "Policyholder counts (20,679 SFFC / 360,274 SFM) are not in the SERFF PDFs; "
                             "AM Best sources those from licensed NAIC statutory filings."),
        ("Indicated vs. Impact insight (NEW)",
                            "Tracking BOTH overall_indicated_change (actuarial recommendation) and "
                            "overall_rate_impact (filed proposal) reveals strategic pricing behavior. "
                            "Example from the 4 complete CO State Farm rows: "
                            "SFMA-134532998 SFM filed -3.724% impact vs. 0.0% indicated; SFFC filed -2.872% "
                            "vs. +13.6% indicated. SFMA-134702926 SFM filed -7.822% vs. +0.3% indicated; "
                            "SFFC filed -8.959% vs. +9.2% indicated. In all four cases State Farm filed "
                            "rate DECREASES despite actuarial analysis indicating flat-to-significant-increase "
                            "rates — a deliberate competitive-positioning choice the carrier is making against "
                            "its own actuarial recommendation. Without the indicated column, this gap is invisible."),
        (None, None),
        ("DATA AVAILABILITY (NEW)", None),
        ("Form A filers", "Carriers that consistently include the NAIC Form A schedule (or equivalent "
                          "filing memo with labeled rate impact / policyholder fields) in their public "
                          "SERFF PDFs: STATE FARM (all 8 filings), ALLSTATE (3 of 6 filings — others "
                          "are rule/structural revisions). For these carriers, the 3 new fields are "
                          "extractable from public PDFs."),
        ("Minimum filers", "Carriers whose public SERFF PDFs do NOT include Form A or any extractable "
                           "rate impact / policyholder counts: GEICO (4 of 4 filings — RV model year "
                           "factor, symbol updates), TRAVELERS (2 of 2 — Quantum Home rule revisions), "
                           "PROGRESSIVE (1 of 1 — Symbol Set update), LIBERTY MUTUAL (1 of 1). "
                           "For these filings, the 3 new fields are blank — values would require "
                           "licensed NAIC statutory data access (AM Best, S&P Capital IQ, etc.)."),
        ("Why minimum filers exist", "These filings are typically RULE revisions, SYMBOL/MODEL YEAR "
                                     "updates, or COVERAGE/MANUAL changes — not standalone rate revisions. "
                                     "Form A's Section 12 (OVERALL % RATE IMPACT) doesn't apply when the "
                                     "filing changes structure (e.g., GEICO's WA RV filing: 'Base rates "
                                     "have been offset to be premium neutral.')."),
        (None, None),
        ("LIMITATIONS", None),
        ("Effective date coverage", "SERFF Filing Access does NOT expose Rate Data / per-company effective "
                                    "dates. Blank effective_date is the honest state of the data."),
        ("Premium dollar coverage", "Most filings disclose only percent change, not dollar baseline."),
        ("Authoritative reference", "https://content.naic.org/sites/default/files/inline-files/Property%20%26%20Casualty%20Product%20Coding%20Matrix.pdf"),
    ]
    return base


def _write_data_availability(ws, rows: list[dict]) -> None:
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")
    headers = ["serff_tracking_number", "carrier", "state", "company_name",
               "carrier_type", "rate_effect_value", "overall_rate_impact",
               "overall_indicated_change", "policyholders_affected",
               "data_status"]
    for c_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=c_idx, value=h)
        c.font = bold; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, 2):
        carrier = (row.get("carrier") or "").strip()
        ctype = ("minimum_filer" if carrier in MINIMUM_FILERS
                 else "form_a_filer" if carrier in FORM_A_FILERS
                 else "unknown")
        impact = row.get("overall_rate_impact")
        ind = row.get("overall_indicated_change")
        pol = row.get("policyholders_affected")
        if impact and ind and pol:
            status = "complete"
        elif impact or ind or pol:
            status = "partial"
        elif ctype == "minimum_filer":
            status = "no_form_a_in_public_pdfs"
        else:
            status = "form_a_missing_or_blank"

        vals = [row.get("serff_tracking_number"), carrier, row.get("state"),
                row.get("company_name"), ctype, row.get("rate_effect_value"),
                impact, ind, pol, status]
        for c_idx, v in enumerate(vals, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    widths = [22, 11, 6, 32, 14, 14, 16, 18, 16, 26]
    for c_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w
    ws.freeze_panes = "A2"


def main() -> None:
    headers, existing = load_existing()
    subs = load_subs()
    rows = build_rows(headers, existing, subs)
    write_workbook(headers, rows)

    print(f"[write] {DST}")
    print(f"  rows: {len(rows)} (was {len(existing)})")

    # Coverage report
    impact_filled = sum(1 for r in rows if r.get("overall_rate_impact"))
    ind_filled = sum(1 for r in rows if r.get("overall_indicated_change"))
    pol_filled = sum(1 for r in rows if r.get("policyholders_affected"))
    print(f"  overall_rate_impact filled:        {impact_filled}/{len(rows)} rows")
    print(f"  overall_indicated_change filled:   {ind_filled}/{len(rows)} rows")
    print(f"  policyholders_affected filled:     {pol_filled}/{len(rows)} rows")

    # By carrier breakdown
    by_carrier: dict[str, list[dict]] = {}
    for r in rows:
        by_carrier.setdefault(r.get("carrier") or "?", []).append(r)
    print("\n  Per-carrier coverage:")
    for c, rs in sorted(by_carrier.items()):
        ip = sum(1 for r in rs if r.get("overall_rate_impact"))
        ind = sum(1 for r in rs if r.get("overall_indicated_change"))
        pol = sum(1 for r in rs if r.get("policyholders_affected"))
        ct = ("minimum" if c in MINIMUM_FILERS else "form_a" if c in FORM_A_FILERS else "?")
        print(f"    {c:12s} ({ct:8s})  rows={len(rs):2d}  impact={ip:2d} ind={ind:2d} pol={pol:2d}")


if __name__ == "__main__":
    main()
